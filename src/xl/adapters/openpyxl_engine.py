"""openpyxl-based workbook operations for table mutations, cell/range ops."""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from xl.contracts.common import ChangeRecord, WarningDetail
from xl.engine.context import WorkbookContext


def _parse_ref(ref: str) -> tuple[int, int, int, int]:
    """Parse A1:B2 style ref into (min_row, min_col, max_row, max_col)."""
    parts = ref.replace("$", "").split(":")
    m1 = re.match(r"([A-Z]+)(\d+)", parts[0])
    if not m1:
        raise ValueError(f"Invalid ref: {ref}")
    min_col = column_index_from_string(m1.group(1))
    min_row = int(m1.group(2))
    if len(parts) == 2:
        m2 = re.match(r"([A-Z]+)(\d+)", parts[1])
        if not m2:
            raise ValueError(f"Invalid ref: {ref}")
        max_col = column_index_from_string(m2.group(1))
        max_row = int(m2.group(2))
    else:
        max_col = min_col
        max_row = min_row
    return min_row, min_col, max_row, max_col


def table_add_column(
    ctx: WorkbookContext,
    table_name: str,
    column_name: str,
    *,
    formula: str | None = None,
    default_value: Any | None = None,
    position: str = "append",
) -> ChangeRecord:
    """Add a column to an Excel Table."""
    result = ctx.find_table(table_name)
    if result is None:
        raise ValueError(f"Table not found: {table_name}")
    ws, tbl = result

    existing_names = {tc.name.casefold() for tc in tbl.tableColumns if tc.name}
    # Fallback: also check header row cells in case tableColumns is not yet
    # populated (openpyxl only populates them after a save/reload roundtrip).
    if not existing_names and tbl.ref:
        hdr_min_row, hdr_min_col, _, hdr_max_col = _parse_ref(tbl.ref)
        for c in range(hdr_min_col, hdr_max_col + 1):
            v = ws.cell(row=hdr_min_row, column=c).value
            if v:
                existing_names.add(str(v).casefold())
    if column_name.casefold() in existing_names:
        raise ValueError(f"Column '{column_name}' already exists in table '{table_name}'")

    ref = tbl.ref
    min_row, min_col, max_row, max_col = _parse_ref(ref)
    new_col_idx = max_col + 1
    new_col_letter = get_column_letter(new_col_idx)

    # Write header
    ws.cell(row=min_row, column=new_col_idx, value=column_name)

    # Fill data rows
    for row in range(min_row + 1, max_row + 1):
        if formula:
            ws.cell(row=row, column=new_col_idx, value=formula)
        elif default_value is not None:
            ws.cell(row=row, column=new_col_idx, value=default_value)

    # Update table ref
    new_ref = f"{get_column_letter(min_col)}{min_row}:{new_col_letter}{max_row}"
    tbl.ref = new_ref

    # Add table column
    new_tc = TableColumn(id=len(tbl.tableColumns) + 1, name=column_name)
    tbl.tableColumns.append(new_tc)

    return ChangeRecord(
        type="table.add_column",
        target=f"{table_name}[{column_name}]",
        after={"column": column_name, "formula": formula, "default_value": default_value},
        impact={"rows": max_row - min_row, "cells": max_row - min_row},
    )


def table_append_rows(
    ctx: WorkbookContext,
    table_name: str,
    rows: list[dict[str, Any]],
    *,
    schema_mode: str = "strict",
) -> ChangeRecord:
    """Append rows to an Excel Table."""
    result = ctx.find_table(table_name)
    if result is None:
        raise ValueError(f"Table not found: {table_name}")
    ws, tbl = result

    ref = tbl.ref
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    # Get column names from table
    col_names = [tc.name for tc in tbl.tableColumns]

    # Detect formula columns from first data row
    formula_cols: set[str] = set()
    formula_templates: dict[str, str] = {}
    first_data_row = min_row + 1
    if first_data_row <= max_row:
        for col_idx, col_name in enumerate(col_names, start=min_col):
            cell_val = ws.cell(row=first_data_row, column=col_idx).value
            if isinstance(cell_val, str) and cell_val.startswith("="):
                formula_cols.add(col_name)
                formula_templates[col_name] = cell_val

    warnings: list[WarningDetail] = []

    for row_data in rows:
        if schema_mode == "strict":
            extra = set(row_data.keys()) - set(col_names)
            missing = (set(col_names) - formula_cols) - set(row_data.keys())
            if extra:
                raise ValueError(f"Extra columns not in table schema: {extra}")
            if missing:
                raise ValueError(f"Missing columns in row data: {missing}")
        elif schema_mode == "allow-missing-null":
            extra = set(row_data.keys()) - set(col_names)
            if extra:
                raise ValueError(f"Extra columns not in table schema: {extra}")
        elif schema_mode == "map-by-header":
            pass  # no validation — best-effort case-insensitive mapping below
        else:
            raise ValueError(f"Unknown schema_mode: {schema_mode}")

        max_row += 1
        if schema_mode == "map-by-header":
            # Case-insensitive lookup
            row_lower = {k.casefold(): v for k, v in row_data.items()}
            for col_idx, col_name in enumerate(col_names, start=min_col):
                if col_name in formula_cols and col_name.casefold() not in row_lower:
                    ws.cell(row=max_row, column=col_idx, value=formula_templates[col_name])
                else:
                    val = row_lower.get(col_name.casefold())
                    ws.cell(row=max_row, column=col_idx, value=val)
        else:
            for col_idx, col_name in enumerate(col_names, start=min_col):
                if col_name in formula_cols and col_name not in row_data:
                    ws.cell(row=max_row, column=col_idx, value=formula_templates[col_name])
                else:
                    val = row_data.get(col_name)
                    ws.cell(row=max_row, column=col_idx, value=val)

    # Update table ref
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    tbl.ref = new_ref

    return ChangeRecord(
        type="table.append_rows",
        target=table_name,
        after={"rows_added": len(rows)},
        impact={"rows": len(rows), "cells": len(rows) * len(col_names)},
        warnings=warnings,
    )


_TABLE_NAME_RE = re.compile(r"^[A-Za-z_]\w*$")


def table_create(
    ctx: WorkbookContext,
    sheet_name: str,
    table_name: str,
    ref: str,
    *,
    columns: list[str] | None = None,
    style: str = "TableStyleMedium2",
) -> ChangeRecord:
    """Create an Excel Table (ListObject) from a cell range."""
    # Validate table name
    if not _TABLE_NAME_RE.match(table_name):
        raise ValueError(
            f"Invalid table name: '{table_name}'. "
            "Must start with a letter or underscore and contain only letters, digits, and underscores."
        )

    # Check uniqueness
    if ctx.find_table(table_name) is not None:
        raise ValueError(f"Table '{table_name}' already exists")

    # Get worksheet
    ws = ctx.get_sheet(sheet_name)

    # Parse ref
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    # Check overlap with existing tables
    for existing_tbl in ws._tables.values():
        t_min_row, t_min_col, t_max_row, t_max_col = _parse_ref(existing_tbl.ref)
        if not (max_row < t_min_row or min_row > t_max_row or
                max_col < t_min_col or min_col > t_max_col):
            raise ValueError(
                f"Range {ref} overlaps existing table '{existing_tbl.displayName}' at {existing_tbl.ref}"
            )

    # Determine column names
    header_row_empty = all(
        ws.cell(row=min_row, column=c).value is None
        for c in range(min_col, max_col + 1)
    )

    if columns and header_row_empty:
        # Write column headers
        for i, col_name in enumerate(columns, start=min_col):
            ws.cell(row=min_row, column=i, value=col_name)
        actual_col_names = list(columns)
    elif columns and not header_row_empty:
        # Verify headers match
        existing_headers = [
            ws.cell(row=min_row, column=c).value
            for c in range(min_col, max_col + 1)
        ]
        if existing_headers != columns:
            raise ValueError(
                f"Header row values {existing_headers} do not match provided columns {columns}"
            )
        actual_col_names = list(columns)
    else:
        # Read existing headers
        actual_col_names = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=min_row, column=c).value
            if v is None:
                raise ValueError(
                    f"Header row has empty cell at column {get_column_letter(c)}; "
                    "provide --columns to fill them"
                )
            actual_col_names.append(str(v))

    # Create Table object
    style_info = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = style_info
    for i, col_name in enumerate(actual_col_names):
        tbl.tableColumns.append(TableColumn(id=i + 1, name=col_name))

    ws.add_table(tbl)

    data_rows = max_row - min_row  # rows excluding header
    return ChangeRecord(
        type="table.create",
        target=f"{sheet_name}!{table_name}",
        after={
            "table": table_name,
            "sheet": sheet_name,
            "ref": ref,
            "columns": actual_col_names,
            "style": style,
        },
        impact={
            "rows": data_rows,
            "cells": data_rows * (max_col - min_col + 1),
        },
    )


def cell_set(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
    value: Any,
    *,
    cell_type: str | None = None,
    force_overwrite_formulas: bool = False,
) -> ChangeRecord:
    """Set a single cell value."""
    ws = ctx.get_sheet(sheet_name)
    m = re.match(r"([A-Z]+)(\d+)", ref.replace("$", ""))
    if not m:
        raise ValueError(f"Invalid cell ref: {ref}")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))

    cell = ws.cell(row=row, column=col)
    old_val = cell.value

    # Check formula overwrite
    if isinstance(old_val, str) and old_val.startswith("=") and not force_overwrite_formulas:
        if not (isinstance(value, str) and value.startswith("=")):
            raise ValueError(
                f"Cell {ref} contains formula '{old_val}'. "
                "Use --force-overwrite-formulas to overwrite."
            )

    # Type coercion
    if cell_type == "number" and not isinstance(value, (int, float)):
        value = float(value)
    elif cell_type == "bool":
        value = bool(value)
    elif cell_type == "date" and isinstance(value, str):
        for _fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                value = datetime.strptime(value, _fmt)
                break
            except ValueError:
                continue

    cell.value = value

    if cell_type == "date":
        cell.number_format = "YYYY-MM-DD"

    return ChangeRecord(
        type="cell.set",
        target=f"{sheet_name}!{ref}",
        before=old_val,
        after=value,
        impact={"cells": 1},
    )


def format_number(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
    *,
    style: str = "number",
    decimals: int = 2,
) -> ChangeRecord:
    """Apply number format to a range."""
    ws = ctx.get_sheet(sheet_name)
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    _VALID_STYLES = frozenset({"number", "percent", "currency", "date", "text"})
    if style not in _VALID_STYLES:
        raise ValueError(f"Unknown format style '{style}'. Valid: {', '.join(sorted(_VALID_STYLES))}")

    dec_part = f".{'0' * decimals}" if decimals > 0 else ""
    fmt_map = {
        "number": f"#,##0{dec_part}",
        "percent": f"0{dec_part}%",
        "currency": f"$#,##0{dec_part}",
        "date": "YYYY-MM-DD",
        "text": "@",
    }
    fmt = fmt_map[style]

    cells_touched = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).number_format = fmt
            cells_touched += 1

    return ChangeRecord(
        type="format.number",
        target=f"{sheet_name}!{ref}",
        after={"format": fmt, "style": style, "decimals": decimals},
        impact={"cells": cells_touched},
    )


def resolve_table_column_ref(
    ctx: WorkbookContext,
    ref: str,
    *,
    include_header: bool = True,
) -> tuple[str, str] | None:
    """Resolve 'TableName[ColumnName]' to (sheet_name, A1_range)."""
    m = re.match(r"(\w+)\[(\w+)\]", ref)
    if not m:
        return None
    table_name, col_name = m.group(1), m.group(2)
    result = ctx.find_table(table_name)
    if result is None:
        return None
    ws, tbl = result
    sheet_name = ws.title

    tbl_ref = tbl.ref
    min_row, min_col, max_row, max_col = _parse_ref(tbl_ref)
    for i, tc in enumerate(tbl.tableColumns):
        if tc.name == col_name:
            col_idx = min_col + i
            col_letter = get_column_letter(col_idx)
            start_row = min_row if include_header else min_row + 1
            if start_row > max_row:
                # No data rows exist — nothing to target.
                return None
            return sheet_name, f"{col_letter}{start_row}:{col_letter}{max_row}"
    return None


# ---------------------------------------------------------------------------
# formula ref adjustment (for relative fill)
# ---------------------------------------------------------------------------
_CELL_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_(])"
)


def _adjust_formula_refs(formula: str, row_delta: int, col_delta: int) -> str:
    """Adjust A1-style references in a formula by row_delta and col_delta.

    - $-prefixed axes are absolute and not adjusted.
    - Content inside double-quoted string literals is skipped.
    - Cross-sheet refs like Sheet1!A1 are handled (the Sheet1! prefix is
      not a letter preceding the column, so the lookbehind passes).
    """
    # Split on double quotes to avoid adjusting refs inside string literals.
    # Odd-indexed segments are inside quotes.
    parts = formula.split('"')
    for i in range(0, len(parts), 2):  # process only outside-string segments
        parts[i] = _CELL_REF_RE.sub(
            lambda m: _adjust_match(m, row_delta, col_delta),
            parts[i],
        )
    return '"'.join(parts)


def _adjust_match(m: re.Match, row_delta: int, col_delta: int) -> str:
    col_abs, col_letters, row_abs, row_num = m.group(1), m.group(2), m.group(3), m.group(4)

    # Adjust column if not absolute
    if col_abs != "$" and col_delta != 0:
        col_idx = column_index_from_string(col_letters) + col_delta
        if col_idx < 1:
            col_idx = 1
        col_letters = get_column_letter(col_idx)

    # Adjust row if not absolute
    if row_abs != "$" and row_delta != 0:
        new_row = int(row_num) + row_delta
        if new_row < 1:
            new_row = 1
        row_num = str(new_row)

    return f"{col_abs}{col_letters}{row_abs}{row_num}"


# ---------------------------------------------------------------------------
# formula set
# ---------------------------------------------------------------------------
def formula_set(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
    formula: str,
    *,
    force_overwrite_values: bool = False,
    force_overwrite_formulas: bool = False,
    fill_mode: str = "relative",
) -> ChangeRecord:
    """Set a formula on a cell or range.

    fill_mode: "fixed" copies the formula literally; "relative" adjusts
    A1-style references relative to the top-left cell of the range.
    """
    ws = ctx.get_sheet(sheet_name)
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    cells_touched = 0
    blocked: list[str] = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            old_val = cell.value
            cell_ref = f"{get_column_letter(col)}{row}"
            # Guard: existing formula
            if isinstance(old_val, str) and old_val.startswith("=") and not force_overwrite_formulas:
                blocked.append(f"{cell_ref} has formula '{old_val}'")
                continue
            # Guard: existing non-empty value
            if old_val is not None and not (isinstance(old_val, str) and old_val.startswith("=")) and not force_overwrite_values:
                blocked.append(f"{cell_ref} has value '{old_val}'")
                continue
            if fill_mode == "relative":
                adjusted = _adjust_formula_refs(formula, row - min_row, col - min_col)
            else:
                adjusted = formula
            cell.value = adjusted
            cells_touched += 1

    if blocked and cells_touched == 0:
        raise ValueError(
            f"All cells blocked: {'; '.join(blocked[:5])}. "
            "Use --force-overwrite-values or --force-overwrite-formulas."
        )

    return ChangeRecord(
        type="formula.set",
        target=f"{sheet_name}!{ref}",
        after={"formula": formula, "cells_touched": cells_touched, "blocked": len(blocked)},
        impact={"cells": cells_touched},
        warnings=[WarningDetail(code="WARN_CELLS_BLOCKED", message=msg) for msg in blocked[:5]],
    )


# ---------------------------------------------------------------------------
# formula lint
# ---------------------------------------------------------------------------
_VOLATILE_FUNCS = re.compile(r"\b(OFFSET|INDIRECT|NOW|TODAY|RAND|RANDBETWEEN)\b", re.IGNORECASE)
_BROKEN_REF = re.compile(r"#REF!")


def formula_lint(
    ctx: WorkbookContext,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Heuristic lint checks on formulas. Returns list of findings."""
    findings: list[dict[str, Any]] = []
    sheets = [sheet_name] if sheet_name else ctx.wb.sheetnames

    for sname in sheets:
        ws = ctx.wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                cell_ref = f"{sname}!{cell.coordinate}"
                # Volatile functions
                m = _VOLATILE_FUNCS.search(val)
                if m:
                    findings.append({
                        "ref": cell_ref,
                        "category": "volatile_function",
                        "severity": "warning",
                        "message": f"Uses volatile function {m.group(0).upper()}",
                        "formula": val,
                    })
                # Broken refs
                if _BROKEN_REF.search(val):
                    findings.append({
                        "ref": cell_ref,
                        "category": "broken_ref",
                        "severity": "error",
                        "message": "Contains #REF! error reference",
                        "formula": val,
                    })

    return findings


# ---------------------------------------------------------------------------
# formula find
# ---------------------------------------------------------------------------
def formula_find(
    ctx: WorkbookContext,
    pattern: str,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Search workbook for formulas matching a regex pattern."""
    regex = re.compile(pattern, re.IGNORECASE)
    matches: list[dict[str, Any]] = []
    sheets = [sheet_name] if sheet_name else ctx.wb.sheetnames

    for sname in sheets:
        ws = ctx.wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                m = regex.search(val)
                if m:
                    matches.append({
                        "ref": f"{sname}!{cell.coordinate}",
                        "formula": val,
                        "match": m.group(0),
                    })

    return matches


# ---------------------------------------------------------------------------
# cell get
# ---------------------------------------------------------------------------
def cell_get(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
) -> dict[str, Any]:
    """Read a cell value and metadata."""
    ws = ctx.get_sheet(sheet_name)
    m = re.match(r"([A-Z]+)(\d+)", ref.replace("$", ""))
    if not m:
        raise ValueError(f"Invalid cell ref: {ref}")
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))

    cell = ws.cell(row=row, column=col)
    val = cell.value
    val_type = "empty"
    formula_text = None
    if val is None:
        val_type = "empty"
    elif isinstance(val, str) and val.startswith("="):
        val_type = "formula"
        formula_text = val
    elif isinstance(val, bool):
        val_type = "bool"
    elif isinstance(val, datetime):
        val_type = "date"
        val = val.isoformat()
    elif isinstance(val, (int, float)):
        val_type = "number"
    elif isinstance(val, str):
        val_type = "text"
    else:
        val_type = type(val).__name__

    return {
        "ref": f"{sheet_name}!{ref}",
        "value": val,
        "type": val_type,
        "formula": formula_text,
        "number_format": cell.number_format,
    }


# ---------------------------------------------------------------------------
# range stat
# ---------------------------------------------------------------------------
def range_stat(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
) -> dict[str, Any]:
    """Compute statistics for a range."""
    ws = ctx.get_sheet(sheet_name)
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    row_count = max_row - min_row + 1
    col_count = max_col - min_col + 1
    non_empty = 0
    numeric_count = 0
    formula_count = 0
    numeric_vals: list[float] = []

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                non_empty += 1
            if isinstance(val, str) and val.startswith("="):
                formula_count += 1
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                numeric_count += 1
                numeric_vals.append(float(val))

    stats: dict[str, Any] = {
        "ref": f"{sheet_name}!{ref}",
        "row_count": row_count,
        "col_count": col_count,
        "non_empty_count": non_empty,
        "numeric_count": numeric_count,
        "formula_count": formula_count,
    }
    if numeric_vals:
        stats["min"] = min(numeric_vals)
        stats["max"] = max(numeric_vals)
        stats["sum"] = sum(numeric_vals)
        stats["avg"] = sum(numeric_vals) / len(numeric_vals)

    return stats


# ---------------------------------------------------------------------------
# range clear
# ---------------------------------------------------------------------------
def range_clear(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str,
    *,
    contents: bool = True,
    formats: bool = False,
) -> ChangeRecord:
    """Clear a range of cells."""
    ws = ctx.get_sheet(sheet_name)
    min_row, min_col, max_row, max_col = _parse_ref(ref)

    cells_cleared = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if contents:
                cell.value = None
            if formats:
                cell.number_format = "General"
            cells_cleared += 1

    return ChangeRecord(
        type="range.clear",
        target=f"{sheet_name}!{ref}",
        after={"cells_cleared": cells_cleared, "contents": contents, "formats": formats},
        impact={"cells": cells_cleared},
    )


# ---------------------------------------------------------------------------
# format width
# ---------------------------------------------------------------------------
def format_width(
    ctx: WorkbookContext,
    sheet_name: str,
    columns: list[str],
    width: float,
) -> ChangeRecord:
    """Set column widths."""
    ws = ctx.get_sheet(sheet_name)
    for col_letter in columns:
        ws.column_dimensions[col_letter].width = width

    return ChangeRecord(
        type="format.width",
        target=f"{sheet_name}![{','.join(columns)}]",
        after={"columns": columns, "width": width},
        impact={"cells": len(columns)},
    )


# ---------------------------------------------------------------------------
# format freeze
# ---------------------------------------------------------------------------
def format_freeze(
    ctx: WorkbookContext,
    sheet_name: str,
    ref: str | None,
) -> ChangeRecord:
    """Freeze panes at a given cell ref, or unfreeze if ref is None."""
    ws = ctx.get_sheet(sheet_name)
    old_freeze = ws.freeze_panes
    ws.freeze_panes = ref

    return ChangeRecord(
        type="format.freeze",
        target=f"{sheet_name}",
        before=old_freeze,
        after=ref,
        impact={"cells": 0},
    )


# ---------------------------------------------------------------------------
# Sheet delete / rename
# ---------------------------------------------------------------------------

def sheet_delete(
    ctx: WorkbookContext,
    sheet_name: str,
) -> ChangeRecord:
    """Delete a sheet from the workbook."""
    if sheet_name not in ctx.wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    if len(ctx.wb.sheetnames) <= 1:
        raise ValueError(
            f"Cannot delete sheet '{sheet_name}': workbook must retain at least one sheet"
        )
    ws = ctx.wb[sheet_name]

    # Check for tables on the sheet
    table_names = [t.displayName for t in ws._tables.values()]
    warnings = []
    if table_names:
        warnings.append(WarningDetail(
            code="WARN_TABLES_ON_SHEET",
            message=f"Sheet '{sheet_name}' contains {len(table_names)} table(s) that will be destroyed: {', '.join(table_names)}",
        ))

    used_range = ws.dimensions if ws.dimensions else None
    ctx.wb.remove(ws)
    return ChangeRecord(
        type="sheet.delete",
        target=sheet_name,
        before={"sheet": sheet_name, "used_range": used_range, "tables": table_names},
        after=None,
        impact={"sheets": 1, "tables": len(table_names)},
        warnings=warnings,
    )


def sheet_rename(
    ctx: WorkbookContext,
    old_name: str,
    new_name: str,
) -> ChangeRecord:
    """Rename a sheet in the workbook."""
    if old_name not in ctx.wb.sheetnames:
        raise KeyError(f"Sheet not found: {old_name}")
    if new_name in ctx.wb.sheetnames:
        raise ValueError(f"Sheet '{new_name}' already exists")
    ws = ctx.wb[old_name]
    ws.title = new_name
    return ChangeRecord(
        type="sheet.rename",
        target=old_name,
        before={"sheet": old_name},
        after={"sheet": new_name},
        impact={"sheets": 1},
    )


# ---------------------------------------------------------------------------
# Table delete / column delete
# ---------------------------------------------------------------------------

def table_delete(
    ctx: WorkbookContext,
    table_name: str,
) -> ChangeRecord:
    """Delete an Excel Table definition (preserves cell data)."""
    result = ctx.find_table(table_name)
    if result is None:
        raise ValueError(f"Table not found: {table_name}")
    ws, tbl = result

    ref = tbl.ref
    col_names = [tc.name for tc in tbl.tableColumns]

    # Remove table from worksheet table list (TableList is dict-like, keyed by name)
    del ws._tables[table_name]

    return ChangeRecord(
        type="table.delete",
        target=table_name,
        before={"table": table_name, "ref": ref, "columns": col_names},
        after=None,
        impact={"tables": 1},
    )


def table_delete_column(
    ctx: WorkbookContext,
    table_name: str,
    column_name: str,
) -> ChangeRecord:
    """Delete a column from an Excel Table, shifting remaining columns left."""
    result = ctx.find_table(table_name)
    if result is None:
        raise ValueError(f"Table not found: {table_name}")
    ws, tbl = result

    # Find column index
    col_idx = None
    for i, tc in enumerate(tbl.tableColumns):
        if tc.name == column_name:
            col_idx = i
            break
    if col_idx is None:
        raise ValueError(f"Column '{column_name}' not found in table '{table_name}'")

    ref = tbl.ref
    min_row, min_col, max_row, max_col = _parse_ref(ref)
    target_col = min_col + col_idx

    # Shift cells left from the column after the deleted one
    for row in range(min_row, max_row + 1):
        for col in range(target_col, max_col):
            src = ws.cell(row=row, column=col + 1)
            dst = ws.cell(row=row, column=col)
            dst.value = src.value
            dst.number_format = src.number_format
        # Clear the last column cell
        ws.cell(row=row, column=max_col).value = None
        ws.cell(row=row, column=max_col).number_format = "General"

    # Remove from tableColumns
    del tbl.tableColumns[col_idx]

    # Update table ref (one column narrower)
    new_max_col_letter = get_column_letter(max_col - 1)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{new_max_col_letter}{max_row}"
    tbl.ref = new_ref

    # Re-index table column IDs
    for i, tc in enumerate(tbl.tableColumns):
        tc.id = i + 1

    rows_affected = max_row - min_row  # data rows (excluding header)
    return ChangeRecord(
        type="table.delete_column",
        target=f"{table_name}[{column_name}]",
        before={"column": column_name, "table": table_name},
        after=None,
        impact={"rows": rows_affected, "cells": rows_affected + 1},
    )
