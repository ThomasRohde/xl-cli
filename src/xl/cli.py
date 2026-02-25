"""Typer CLI application — top-level commands and subcommand groups."""

from __future__ import annotations

import json
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated, Any, Optional

import typer

from xl.help import patch_typer_help
from xl.help.custom_types import patch_typer_errors

patch_typer_help()
patch_typer_errors()

import xl
from xl.contracts.common import ChangeRecord, ErrorDetail, Target, WarningDetail, WorkbookCorruptError
from xl.contracts.plans import (
    Operation,
    PatchPlan,
    PlanTarget,
    Postcondition,
    Precondition,
)
from xl.contracts.responses import ApplyResult, ValidationResult
from xl.engine.dispatcher import (
    error_envelope,
    exit_code_for,
    print_response,
    success_envelope,
)
from xl.io.fileops import read_text_safe
from xl.observe.events import Timer

# ---------------------------------------------------------------------------
# App & subcommand groups
# ---------------------------------------------------------------------------

_MAIN_HELP = """\
Agent-first CLI for reading, transforming, and validating Excel workbooks (.xlsx/.xlsm).

**Recommended workflow:**  inspect → plan → validate → apply → verify

1. `xl wb inspect -f data.xlsx`  — discover sheets, tables, named ranges, fingerprint
2. `xl plan add-column -f data.xlsx -t Sales -n Margin --formula "=[@Revenue]-[@Cost]" --out plan.json`
3. `xl validate plan -f data.xlsx --plan plan.json`
4. `xl apply -f data.xlsx --plan plan.json --dry-run`  — preview changes
5. `xl apply -f data.xlsx --plan plan.json --backup`   — apply with backup
6. `xl verify assert -f data.xlsx --assertions '[{"type":"table.column_exists","table":"Sales","column":"Margin"}]'`

**Every command** returns a JSON `ResponseEnvelope`:
`{"ok": bool, "command": "...", "result": {...}, "errors": [...], "warnings": [...], "metrics": {"duration_ms": N}}`

**Ref syntax** used by cell/range/formula/format commands:
- Cell: `Sheet1!B2` — Range: `Sheet1!A1:D10` — Table column: `TableName[ColumnName]`

**Safety rails** — all mutating commands support:
- `--dry-run` previews changes without writing
- `--backup` creates a timestamped .bak copy before writing
- Fingerprint-based conflict detection prevents stale overwrites

**Exit codes:** 0=success, 10=validation, 20=protection, 30=formula, 40=conflict, 50=io, 60=recalc, 70=unsupported, 90=internal

Run `xl guide` for a comprehensive machine-readable orientation.
"""

_WB_EPILOG = """\
**Examples:**

`xl wb inspect -f data.xlsx`  — returns sheets, tables, named ranges, fingerprint

`xl wb lock-status -f data.xlsx`  — check if another process holds a lock

The **fingerprint** (xxhash of file contents) is used by patch plans for conflict detection.
Use `xl wb inspect` as the first step to understand any workbook.
"""

_SHEET_EPILOG = """\
**Examples:**

`xl sheet ls -f data.xlsx`  — list all sheets with dimensions and visibility

Sheet names are required in ref syntax for cell/range commands (e.g. `Sheet1!A1`).
Use this command to discover valid sheet names before other operations.
"""

_TABLE_EPILOG = """\
**Examples:**

`xl table ls -f data.xlsx`  — list all tables with columns, row counts, sheet locations

`xl table create -f data.xlsx -t Sales -s Revenue --ref A1:D5`  — promote range to table

`xl table create -f data.xlsx -t Metrics -s Sheet1 --ref A1:C1 --columns "Name,Value,Date"`

`xl table add-column -f data.xlsx -t Sales -n Profit --formula "=[@Revenue]-[@Cost]"`

`xl table append-rows -f data.xlsx -t Sales --data '[{"Revenue":100,"Cost":60}]'`

**Table column refs** use structured references: `TableName[ColumnName]` (usable in formula and format commands).
Prefer table-level operations over raw cell manipulation when data is in Excel tables.
"""

_CELL_EPILOG = """\
**Examples:**

`xl cell get -f data.xlsx --ref "Sheet1!B2"`

`xl cell set -f data.xlsx --ref "Sheet1!B2" --value 42 --type number`

`xl cell set -f data.xlsx --ref "Sheet1!B2" --value "Hello" --type text`

**Ref format:** always include the sheet name — `SheetName!CellRef` (e.g. `Sheet1!B2`).
**Formula protection:** setting a cell that contains a formula requires `--force-overwrite-formulas`.
"""

_RANGE_EPILOG = """\
**Examples:**

`xl range stat -f data.xlsx --ref "Sheet1!C2:C100"`  — min, max, mean, sum, count, stddev

`xl range clear -f data.xlsx --ref "Sheet1!A1:D10" --contents`  — clear values only

`xl range clear -f data.xlsx --ref "Sheet1!A1:D10" --all`  — clear values and formatting

**Ref format:** `SheetName!StartCell:EndCell` (e.g. `Sheet1!A1:D10`).
"""

_FORMULA_EPILOG = """\
**Examples:**

`xl formula set -f data.xlsx --ref "Sheet1!E2" --formula "=C2-D2"`

`xl formula set -f data.xlsx --ref "Sales[Margin]" --formula "=[@Revenue]-[@Cost]"`  — table column

`xl formula lint -f data.xlsx`  — find volatile functions, broken refs, common issues

`xl formula find -f data.xlsx --pattern "VLOOKUP"`  — search by regex

**Safety:** overwriting existing formulas requires `--force-overwrite-formulas`.
Overwriting existing values requires `--force-overwrite-values`.
"""

_FORMAT_EPILOG = """\
**Examples:**

`xl format number -f data.xlsx --ref "Sheet1!C2:C100" --style currency --decimals 2`

`xl format number -f data.xlsx --ref "Sales[Revenue]" --style number --decimals 0`  — table column ref

`xl format width -f data.xlsx --sheet Sheet1 --columns A,B,C --width 15`

`xl format freeze -f data.xlsx --sheet Sheet1 --ref B2`  — freeze rows above and columns left of B2

`xl format freeze -f data.xlsx --sheet Sheet1 --unfreeze`  — remove freeze

**Styles:** number, percent, currency, date, text.
**Refs:** `Sheet!Range` or `TableName[Column]` for number formatting.
"""

_VALIDATE_EPILOG = """\
**Examples:**

`xl validate workbook -f data.xlsx`  — check workbook health (corrupt refs, missing sheets)

`xl validate plan -f data.xlsx --plan plan.json`  — check plan against workbook (fingerprint, tables, columns)

`xl validate refs -f data.xlsx --ref "Sheet1!A1:D10"`  — verify a ref points to valid cells

Run validation **before** `xl apply` to catch issues early.
"""

_PLAN_EPILOG = """\
**Patch plans** are JSON files describing changes to apply atomically.
Generate → compose → validate → apply (with --dry-run first).

**Examples:**

`xl plan add-column -f data.xlsx -t Sales -n Margin --formula "=[@Revenue]-[@Cost]" --out plan.json` — generate raw plan file

`xl plan set-cells -f data.xlsx --ref "Sheet1!B2" --value 42 --type number`

`xl plan format -f data.xlsx --ref "Sales[Revenue]" --style currency --decimals 2`

`xl plan compose --plan plan1.json --plan plan2.json`  — merge multiple plans

`xl plan show --plan plan.json`  — inspect a plan

Plans include a **fingerprint** of the target workbook for conflict detection.
Use `--out` to write raw plan files and `--append plan.json` to incrementally build multi-operation plans.
"""

_VERIFY_EPILOG = """\
**Examples:**

`xl verify assert -f data.xlsx --assertions '[{"type":"table.column_exists","table":"Sales","column":"Margin"}]'`

`xl verify assert -f data.xlsx --assertions '[{"type":"cell.value_equals","ref":"Sheet1!B2","expected":42}]'`

`xl verify assert -f data.xlsx --assertions-file assertions.json`

**Assertion types and required fields:**

- `table.exists` — fields: `table`
- `table.column_exists` — fields: `table`, `column`
- `cell.value_equals` — fields: `ref` (Sheet!Cell), `expected` (value to match; legacy alias: `value`)
- `cell.not_empty` — fields: `ref` (Sheet!Cell)
- `cell.value_type` — fields: `ref` (Sheet!Cell), `expected_type` (one of: number, text, formula, bool, empty)
- `table.row_count` — fields: `table`, plus one of: `expected` (exact), `min`, `max`
- `table.row_count.gte` / `row_count.gte` — fields: `table`, `min_rows` (or `min`)

Run after `xl apply` to confirm the workbook is in the expected state.
"""

_DIFF_EPILOG = """\
**Examples:**

`xl diff compare --file-a original.xlsx --file-b modified.xlsx`

`xl diff compare --file-a original.xlsx --file-b modified.xlsx --sheet Revenue`  — single sheet

Returns cell-level changes, sheets added/removed, and fingerprint comparison.
Useful for reviewing changes after `xl apply`.
"""

def _version_callback(value: bool) -> None:
    if value:
        typer.echo(xl.__version__)
        raise typer.Exit()


app = typer.Typer(
    name="xl",
    help=_MAIN_HELP,
    no_args_is_help=True,
    rich_markup_mode="markdown",
)

wb_app = typer.Typer(
    name="wb", help="Workbook-level inspection and status.",
    epilog=_WB_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
sheet_app = typer.Typer(
    name="sheet", help="Sheet listing and discovery.",
    epilog=_SHEET_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
table_app = typer.Typer(
    name="table", help="Table operations — create, list, add columns, append rows.",
    epilog=_TABLE_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
cell_app = typer.Typer(
    name="cell", help="Read and write individual cell values.",
    epilog=_CELL_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
range_app = typer.Typer(
    name="range", help="Range statistics and clearing.",
    epilog=_RANGE_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
formula_app = typer.Typer(
    name="formula", help="Set, lint, and search formulas.",
    epilog=_FORMULA_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
format_app = typer.Typer(
    name="format", help="Number formatting, column widths, and freeze panes.",
    epilog=_FORMAT_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
query_app = typer.Typer(
    name="query", help="SQL queries over table data (via DuckDB).",
    rich_markup_mode="markdown",
)
validate_app = typer.Typer(
    name="validate", help="Validate workbook health, plans, and references.",
    epilog=_VALIDATE_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
plan_app = typer.Typer(
    name="plan", help="Generate, inspect, and compose patch plans.",
    epilog=_PLAN_EPILOG,
    no_args_is_help=True, rich_markup_mode="markdown",
)
apply_app = typer.Typer(
    name="apply", help="Apply patch plans to workbooks.",
    rich_markup_mode="markdown",
)
verify_app = typer.Typer(
    name="verify", help="Run post-apply assertions to confirm expected state.",
    epilog=_VERIFY_EPILOG,
    rich_markup_mode="markdown",
)
diff_app = typer.Typer(
    name="diff", help="Compare two workbook files cell by cell.",
    epilog=_DIFF_EPILOG,
    rich_markup_mode="markdown",
)

app.add_typer(wb_app)
app.add_typer(sheet_app)
app.add_typer(table_app)
app.add_typer(cell_app)
app.add_typer(range_app)
app.add_typer(formula_app)
app.add_typer(format_app)
app.add_typer(validate_app)
app.add_typer(plan_app)
app.add_typer(verify_app)
app.add_typer(diff_app)


@app.callback(invoke_without_command=True)
def main(
    version: Annotated[
        bool, typer.Option("--version", "-V", help="Print version and exit.", is_eager=True)
    ] = False,
    human: Annotated[
        bool, typer.Option("--human", help="Force human-readable help (overrides LLM=true).", is_eager=True)
    ] = False,
) -> None:
    if version:
        _version_callback(True)


# Type aliases for common options
FilePath = Annotated[str, typer.Option("--file", "-f", help="Path to .xlsx/.xlsm workbook file")]
JsonFlag = Annotated[bool, typer.Option("--json", help="JSON output (always on — all output is JSON)")]
SheetOpt = Annotated[Optional[str], typer.Option("--sheet", "-s", help="Sheet name (as shown by 'xl sheet ls')")]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _load_ctx(file: str, *, data_only: bool = False):
    from xl.engine.context import WorkbookContext
    return WorkbookContext(file, data_only=data_only)


def _emit(envelope, code=None):
    print_response(envelope)
    raise typer.Exit(code if code is not None else exit_code_for(envelope))


def _load_ctx_or_emit(file: str, cmd: str, *, data_only: bool = False):
    """Load a WorkbookContext, or emit an error envelope and return None."""
    try:
        return _load_ctx(file, data_only=data_only)
    except FileNotFoundError:
        env = error_envelope(cmd, "ERR_WORKBOOK_NOT_FOUND", f"File not found: {file}", target=Target(file=file))
        _emit(env)
    except WorkbookCorruptError as e:
        env = error_envelope(cmd, "ERR_WORKBOOK_CORRUPT", str(e), target=Target(file=file))
        _emit(env)


def _load_patch_plan(
    plan_path: str,
) -> PatchPlan:
    """Load and validate a raw PatchPlan JSON file."""
    try:
        data = json.loads(read_text_safe(plan_path))
    except Exception as e:
        raise ValueError(f"Cannot parse plan: {e}") from e

    if not isinstance(data, dict):
        raise ValueError("Plan file must contain a JSON object.")

    if {"ok", "command", "result"}.issubset(data):
        # Auto-extract the plan body from a ResponseEnvelope for convenience.
        inner = data.get("result")
        if isinstance(inner, dict) and "target" in inner and "operations" in inner:
            data = inner
        else:
            raise ValueError(
                "Plan file contains a ResponseEnvelope, but its 'result' is not a "
                "valid patch plan body. Use --out to write raw plan files."
            )

    missing = [k for k in ("target", "operations") if k not in data]
    if missing:
        raise ValueError(f"Plan file missing required keys: {', '.join(missing)}")

    try:
        return PatchPlan(**data)
    except Exception as e:
        raise ValueError(f"Cannot parse plan: {e}") from e


def _emit_invalid_plan(command: str, file: str, message: str) -> None:
    env = error_envelope(command, "ERR_PLAN_INVALID", message, target=Target(file=file))
    _emit(env)


def _write_plan(path: str, plan: PatchPlan) -> None:
    Path(path).write_text(json.dumps(plan.model_dump(mode="json"), indent=2))


# ---------------------------------------------------------------------------
# xl version
# ---------------------------------------------------------------------------
@app.command()
def version():
    """Print the xl CLI version.

    Example: `xl version`
    """
    env = success_envelope("version", {"version": xl.__version__})
    _emit(env)


# ---------------------------------------------------------------------------
# xl guide
# ---------------------------------------------------------------------------
@app.command()
def guide(
    json_out: JsonFlag = True,
):
    """Print the complete agent integration guide as structured JSON.

    Returns a comprehensive guide covering all commands, workflows,
    ref syntax, response format, error codes, and examples.
    Designed for AI agent onboarding — run this first when working
    with an unfamiliar workbook or learning the CLI.

    Example: `xl guide`
    """
    guide_data = {
        "overview": (
            "xl is an agent-first CLI for reading, transforming, and validating "
            "Excel workbooks (.xlsx/.xlsm). It provides a transactional "
            "spreadsheet execution layer with JSON outputs, patch plans, "
            "dry-run/validation, and safety rails."
        ),
        "workflow": {
            "description": "The recommended workflow for safe workbook mutations.",
            "steps": [
                {"step": 1, "command": "xl wb inspect -f <file>", "purpose": "Discover sheets, tables, named ranges, fingerprint"},
                {"step": 2, "command": "xl table ls -f <file>", "purpose": "List tables with columns and row counts"},
                {"step": 3, "command": "xl plan add-column -f <file> -t <table> -n <col> --formula <formula> --out plan.json", "purpose": "Generate a raw patch plan file"},
                {"step": 4, "command": "xl validate plan -f <file> --plan plan.json", "purpose": "Validate plan against workbook"},
                {"step": 5, "command": "xl apply -f <file> --plan plan.json --dry-run", "purpose": "Preview changes without writing"},
                {"step": 6, "command": "xl apply -f <file> --plan plan.json --backup", "purpose": "Apply changes with backup"},
                {"step": 7, "command": "xl verify assert -f <file> --assertions <json>", "purpose": "Confirm expected state"},
            ],
        },
        "commands": {
            "inspection": {
                "xl wb inspect": "Discover workbook structure — sheets, tables, named ranges, fingerprint",
                "xl wb lock-status": "Check if file is locked by another process",
                "xl sheet ls": "List all sheets with dimensions",
                "xl table ls": "List all tables with columns and row counts",
            },
            "reading": {
                "xl cell get": "Read a single cell value, type, and formula",
                "xl range stat": "Compute statistics (min, max, mean, sum, count, stddev) for a range",
                "xl query": "Query table data using SQL via DuckDB",
                "xl formula find": "Search formulas by regex pattern",
                "xl formula lint": "Lint formulas for common issues",
            },
            "plan_generation": {
                "xl plan add-column": "Generate a plan to add a table column (non-mutating)",
                "xl plan create-table": "Generate a plan to create an Excel Table from a range (non-mutating)",
                "xl plan set-cells": "Generate a plan to set cell values (non-mutating)",
                "xl plan format": "Generate a plan for number formatting (non-mutating)",
                "xl plan compose": "Merge multiple plan files into one",
                "xl plan show": "Display a plan's contents",
            },
            "validation": {
                "xl validate workbook": "Check workbook health",
                "xl validate plan": "Validate a plan against a workbook (fingerprint, tables, columns)",
                "xl validate refs": "Verify a cell/range reference is valid",
            },
            "mutation": {
                "xl apply": "Apply a patch plan to a workbook (supports --dry-run and --backup)",
                "xl cell set": "Set a single cell value",
                "xl table create": "Create an Excel Table from a cell range (promote range to ListObject)",
                "xl table add-column": "Add a column to an Excel table",
                "xl table append-rows": "Append rows to an Excel table",
                "xl formula set": "Set a formula on a cell, range, or table column",
                "xl format number": "Apply number format to a range or table column",
                "xl format width": "Set column widths",
                "xl format freeze": "Freeze or unfreeze panes",
                "xl range clear": "Clear cell contents and/or formatting",
            },
            "verification": {
                "xl verify assert": "Run post-apply assertions (table.exists, table.column_exists, cell.value_equals, cell.not_empty, cell.value_type, table.row_count, table.row_count.gte / row_count.gte)",
                "xl diff compare": "Compare two workbook files cell by cell",
            },
            "automation": {
                "xl run": "Execute a multi-step YAML workflow",
                "xl serve --stdio": "Start stdio server for agent tool integration (MCP/ACP)",
            },
        },
        "workflow_commands": {
            "description": "Step commands supported in 'xl run' YAML workflows (steps[].run values).",
            "inspection": ["wb.inspect", "sheet.ls", "table.ls", "cell.get", "range.stat", "query", "formula.find", "formula.lint"],
            "mutation": ["table.create", "table.add_column", "table.append_rows", "cell.set", "formula.set", "format.number", "format.width", "format.freeze", "range.clear"],
            "validation": ["validate.plan", "validate.workbook", "validate.refs", "verify.assert"],
            "other": ["apply", "diff.compare"],
        },
        "ref_syntax": {
            "description": "Reference formats used by cell, range, formula, and format commands.",
            "formats": [
                {"pattern": "SheetName!CellRef", "example": "Sheet1!B2", "usage": "Single cell (cell get/set)"},
                {"pattern": "SheetName!Start:End", "example": "Sheet1!A1:D10", "usage": "Cell range (range stat/clear, formula set)"},
                {"pattern": "TableName[ColumnName]", "example": "Sales[Revenue]", "usage": "Table column (formula set, format number)"},
                {"pattern": "=[@ColumnName]", "example": "=[@Revenue]-[@Cost]", "usage": "Structured ref inside table formulas"},
            ],
        },
        "response_format": {
            "description": "Every command returns a JSON ResponseEnvelope.",
            "fields": {
                "ok": "bool — true if command succeeded",
                "command": "string — command identifier (e.g. 'table.ls')",
                "target": "object — {file, sheet, table, ref} identifying the target",
                "result": "any — command-specific payload",
                "changes": "array — list of ChangeRecord objects describing mutations",
                "warnings": "array — non-fatal issues [{code, message}]",
                "errors": "array — error details [{code, message, details}]",
                "metrics": "object — {duration_ms: int}",
                "recalc": "object — {mode: 'cached', performed: false}",
            },
            "parsing_rules": [
                "Check 'ok' first — false means the command failed",
                "Read errors[0].code for machine-readable error category",
                "Check warnings[] for non-fatal issues",
                "Use result for the command-specific data",
                "Use the process exit code for scripting (0=success, non-zero=error)",
            ],
        },
        "error_codes": {
            "ERR_WORKBOOK_NOT_FOUND": "File does not exist (exit 50)",
            "ERR_WORKBOOK_CORRUPT": "File cannot be parsed as a workbook (exit 50)",
            "ERR_SHEET_NOT_FOUND": "Named sheet not found in workbook (exit 10)",
            "ERR_TABLE_NOT_FOUND": "Named table not found in workbook (exit 50)",
            "ERR_RANGE_INVALID": "Reference format is invalid or sheet not found (exit 10)",
            "ERR_PATTERN_INVALID": "Regex pattern is invalid (exit 10)",
            "ERR_COLUMN_EXISTS": "Column already exists in target table (exit 10)",
            "ERR_TABLE_EXISTS": "Table name already exists in workbook (exit 10)",
            "ERR_TABLE_OVERLAP": "Table range overlaps an existing table (exit 10)",
            "ERR_INVALID_ARGUMENT": "Conflicting or malformed arguments provided (exit 10)",
            "ERR_SCHEMA_MISMATCH": "Row data columns don't match table schema (exit 10)",
            "ERR_FORMULA_OVERWRITE_BLOCKED": "Cell contains a formula; use --force-overwrite-formulas (exit 30)",
            "ERR_FORMULA_BLOCKED": "Formula write blocked by safety check (exit 30)",
            "ERR_PLAN_INVALID": "Plan JSON cannot be parsed (exit 10)",
            "ERR_PLAN_FINGERPRINT_CONFLICT": "Workbook changed since plan was created (exit 40)",
            "ERR_VALIDATION_FAILED": "Plan validation checks failed (exit 10)",
            "ERR_LOCK_HELD": "File is locked by another process (exit 50)",
            "ERR_MISSING_DATA": "Required data argument not provided (exit 10)",
            "ERR_MISSING_PARAM": "Required parameter not provided (exit 10)",
            "ERR_ASSERTION_FAILED": "One or more assertions failed (exit 10)",
            "ERR_QUERY_FAILED": "SQL query execution failed (exit 90)",
            "ERR_OPERATION_FAILED": "Plan operation failed during apply (exit 90)",
            "ERR_PROTECTED_RANGE": "Range is protected by policy (exit 20)",
            "ERR_PLAN_TARGET_MISMATCH": "Plan compose targets different workbooks (exit 10)",
            "ERR_WORKFLOW_STEP_FAILED": "A workflow step failed during execution (exit 90)",
            "ERR_USAGE": "CLI usage error — missing or invalid arguments (exit 10)",
        },
        "exit_codes": {
            "0": "Success",
            "10": "Validation error (bad input, schema mismatch, invalid plan)",
            "20": "Protection/permission error (protected range or sheet)",
            "30": "Formula error (overwrite blocked, parse failure)",
            "40": "Conflict (fingerprint mismatch — workbook changed)",
            "50": "IO error (file not found, locked, permission denied)",
            "60": "Recalculation error",
            "70": "Unsupported feature or operation",
            "90": "Internal error (unexpected failure)",
        },
        "safety": {
            "dry_run": "All mutating commands support --dry-run to preview changes without writing",
            "backup": "All mutating commands support --backup to create a timestamped .bak copy",
            "fingerprint": "Patch plans record the workbook's xxhash; apply rejects if the file changed",
            "formula_protection": "cell set and formula set refuse to overwrite formulas unless --force-overwrite-formulas is used",
            "policy": "An optional xl-policy.yaml can restrict protected sheets, ranges, and mutation thresholds",
        },
        "examples": [
            {
                "name": "Add a calculated column to a table",
                "steps": [
                    "xl wb inspect -f budget.xlsx",
                    "xl table ls -f budget.xlsx",
                    "xl plan add-column -f budget.xlsx -t Sales -n GrossMarginPct --formula \"=[@GrossMargin]/[@Revenue]\" --out plan.json",
                    "xl validate plan -f budget.xlsx --plan plan.json",
                    "xl apply -f budget.xlsx --plan plan.json --dry-run",
                    "xl apply -f budget.xlsx --plan plan.json --backup",
                    "xl verify assert -f budget.xlsx --assertions '[{\"type\":\"table.column_exists\",\"table\":\"Sales\",\"column\":\"GrossMarginPct\"}]'",
                ],
            },
            {
                "name": "Query and analyze table data",
                "steps": [
                    "xl table ls -f sales.xlsx",
                    "xl query -f sales.xlsx --sql \"SELECT Region, SUM(Revenue) as Total FROM Sales GROUP BY Region ORDER BY Total DESC\"",
                    "xl range stat -f sales.xlsx --ref \"Sheet1!C2:C100\"",
                ],
            },
            {
                "name": "Format and freeze a report",
                "steps": [
                    "xl format number -f report.xlsx --ref \"Sales[Revenue]\" --style currency --decimals 2",
                    "xl format number -f report.xlsx --ref \"Sales[GrossMarginPct]\" --style percent --decimals 1",
                    "xl format width -f report.xlsx --sheet Sheet1 --columns A,B,C,D --width 15",
                    "xl format freeze -f report.xlsx --sheet Sheet1 --ref B2",
                ],
            },
            {
                "name": "Compare workbooks before and after changes",
                "steps": [
                    "xl diff compare --file-a original.xlsx --file-b modified.xlsx",
                ],
            },
        ],
    }
    env = success_envelope("guide", guide_data)
    _emit(env)


# ---------------------------------------------------------------------------
# xl wb inspect
# ---------------------------------------------------------------------------
@wb_app.command("inspect")
def wb_inspect(
    file: FilePath,
    json_out: JsonFlag = True,
):
    """Inspect workbook metadata — sheets, tables, named ranges, fingerprint.

    Returns the full structure of a workbook: sheet names and dimensions,
    all Excel tables with columns and row counts, named ranges, and a
    fingerprint hash for conflict detection. Start here when working
    with an unfamiliar workbook.

    Example: `xl wb inspect -f data.xlsx`

    See also: `xl sheet ls`, `xl table ls` for focused listing.
    """
    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "wb.inspect")
        meta = ctx.get_workbook_meta()
        ctx.close()

    env = success_envelope(
        "wb.inspect",
        meta.model_dump(),
        target=Target(file=file),
        duration_ms=t.elapsed_ms,
    )
    if meta.warnings:
        env.warnings = [WarningDetail(code="WORKBOOK_WARNING", message=w) for w in meta.warnings]
    _emit(env)


# ---------------------------------------------------------------------------
# xl wb create
# ---------------------------------------------------------------------------
@wb_app.command("create")
def wb_create(
    file: FilePath,
    sheets: Annotated[Optional[str], typer.Option("--sheets", help="Comma-separated sheet names (e.g. 'Revenue,Summary')")] = None,
    force: Annotated[bool, typer.Option("--force", help="Overwrite file if it already exists")] = False,
    json_out: JsonFlag = True,
):
    """Create a new empty workbook. Non-mutating (creates a new file).

    Creates a new .xlsx file. Errors if file already exists unless `--force`.
    Use `--sheets` to specify initial sheet names (default: one sheet named 'Sheet').

    Example: `xl wb create -f new.xlsx`

    Example: `xl wb create -f report.xlsx --sheets Revenue,Summary,Costs`

    See also: `xl sheet create` to add sheets to an existing workbook.
    """
    from xl.engine.context import WorkbookContext
    from xl.io.fileops import fingerprint

    p = Path(file).resolve()
    sheet_list = [s.strip() for s in sheets.split(",") if s.strip()] if sheets else None

    with Timer() as t:
        if p.exists() and not force:
            env = error_envelope(
                "wb.create", "ERR_FILE_EXISTS",
                f"File already exists: {p}. Use --force to overwrite.",
                target=Target(file=file),
            )
            _emit(env)
            return

        if p.exists() and force:
            p.unlink()

        try:
            ctx = WorkbookContext.create(p, sheets=sheet_list)
            meta = ctx.get_workbook_meta()
            ctx.close()
        except Exception as e:
            env = error_envelope("wb.create", "ERR_IO", str(e), target=Target(file=file))
            _emit(env)
            return

    result = {
        "path": str(p),
        "fingerprint": fingerprint(p),
        "sheets": [s.name for s in meta.sheets],
    }
    env = success_envelope("wb.create", result, target=Target(file=file), duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl sheet create
# ---------------------------------------------------------------------------
@sheet_app.command("create")
def sheet_create(
    file: FilePath,
    name: Annotated[str, typer.Option("--name", "-n", help="Name for the new sheet")],
    position: Annotated[Optional[int], typer.Option("--position", help="Zero-based index to insert the sheet at (default: append at end)")] = None,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Add a new sheet to an existing workbook. Mutating.

    Creates a new worksheet with the given name. Use `--position` to control
    where the sheet is inserted (0 = first). Errors if a sheet with that
    name already exists.

    Example: `xl sheet create -f data.xlsx --name Costs`

    Example: `xl sheet create -f data.xlsx --name Summary --position 0 --backup`

    See also: `xl sheet ls` to list existing sheets.
    """
    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "sheet.create")

        if name in ctx.wb.sheetnames:
            ctx.close()
            env = error_envelope(
                "sheet.create", "ERR_SHEET_EXISTS",
                f"Sheet '{name}' already exists in workbook",
                target=Target(file=file, sheet=name),
            )
            _emit(env)
            return

        ctx.wb.create_sheet(name, index=position)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {
        "dry_run": dry_run,
        "backup_path": backup_path,
        "sheet": name,
        "position": position,
    }
    env = success_envelope(
        "sheet.create", result,
        target=Target(file=file, sheet=name),
        changes=[ChangeRecord(
            type="sheet.create",
            target=name,
            after={"sheet": name, "position": position},
            impact={"cells": 0},
        )],
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl sheet ls
# ---------------------------------------------------------------------------
@sheet_app.command("ls")
def sheet_ls(
    file: FilePath,
    json_out: JsonFlag = True,
):
    """List all sheets in a workbook with name, index, and dimensions.

    Use this to discover valid sheet names for ref syntax (`SheetName!A1`).

    Example: `xl sheet ls -f data.xlsx`
    """
    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "sheet.ls")
        sheets = ctx.list_sheets()
        ctx.close()

    env = success_envelope(
        "sheet.ls",
        [s.model_dump() for s in sheets],
        target=Target(file=file),
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl table ls
# ---------------------------------------------------------------------------
@table_app.command("ls")
def table_ls(
    file: FilePath,
    sheet: SheetOpt = None,
    json_out: JsonFlag = True,
):
    """List all Excel tables with columns, row counts, and sheet locations.

    Returns table name, sheet, cell range, column names, and row count for
    each table. Use `--sheet` to filter to a single sheet.

    Example: `xl table ls -f data.xlsx`

    Example: `xl table ls -f data.xlsx --sheet Revenue`
    """
    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "table.ls")
        try:
            tables = ctx.list_tables(sheet)
        except ValueError as e:
            ctx.close()
            env = error_envelope("table.ls", "ERR_SHEET_NOT_FOUND", str(e), target=Target(file=file, sheet=sheet))
            _emit(env)
            return
        ctx.close()

    env = success_envelope(
        "table.ls",
        [tb.model_dump() for tb in tables],
        target=Target(file=file, sheet=sheet),
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl table add-column
# ---------------------------------------------------------------------------
@table_app.command("add-column")
def table_add_column_cmd(
    file: FilePath,
    table: Annotated[str, typer.Option("--table", "-t", help="Table name (as shown by 'xl table ls')")],
    name: Annotated[str, typer.Option("--name", "-n", help="New column name to add")],
    formula: Annotated[Optional[str], typer.Option("--formula", help="Column formula using structured refs (e.g. '=[@Revenue]-[@Cost]')")] = None,
    default_value: Annotated[Optional[str], typer.Option("--default", help="Default static value to fill in all rows")] = None,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Add a new column to an Excel table. Mutating.

    Appends a column to the right of existing columns. Use `--formula` for
    calculated columns with structured refs (`=[@ColName]`). Use `--default`
    for a static fill value. Provide both `--formula` and `--default` to set
    a formula with a fallback.

    Example: `xl table add-column -f data.xlsx -t Sales -n Margin --formula "=[@Revenue]-[@Cost]"`

    Example: `xl table add-column -f data.xlsx -t Sales -n Status --default "Active" --dry-run`

    See also: `xl plan add-column` to generate a plan instead of mutating directly.
    """
    from xl.adapters.openpyxl_engine import table_add_column
    from xl.io.fileops import backup as make_backup

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "table.add_column")

        try:
            change = table_add_column(ctx, table, name, formula=formula, default_value=default_value)
        except ValueError as e:
            ctx.close()
            err = str(e).lower()
            if "already exists" in err:
                code = "ERR_COLUMN_EXISTS"
            else:
                code = "ERR_TABLE_NOT_FOUND"
            env = error_envelope("table.add_column", code, str(e), target=Target(file=file, table=table))
            _emit(env)

        backup_path = None
        if not dry_run:
            if backup:
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {
        "dry_run": dry_run,
        "backup_path": backup_path,
    }
    env = success_envelope(
        "table.add_column",
        result,
        target=Target(file=file, table=table),
        changes=[change],
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl table append-rows
# ---------------------------------------------------------------------------
@table_app.command("append-rows")
def table_append_rows_cmd(
    file: FilePath,
    table: Annotated[str, typer.Option("--table", "-t", help="Table name (as shown by 'xl table ls')")],
    data: Annotated[Optional[str], typer.Option("--data", help="Inline JSON array of row objects, e.g. '[{\"Col1\":\"val\"}]'")] = None,
    data_file: Annotated[Optional[str], typer.Option("--data-file", help="Path to JSON file containing an array of row objects")] = None,
    schema_mode: Annotated[str, typer.Option("--schema-mode", help="'strict' (exact match), 'allow-missing-null' (missing cols→null), 'map-by-header' (match by name)")] = "strict",
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Append rows to an Excel table. Mutating.

    Provide row data as inline JSON (`--data`) or from a file (`--data-file`).
    Each row is an object with column names as keys. The `--schema-mode`
    controls how column mismatches are handled.

    Example: `xl table append-rows -f data.xlsx -t Sales --data '[{"Revenue":100,"Cost":60}]'`

    Example: `xl table append-rows -f data.xlsx -t Sales --data-file rows.json --schema-mode allow-missing-null`
    """
    from xl.adapters.openpyxl_engine import table_append_rows

    # Parse row data
    if data:
        try:
            rows = json.loads(data)
        except json.JSONDecodeError as e:
            env = error_envelope("table.append_rows", "ERR_INVALID_ARGUMENT", f"Malformed JSON in --data: {e}", target=Target(file=file, table=table))
            _emit(env)
            return
    elif data_file:
        try:
            rows = json.loads(read_text_safe(data_file))
        except (json.JSONDecodeError, OSError) as e:
            env = error_envelope("table.append_rows", "ERR_INVALID_ARGUMENT", f"Cannot read --data-file: {e}", target=Target(file=file, table=table))
            _emit(env)
            return
    else:
        env = error_envelope("table.append_rows", "ERR_MISSING_DATA", "Provide --data or --data-file", target=Target(file=file, table=table))
        _emit(env)
        return  # unreachable due to _emit raising

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "table.append_rows")

        try:
            change = table_append_rows(ctx, table, rows, schema_mode=schema_mode)
        except ValueError as e:
            ctx.close()
            code = "ERR_SCHEMA_MISMATCH" if "columns" in str(e).lower() else "ERR_TABLE_NOT_FOUND"
            env = error_envelope("table.append_rows", code, str(e), target=Target(file=file, table=table))
            _emit(env)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "table.append_rows",
        result,
        target=Target(file=file, table=table),
        changes=[change],
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl table create
# ---------------------------------------------------------------------------
@table_app.command("create")
def table_create_cmd(
    file: FilePath,
    table: Annotated[str, typer.Option("--table", "-t", help="Display name for the new table (e.g. 'Sales')")],
    sheet: Annotated[str, typer.Option("--sheet", "-s", help="Sheet name where the table will be created")],
    ref: Annotated[str, typer.Option("--ref", help="Cell range to promote to table (e.g. 'A1:D5')")],
    columns: Annotated[Optional[str], typer.Option("--columns", help="Comma-separated column headers (written if header row is empty)")] = None,
    style: Annotated[str, typer.Option("--style", help="Table style name (e.g. 'TableStyleMedium2')")] = "TableStyleMedium2",
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Create an Excel Table (ListObject) from a cell range. Mutating.

    Promotes an existing data range to an Excel Table object. The first
    row of the range must contain column headers (or use `--columns` to
    write headers to an empty first row).

    Example: `xl table create -f data.xlsx -t Sales -s Revenue --ref A1:D5`

    Example: `xl table create -f data.xlsx -t Metrics -s Sheet1 --ref A1:C1 --columns "Name,Value,Date"`

    See also: `xl table ls` to list existing tables, `xl plan create-table` to generate a plan.
    """
    from xl.adapters.openpyxl_engine import table_create
    from xl.io.fileops import backup as make_backup

    col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else None

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "table.create")

        try:
            change = table_create(ctx, sheet, table, ref, columns=col_list, style=style)
        except (ValueError, KeyError) as e:
            ctx.close()
            err = str(e).lower()
            if "already exists" in err:
                code = "ERR_TABLE_EXISTS"
            elif "overlap" in err:
                code = "ERR_TABLE_OVERLAP"
            elif "not found" in err:
                code = "ERR_SHEET_NOT_FOUND"
            else:
                code = "ERR_INVALID_ARGUMENT"
            env = error_envelope("table.create", code, str(e), target=Target(file=file, sheet=sheet, table=table))
            _emit(env)

        backup_path = None
        if not dry_run:
            if backup:
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {
        "dry_run": dry_run,
        "backup_path": backup_path,
    }
    env = success_envelope(
        "table.create",
        result,
        target=Target(file=file, sheet=sheet, table=table),
        changes=[change],
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl cell set
# ---------------------------------------------------------------------------
@cell_app.command("set")
def cell_set_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Cell reference as SheetName!Cell (e.g. Sheet1!B2)")],
    value: Annotated[str, typer.Option("--value", help="Value to write (coerced according to --type)")],
    cell_type: Annotated[Optional[str], typer.Option("--type", help="Value type: 'number', 'text', or 'bool'")] = None,
    force_overwrite_formulas: Annotated[bool, typer.Option("--force-overwrite-formulas", help="Allow overwriting a cell that contains a formula")] = False,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Set a cell value. Mutating.

    Writes a value to a single cell. The `--type` flag controls coercion:
    `number` parses as float/int, `bool` parses true/false/1/0, `text` keeps as string.
    Refuses to overwrite a formula unless `--force-overwrite-formulas` is set.

    Example: `xl cell set -f data.xlsx --ref "Sheet1!B2" --value 42 --type number`

    Example: `xl cell set -f data.xlsx --ref "Sheet1!A1" --value "Hello" --type text --backup`

    See also: `xl cell get` to read, `xl plan set-cells` to generate a plan instead.
    """
    from xl.adapters.openpyxl_engine import cell_set

    # Parse sheet!ref
    if "!" in ref:
        sheet_name, cell_ref = ref.split("!", 1)
    else:
        env = error_envelope("cell.set", "ERR_RANGE_INVALID", "Ref must include sheet name (e.g. Sheet1!B2)", target=Target(file=file))
        _emit(env)
        return

    # Coerce value
    parsed_value: Any = value
    if cell_type == "number":
        try:
            parsed_value = float(value)
            if parsed_value == int(parsed_value):
                parsed_value = int(parsed_value)
        except ValueError:
            pass
    elif cell_type == "bool":
        parsed_value = value.lower() in ("true", "1", "yes")

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "cell.set")

        try:
            change = cell_set(ctx, sheet_name, cell_ref, parsed_value, cell_type=cell_type, force_overwrite_formulas=force_overwrite_formulas)
        except (ValueError, KeyError) as e:
            ctx.close()
            code = "ERR_FORMULA_OVERWRITE_BLOCKED" if "formula" in str(e).lower() else "ERR_RANGE_INVALID"
            env = error_envelope("cell.set", code, str(e), target=Target(file=file, ref=ref))
            _emit(env)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "cell.set",
        result,
        target=Target(file=file, ref=ref),
        changes=[change],
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl validate workbook
# ---------------------------------------------------------------------------
@validate_app.command("workbook")
def validate_workbook_cmd(
    file: FilePath,
    json_out: JsonFlag = True,
):
    """Validate workbook health — check for corruption, broken refs, issues.

    Returns a list of checks with pass/fail status. Use this to verify a
    workbook is in good shape before performing operations.

    Example: `xl validate workbook -f data.xlsx`
    """
    from xl.validation.validators import validate_workbook

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "validate.workbook")
        result = validate_workbook(ctx)
        ctx.close()

    env = success_envelope(
        "validate.workbook",
        result.model_dump(),
        target=Target(file=file),
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl validate plan
# ---------------------------------------------------------------------------
@validate_app.command("plan")
def validate_plan_cmd(
    file: FilePath,
    plan_path: Annotated[str, typer.Option("--plan", help="Path to raw patch plan JSON file (use --out or --append)")],
    json_out: JsonFlag = True,
):
    """Validate a patch plan against a workbook before applying.

    Checks fingerprint match, table/column existence, and schema compatibility.
    Returns `ok: false` with detailed error checks if validation fails.
    Always run this before `xl apply`.

    Example: `xl validate plan -f data.xlsx --plan plan.json`

    See also: `xl apply --dry-run` for a preview of actual changes.
    """
    from xl.validation.validators import validate_plan

    try:
        plan = _load_patch_plan(plan_path)
    except ValueError as e:
        _emit_invalid_plan("validate.plan", file, str(e))
        return

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "validate.plan")
        result = validate_plan(ctx, plan)
        ctx.close()

    env = success_envelope(
        "validate.plan",
        result.model_dump(),
        target=Target(file=file),
        duration_ms=t.elapsed_ms,
    )
    if not result.valid:
        failed_checks = [c for c in result.checks if not c.get("passed", True)]
        env.ok = False
        env.errors = [
            ErrorDetail(
                code="ERR_VALIDATION_FAILED",
                message="Plan validation failed",
                details={"checks": failed_checks},
            )
        ]
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan show
# ---------------------------------------------------------------------------
@plan_app.command("show")
def plan_show(
    plan_path: Annotated[str, typer.Option("--plan", help="Path to raw patch plan JSON file (use --out or --append)")],
    json_out: JsonFlag = True,
):
    """Display the contents of a patch plan — operations, preconditions, target.

    Example: `xl plan show --plan plan.json`

    See also: `xl validate plan` to check the plan against a workbook.
    """
    try:
        plan = _load_patch_plan(plan_path)
    except ValueError as e:
        env = error_envelope("plan.show", "ERR_PLAN_INVALID", str(e), target=Target(file=plan_path))
        _emit(env)
        return

    env = success_envelope("plan.show", plan.model_dump())
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan add-column
# ---------------------------------------------------------------------------
@plan_app.command("add-column")
def plan_add_column(
    file: FilePath,
    table: Annotated[str, typer.Option("--table", "-t", help="Table name (as shown by 'xl table ls')")],
    name: Annotated[str, typer.Option("--name", "-n", help="New column name to add")],
    formula: Annotated[Optional[str], typer.Option("--formula", help="Column formula using structured refs (e.g. '=[@Col1]+[@Col2]')")] = None,
    default_value: Annotated[Optional[str], typer.Option("--default", help="Default static value to fill in all rows")] = None,
    append: Annotated[Optional[str], typer.Option("--append", help="Path to existing plan file to append this operation to")] = None,
    out: Annotated[Optional[str], typer.Option("--out", "-o", help="Write raw patch plan JSON to this file")] = None,
    json_out: JsonFlag = True,
):
    """Generate a plan to add a column to a table. Non-mutating.

    Outputs a patch plan envelope to stdout.
    Use `--out` to write a raw patch plan file for `xl validate plan` / `xl apply`.
    Use `--append` to append to (or create) a raw patch plan file.

    Example: `xl plan add-column -f data.xlsx -t Sales -n Margin --formula "=[@Revenue]-[@Cost]" --out plan.json`

    See also: `xl table add-column` to apply directly, `xl apply` to execute this plan.
    """
    from xl.io.fileops import fingerprint

    fp = fingerprint(file) if Path(file).exists() else None
    plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"

    op = Operation(
        op_id=f"op_{uuid.uuid4().hex[:6]}",
        type="table.add_column",
        table=table,
        name=name,
        formula=formula,
        value=default_value,
    )
    pre = Precondition(type="table_exists", table=table)
    post = Postcondition(type="column_exists", table=table, column=name)

    if append:
        if Path(append).exists():
            try:
                plan = _load_patch_plan(append)
            except ValueError as e:
                _emit_invalid_plan("plan.add_column", file, str(e))
                return
            if plan.target.file and Path(plan.target.file).resolve() != Path(file).resolve():
                env = error_envelope(
                    "plan.add_column",
                    "ERR_VALIDATION_FAILED",
                    f"Append plan target file '{plan.target.file}' does not match '{file}'",
                    target=Target(file=file),
                )
                _emit(env)
                return
        else:
            plan = PatchPlan(
                plan_id=plan_id,
                target=PlanTarget(file=file, fingerprint=fp),
            )
        plan.operations.append(op)
        plan.preconditions.append(pre)
        plan.postconditions.append(post)
        _write_plan(append, plan)
    else:
        plan = PatchPlan(
            plan_id=plan_id,
            target=PlanTarget(file=file, fingerprint=fp),
            preconditions=[pre],
            operations=[op],
            postconditions=[post],
        )

    if out:
        _write_plan(out, plan)

    env = success_envelope("plan.add_column", plan.model_dump(), target=Target(file=file, table=table))
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan create-table
# ---------------------------------------------------------------------------
@plan_app.command("create-table")
def plan_create_table(
    file: FilePath,
    table: Annotated[str, typer.Option("--table", "-t", help="Display name for the new table")],
    sheet: Annotated[str, typer.Option("--sheet", "-s", help="Sheet name where the table will be created")],
    ref: Annotated[str, typer.Option("--ref", help="Cell range for the table (e.g. 'A1:D5')")],
    columns: Annotated[Optional[str], typer.Option("--columns", help="Comma-separated column headers")] = None,
    style: Annotated[str, typer.Option("--style", help="Table style name")] = "TableStyleMedium2",
    append: Annotated[Optional[str], typer.Option("--append", help="Path to existing plan file to append to")] = None,
    out: Annotated[Optional[str], typer.Option("--out", "-o", help="Write raw patch plan JSON to file")] = None,
    json_out: JsonFlag = True,
):
    """Generate a plan to create an Excel Table from a range. Non-mutating.

    Outputs a patch plan envelope to stdout.
    Use `--out` to write a raw patch plan file for `xl validate plan` / `xl apply`.
    Use `--append` to append to (or create) a raw patch plan file.

    Example: `xl plan create-table -f data.xlsx -t Sales -s Revenue --ref A1:D5 --out plan.json`

    See also: `xl table create` to apply directly, `xl apply` to execute this plan.
    """
    from xl.io.fileops import fingerprint

    col_list = [c.strip() for c in columns.split(",") if c.strip()] if columns else None

    fp = fingerprint(file) if Path(file).exists() else None
    plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"

    op = Operation(
        op_id=f"op_{uuid.uuid4().hex[:6]}",
        type="table.create",
        table=table,
        sheet=sheet,
        ref=ref,
        columns=col_list,
        style=style,
    )
    pre = Precondition(type="sheet_exists", sheet=sheet)
    post = Postcondition(type="table_exists", table=table)

    if append:
        if Path(append).exists():
            try:
                plan = _load_patch_plan(append)
            except ValueError as e:
                _emit_invalid_plan("plan.create_table", file, str(e))
                return
            if plan.target.file and Path(plan.target.file).resolve() != Path(file).resolve():
                env = error_envelope(
                    "plan.create_table",
                    "ERR_VALIDATION_FAILED",
                    f"Append plan target file '{plan.target.file}' does not match '{file}'",
                    target=Target(file=file),
                )
                _emit(env)
                return
        else:
            plan = PatchPlan(
                plan_id=plan_id,
                target=PlanTarget(file=file, fingerprint=fp),
            )
        plan.operations.append(op)
        plan.preconditions.append(pre)
        plan.postconditions.append(post)
        _write_plan(append, plan)
    else:
        plan = PatchPlan(
            plan_id=plan_id,
            target=PlanTarget(file=file, fingerprint=fp),
            preconditions=[pre],
            operations=[op],
            postconditions=[post],
        )

    if out:
        _write_plan(out, plan)

    env = success_envelope("plan.create_table", plan.model_dump(), target=Target(file=file, sheet=sheet, table=table))
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan set-cells
# ---------------------------------------------------------------------------
@plan_app.command("set-cells")
def plan_set_cells(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Cell reference as SheetName!Cell (e.g. Sheet1!B2)")],
    value: Annotated[str, typer.Option("--value", help="Value to set (coerced according to --type)")],
    cell_type: Annotated[Optional[str], typer.Option("--type", help="Value type: 'number', 'text', or 'bool'")] = None,
    append: Annotated[Optional[str], typer.Option("--append", help="Path to existing plan file to append this operation to")] = None,
    out: Annotated[Optional[str], typer.Option("--out", "-o", help="Write raw patch plan JSON to this file")] = None,
    json_out: JsonFlag = True,
):
    """Generate a plan to set cell values. Non-mutating.

    Outputs a patch plan envelope to stdout.
    Use `--out` to write a raw patch plan file.
    Use `--append` to append to (or create) a raw patch plan file.

    Example: `xl plan set-cells -f data.xlsx --ref "Sheet1!B2" --value 42 --type number`

    See also: `xl cell set` to apply directly.
    """
    from xl.io.fileops import fingerprint

    fp = fingerprint(file) if Path(file).exists() else None
    plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"

    sheet_name = ""
    cell_ref = ref
    if "!" in ref:
        sheet_name, cell_ref = ref.split("!", 1)

    parsed_value: Any = value
    if cell_type == "number":
        try:
            parsed_value = float(value)
        except ValueError:
            pass

    op = Operation(
        op_id=f"op_{uuid.uuid4().hex[:6]}",
        type="cell.set",
        sheet=sheet_name,
        ref=cell_ref,
        value=parsed_value,
        cell_type=cell_type,
    )

    if append:
        if Path(append).exists():
            try:
                plan = _load_patch_plan(append)
            except ValueError as e:
                _emit_invalid_plan("plan.set_cells", file, str(e))
                return
            if plan.target.file and Path(plan.target.file).resolve() != Path(file).resolve():
                env = error_envelope(
                    "plan.set_cells",
                    "ERR_VALIDATION_FAILED",
                    f"Append plan target file '{plan.target.file}' does not match '{file}'",
                    target=Target(file=file),
                )
                _emit(env)
                return
        else:
            plan = PatchPlan(
                plan_id=plan_id,
                target=PlanTarget(file=file, fingerprint=fp),
            )
        plan.operations.append(op)
        if sheet_name:
            plan.preconditions.append(Precondition(type="sheet_exists", sheet=sheet_name))
        _write_plan(append, plan)
    else:
        plan = PatchPlan(
            plan_id=plan_id,
            target=PlanTarget(file=file, fingerprint=fp),
            preconditions=[Precondition(type="sheet_exists", sheet=sheet_name)] if sheet_name else [],
            operations=[op],
        )

    if out:
        _write_plan(out, plan)

    env = success_envelope("plan.set_cells", plan.model_dump(), target=Target(file=file, ref=ref))
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan format
# ---------------------------------------------------------------------------
@plan_app.command("format")
def plan_format(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Range as Sheet!A1:D10 or table column as TableName[Column]")],
    style: Annotated[str, typer.Option("--style", help="Format style: number, percent, currency, or date")] = "number",
    decimals: Annotated[int, typer.Option("--decimals", help="Decimal places (default: 2)")] = 2,
    append: Annotated[Optional[str], typer.Option("--append", help="Path to existing plan file to append this operation to")] = None,
    out: Annotated[Optional[str], typer.Option("--out", "-o", help="Write raw patch plan JSON to this file")] = None,
    json_out: JsonFlag = True,
):
    """Generate a plan for number formatting. Non-mutating.

    Outputs a patch plan envelope to stdout.
    Use `--out` to write a raw patch plan file.
    Use `--append` to append to (or create) a raw patch plan file.

    Example: `xl plan format -f data.xlsx --ref "Sales[Revenue]" --style currency --decimals 2`

    See also: `xl format number` to apply directly.
    """
    from xl.io.fileops import fingerprint

    fp = fingerprint(file) if Path(file).exists() else None

    op = Operation(
        op_id=f"op_{uuid.uuid4().hex[:6]}",
        type="format.number",
        ref=ref,
        style=style,
        decimals=decimals,
    )

    if append:
        if Path(append).exists():
            try:
                plan = _load_patch_plan(append)
            except ValueError as e:
                _emit_invalid_plan("plan.format", file, str(e))
                return
            if plan.target.file and Path(plan.target.file).resolve() != Path(file).resolve():
                env = error_envelope(
                    "plan.format",
                    "ERR_VALIDATION_FAILED",
                    f"Append plan target file '{plan.target.file}' does not match '{file}'",
                    target=Target(file=file),
                )
                _emit(env)
                return
        else:
            plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"
            plan = PatchPlan(
                plan_id=plan_id,
                target=PlanTarget(file=file, fingerprint=fp),
            )
        plan.operations.append(op)
        _write_plan(append, plan)
    else:
        plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"
        plan = PatchPlan(
            plan_id=plan_id,
            target=PlanTarget(file=file, fingerprint=fp),
            operations=[op],
        )

    if out:
        _write_plan(out, plan)

    env = success_envelope("plan.format", plan.model_dump(), target=Target(file=file, ref=ref))
    _emit(env)


# ---------------------------------------------------------------------------
# xl plan compose
# ---------------------------------------------------------------------------
@plan_app.command("compose")
def plan_compose(
    plans: Annotated[list[str], typer.Option("--plan", help="Plan files to merge (repeat --plan for each file)")],
    out: Annotated[Optional[str], typer.Option("--out", "-o", help="Write raw patch plan JSON to this file")] = None,
    json_out: JsonFlag = True,
):
    """Merge multiple plan files into a single composed plan.

    Combines operations, preconditions, and postconditions from all input
    plans. The target file and fingerprint are taken from the first plan.

    Example: `xl plan compose --plan step1.json --plan step2.json --plan step3.json`

    See also: `xl plan add-column --append` to build plans incrementally.
    """
    merged_ops: list[Operation] = []
    merged_pre: list[Precondition] = []
    merged_post: list[Postcondition] = []
    target_file = ""
    fp = None

    if not plans:
        env = error_envelope("plan.compose", "ERR_MISSING_PARAM", "Provide at least one --plan file")
        _emit(env)
        return

    for p in plans:
        try:
            plan = _load_patch_plan(p)
        except ValueError as e:
            env = error_envelope("plan.compose", "ERR_PLAN_INVALID", str(e), target=Target(file=p))
            _emit(env)
            return
        if not target_file:
            target_file = plan.target.file
            fp = plan.target.fingerprint
        else:
            if plan.target.fingerprint and fp and plan.target.fingerprint != fp:
                env = error_envelope("plan.compose", "ERR_PLAN_TARGET_MISMATCH",
                    f"Plan '{p}' targets a different workbook fingerprint than '{plans[0]}'",
                    target=Target(file=target_file))
                _emit(env)
                return
        merged_ops.extend(plan.operations)
        merged_pre.extend(plan.preconditions)
        merged_post.extend(plan.postconditions)

    plan_id = f"pln_{datetime.now(timezone.utc).strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}"
    composed = PatchPlan(
        plan_id=plan_id,
        target=PlanTarget(file=target_file, fingerprint=fp),
        preconditions=merged_pre,
        operations=merged_ops,
        postconditions=merged_post,
    )

    if out:
        _write_plan(out, composed)

    env = success_envelope("plan.compose", composed.model_dump(), target=Target(file=target_file))
    _emit(env)


# ---------------------------------------------------------------------------
# xl apply
# ---------------------------------------------------------------------------
@app.command("apply")
def apply_cmd(
    file: FilePath,
    plan_path: Annotated[str, typer.Option("--plan", help="Path to raw patch plan JSON file (use --out or --append)")],
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview all changes without writing to disk")] = False,
    do_backup: Annotated[bool, typer.Option("--backup/--no-backup", help="Create timestamped .bak copy before writing (default: on)")] = True,
    json_out: JsonFlag = True,
):
    """Apply a patch plan to a workbook. Mutating.

    Executes all operations in the plan against the workbook. Validates
    the plan first (fingerprint, preconditions). Use `--dry-run` to preview
    changes, then apply for real. Backup is on by default.

    Example (preview): `xl apply -f data.xlsx --plan plan.json --dry-run`

    Example (apply): `xl apply -f data.xlsx --plan plan.json --backup`

    Workflow: `xl plan ...` → `xl validate plan` → `xl apply --dry-run` → `xl apply`

    See also: `xl validate plan` to check before applying, `xl verify assert` to check after.
    """
    from xl.adapters.openpyxl_engine import (
        cell_set,
        format_number,
        resolve_table_column_ref,
        table_add_column,
        table_append_rows,
        table_create,
    )
    from xl.io.fileops import backup as make_backup
    from xl.io.fileops import fingerprint
    from xl.validation.validators import validate_plan

    # Load plan
    try:
        plan = _load_patch_plan(plan_path)
    except ValueError as e:
        _emit_invalid_plan("apply", file, str(e))
        return

    with Timer() as t:
        # Load workbook
        ctx = _load_ctx_or_emit(file, "apply")

        fp_before = ctx.fp

        # Fingerprint conflict check
        if plan.target.fingerprint and plan.options.fail_on_external_change:
            if plan.target.fingerprint != fp_before:
                ctx.close()
                env = error_envelope(
                    "apply", "ERR_PLAN_FINGERPRINT_CONFLICT",
                    "Workbook fingerprint changed since plan was created",
                    target=Target(file=file),
                    details={"expected": plan.target.fingerprint, "actual": fp_before},
                )
                _emit(env)
                return

        # Validate plan
        val_result = validate_plan(ctx, plan)
        if not val_result.valid:
            ctx.close()
            env = error_envelope(
                "apply", "ERR_VALIDATION_FAILED",
                "Plan validation failed",
                target=Target(file=file),
                details={"checks": val_result.checks},
            )
            _emit(env)
            return

        # Execute operations
        changes: list[ChangeRecord] = []
        for op in plan.operations:
            try:
                if op.type == "table.add_column":
                    change = table_add_column(ctx, op.table, op.name, formula=op.formula, default_value=op.value)
                    changes.append(change)
                elif op.type == "table.append_rows":
                    change = table_append_rows(ctx, op.table, op.rows or [], schema_mode=op.schema_mode or "strict")
                    changes.append(change)
                elif op.type == "cell.set":
                    sheet_name = op.sheet or ""
                    change = cell_set(ctx, sheet_name, op.ref or "", op.value, force_overwrite_formulas=op.force_overwrite_formulas)
                    changes.append(change)
                elif op.type == "format.number":
                    ref_str = op.ref or ""
                    sheet_name = ""
                    actual_ref = ref_str
                    # Resolve table column refs
                    resolved = resolve_table_column_ref(ctx, ref_str)
                    if resolved:
                        sheet_name, actual_ref = resolved
                    elif "!" in ref_str:
                        sheet_name, actual_ref = ref_str.split("!", 1)

                    if sheet_name:
                        change = format_number(ctx, sheet_name, actual_ref, style=op.style or "number", decimals=op.decimals if op.decimals is not None else 2)
                        changes.append(change)
                elif op.type == "table.create":
                    change = table_create(ctx, op.sheet or "", op.table or "", op.ref or "",
                                          columns=op.columns, style=op.style or "TableStyleMedium2")
                    changes.append(change)
                else:
                    changes.append(ChangeRecord(
                        op_id=op.op_id,
                        type=op.type,
                        target=str(op.ref or op.table or ""),
                        after={"status": "skipped", "reason": f"Unsupported operation type: {op.type}"},
                    ))
            except Exception as e:
                ctx.close()
                env = error_envelope(
                    "apply", "ERR_OPERATION_FAILED",
                    f"Operation {op.op_id} failed: {e}",
                    target=Target(file=file),
                )
                _emit(env)
                return

        # Save
        backup_path = None
        fp_after = None
        if not dry_run:
            if do_backup:
                backup_path = make_backup(file)
            ctx.save(file)
            fp_after = fingerprint(file)
        ctx.close()

    result_data = ApplyResult(
        applied=not dry_run,
        dry_run=dry_run,
        backup_path=backup_path,
        operations_applied=len(changes),
        fingerprint_before=fp_before,
        fingerprint_after=fp_after,
    ).model_dump()

    if dry_run and changes:
        from xl.engine.dispatcher import summarize_changes
        result_data["dry_run_summary"] = summarize_changes(changes)

    env = success_envelope(
        "apply",
        result_data,
        target=Target(file=file),
        changes=changes,
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl query
# ---------------------------------------------------------------------------
@app.command("query")
def query_cmd(
    file: FilePath,
    sql: Annotated[Optional[str], typer.Option("--sql", help="Full SQL query (all Excel tables are loaded as DuckDB tables by name)")] = None,
    table: Annotated[Optional[str], typer.Option("--table", "-t", help="Table name for shorthand mode (builds SELECT for you)")] = None,
    where: Annotated[Optional[str], typer.Option("--where", help="WHERE clause for shorthand mode (e.g. \"Revenue > 1000\")")] = None,
    select: Annotated[Optional[str], typer.Option("--select", help="Comma-separated column names for shorthand (default: all)")] = None,
    json_out: JsonFlag = True,
):
    """Query table data using SQL via DuckDB.

    All Excel tables in the workbook are loaded into DuckDB by their table
    name. Use `--sql` for full SQL, or `--table` + `--where` + `--select`
    for a shorthand query builder.

    Example (full SQL): `xl query -f data.xlsx --sql "SELECT Region, SUM(Sales) FROM Sales GROUP BY Region"`

    Example (shorthand): `xl query -f data.xlsx -t Sales --select "Region,Revenue" --where "Revenue > 1000"`

    See also: `xl table ls` to discover available table names.
    """
    import duckdb

    conn = None
    ctx = None

    # Build SQL if not provided directly (before the try block to avoid double-envelope)
    if sql is None:
        if table is None:
            env = error_envelope("query", "ERR_MISSING_PARAM", "Provide --sql or --table", target=Target(file=file))
            _emit(env)
            return
        cols = select if select else "*"
        sql = f"SELECT {cols} FROM {table}"
        if where:
            sql += f" WHERE {where}"

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "query")

        try:
            # Extract all tables to DuckDB
            conn = duckdb.connect()
            tables_found = ctx.list_tables()
            for tbl in tables_found:
                ws = ctx.wb[tbl.sheet]
                from xl.adapters.openpyxl_engine import _parse_ref
                min_row, min_col, max_row, max_col = _parse_ref(tbl.ref)
                col_names = [tc.name for tc in tbl.columns]
                data_rows = []
                for row_idx in range(min_row + 1, max_row + 1):
                    row_data = {}
                    for ci, col_name in enumerate(col_names):
                        cell = ws.cell(row=row_idx, column=min_col + ci)
                        row_data[col_name] = cell.value
                    data_rows.append(row_data)

                if data_rows:
                    # Deduplicate column names — suffix duplicates to avoid DuckDB catalog errors
                    seen: dict[str, int] = {}
                    deduped_col_names: list[str] = []
                    for cn in col_names:
                        if cn in seen:
                            seen[cn] += 1
                            deduped_col_names.append(f"{cn}_{seen[cn]}")
                        else:
                            seen[cn] = 0
                            deduped_col_names.append(cn)
                    col_names = deduped_col_names

                    # Build column defs and insert via parameterized VALUES for DuckDB compat
                    col_defs = []
                    for col_name in col_names:
                        sample = data_rows[0].get(col_name)
                        if isinstance(sample, int):
                            col_defs.append(f'"{col_name}" BIGINT')
                        elif isinstance(sample, float):
                            col_defs.append(f'"{col_name}" DOUBLE')
                        else:
                            col_defs.append(f'"{col_name}" VARCHAR')
                    conn.execute(f'CREATE TABLE "{tbl.name}" ({", ".join(col_defs)})')
                    placeholders = ", ".join(["?"] * len(col_names))
                    insert_sql = f'INSERT INTO "{tbl.name}" VALUES ({placeholders})'
                    rows_to_insert = [tuple(row_data.get(c) for c in col_names) for row_data in data_rows]
                    conn.executemany(insert_sql, rows_to_insert)

            cursor = conn.execute(sql)
            columns = [desc[0] for desc in cursor.description]
            raw_rows = cursor.fetchall()
            rows = [dict(zip(columns, row)) for row in raw_rows]
            row_count = len(rows)
        except (SystemExit, KeyboardInterrupt):
            raise
        except Exception as e:
            env = error_envelope("query", "ERR_QUERY_FAILED", str(e), target=Target(file=file))
            _emit(env)
            return
        finally:
            if conn is not None:
                conn.close()
            if ctx is not None:
                ctx.close()

    from xl.contracts.responses import QueryResult
    qr = QueryResult(columns=columns, rows=rows, row_count=row_count)
    env = success_envelope(
        "query",
        qr.model_dump(),
        target=Target(file=file),
        duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl cell get
# ---------------------------------------------------------------------------
@cell_app.command("get")
def cell_get_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Cell reference as SheetName!Cell (e.g. Sheet1!B2)")],
    data_only: Annotated[bool, typer.Option("--data-only", help="Return last-cached value for formula cells instead of the formula")] = False,
    json_out: JsonFlag = True,
):
    """Read a single cell's value, type, and formula.

    Returns the cell value, data type, and formula (if any). Use `--data-only`
    to get the last-computed value of formula cells rather than the formula text.

    Example: `xl cell get -f data.xlsx --ref "Sheet1!B2"`

    Example: `xl cell get -f data.xlsx --ref "Sheet1!B2" --data-only`

    See also: `xl cell set` to write, `xl range stat` for aggregate stats.
    """
    from xl.adapters.openpyxl_engine import cell_get

    if "!" not in ref:
        env = error_envelope("cell.get", "ERR_RANGE_INVALID", "Ref must include sheet name (e.g. Sheet1!B2)", target=Target(file=file))
        _emit(env)
        return

    sheet_name, cell_ref = ref.split("!", 1)

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "cell.get", data_only=data_only)

        try:
            result = cell_get(ctx, sheet_name, cell_ref)
        except (ValueError, KeyError) as e:
            ctx.close()
            env = error_envelope("cell.get", "ERR_RANGE_INVALID", str(e), target=Target(file=file, ref=ref))
            _emit(env)
            return
        ctx.close()

    warnings: list[WarningDetail] = []
    if data_only and result.get("value") is None and result.get("type") == "empty":
        warnings.append(WarningDetail(
            code="WARN_UNCACHED_FORMULA",
            message=(
                "Cell appears empty in data-only mode. It may contain a formula "
                "whose cached value was not saved by the last application that wrote "
                "this workbook. Re-run without --data-only to see the formula text."
            ),
        ))
    env = success_envelope("cell.get", result, target=Target(file=file, ref=ref), duration_ms=t.elapsed_ms)
    if warnings:
        env.warnings = warnings
    _emit(env)


# ---------------------------------------------------------------------------
# xl range stat
# ---------------------------------------------------------------------------
@range_app.command("stat")
def range_stat_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Range reference as SheetName!Start:End (e.g. Sheet1!C2:C100)")],
    data_only: Annotated[bool, typer.Option("--data-only", help="Use last-cached values for formula cells")] = False,
    json_out: JsonFlag = True,
):
    """Compute statistics for a cell range — min, max, mean, sum, count, stddev.

    Returns descriptive statistics for all numeric values in the range.

    Example: `xl range stat -f data.xlsx --ref "Sheet1!C2:C100"`

    See also: `xl query` for more complex aggregations via SQL.
    """
    from xl.adapters.openpyxl_engine import range_stat

    if "!" not in ref:
        env = error_envelope("range.stat", "ERR_RANGE_INVALID", "Ref must include sheet name", target=Target(file=file))
        _emit(env)
        return

    sheet_name, range_ref = ref.split("!", 1)

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "range.stat", data_only=data_only)
        result = range_stat(ctx, sheet_name, range_ref)
        ctx.close()

    env = success_envelope("range.stat", result, target=Target(file=file, ref=ref), duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl range clear
# ---------------------------------------------------------------------------
@range_app.command("clear")
def range_clear_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Range reference as SheetName!Start:End (e.g. Sheet1!A1:D10)")],
    contents: Annotated[bool, typer.Option("--contents", help="Clear values and formulas")] = False,
    formats: Annotated[bool, typer.Option("--formats", help="Clear number/style formatting")] = False,
    clear_all: Annotated[bool, typer.Option("--all", help="Clear both contents and formatting")] = False,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Clear a range of cells — values, formulas, and/or formatting. Mutating.

    By default clears cell contents (values and formulas). Use `--formats` to
    clear formatting, or `--all` for both.

    Example: `xl range clear -f data.xlsx --ref "Sheet1!A1:D10" --contents`

    Example: `xl range clear -f data.xlsx --ref "Sheet1!A1:D10" --all --backup`
    """
    from xl.adapters.openpyxl_engine import range_clear

    if "!" not in ref:
        env = error_envelope("range.clear", "ERR_RANGE_INVALID", "Ref must include sheet name", target=Target(file=file))
        _emit(env)
        return

    sheet_name, range_ref = ref.split("!", 1)
    if clear_all:
        do_contents = True
        do_formats = True
    elif contents or formats:
        do_contents = contents
        do_formats = formats
    else:
        # Default behavior remains content-only clear for backward compatibility.
        do_contents = True
        do_formats = False

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "range.clear")
        change = range_clear(ctx, sheet_name, range_ref, contents=do_contents, formats=do_formats)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "range.clear", result, target=Target(file=file, ref=ref),
        changes=[change], duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl formula set
# ---------------------------------------------------------------------------
@formula_app.command("set")
def formula_set_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Target: Sheet!A1, Sheet!A1:A10 (range fill), or TableName[Column]")],
    formula: Annotated[str, typer.Option("--formula", help="Excel formula (e.g. '=C2-D2' or '=[@Revenue]-[@Cost]')")],
    force_overwrite_values: Annotated[bool, typer.Option("--force-overwrite-values", help="Allow overwriting cells that contain plain values")] = False,
    force_overwrite_formulas: Annotated[bool, typer.Option("--force-overwrite-formulas", help="Allow overwriting cells that already contain formulas")] = False,
    fill_mode: Annotated[str, typer.Option("--fill-mode", help="'fixed' copies formula literally; 'relative' adjusts A1-refs per cell (like Excel fill-down)")] = "relative",
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Set a formula on a cell, range, or table column. Mutating.

    Supports three ref types:
    - Single cell: `Sheet1!E2` — sets formula on one cell
    - Range: `Sheet1!E2:E100` — fills the formula down the range
    - Table column: `Sales[Margin]` — fills all data rows in the column

    Safety: refuses to overwrite existing values or formulas unless the
    corresponding `--force-overwrite-*` flag is set.

    Example: `xl formula set -f data.xlsx --ref "Sheet1!E2" --formula "=C2-D2"`

    Example: `xl formula set -f data.xlsx --ref "Sales[Margin]" --formula "=[@Revenue]-[@Cost]"`

    See also: `xl formula lint` to check formula health, `xl formula find` to search.
    """
    from xl.adapters.openpyxl_engine import formula_set, resolve_table_column_ref

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "formula.set")

        # Resolve ref
        resolved = resolve_table_column_ref(ctx, ref, include_header=False)
        if resolved:
            sheet_name, cell_ref = resolved
        elif "!" in ref:
            sheet_name, cell_ref = ref.split("!", 1)
        else:
            ctx.close()
            env = error_envelope("formula.set", "ERR_RANGE_INVALID", "Ref must include sheet (Sheet!A1) or be a table column (Table[Col])", target=Target(file=file))
            _emit(env)
            return

        try:
            change = formula_set(ctx, sheet_name, cell_ref, formula,
                                 force_overwrite_values=force_overwrite_values,
                                 force_overwrite_formulas=force_overwrite_formulas,
                                 fill_mode=fill_mode)
        except (ValueError, KeyError) as e:
            ctx.close()
            env = error_envelope("formula.set", "ERR_FORMULA_BLOCKED", str(e), target=Target(file=file, ref=ref))
            _emit(env)
            return

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "formula.set", result, target=Target(file=file, ref=ref),
        changes=[change], duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl formula lint
# ---------------------------------------------------------------------------
@formula_app.command("lint")
def formula_lint_cmd(
    file: FilePath,
    sheet: SheetOpt = None,
    severity: Annotated[Optional[str], typer.Option("--severity", help="Minimum severity filter: info, warning, or error")] = None,
    category: Annotated[Optional[str], typer.Option("--category", help="Comma-separated category filter (e.g. 'volatile_function,broken_ref')")] = None,
    summary: Annotated[bool, typer.Option("--summary", help="Return grouped counts instead of individual findings")] = False,
    json_out: JsonFlag = True,
):
    """Lint formulas for common issues — volatile functions, broken refs, anti-patterns.

    Scans all formulas (or a single sheet with `--sheet`) and returns findings.
    Use `--severity` to filter by minimum severity level.
    Use `--category` to filter to specific categories.
    Use `--summary` to get aggregated counts instead of individual findings.

    Example: `xl formula lint -f data.xlsx`

    Example: `xl formula lint -f data.xlsx --severity error`

    Example: `xl formula lint -f data.xlsx --category volatile_function --summary`
    """
    from xl.adapters.openpyxl_engine import formula_lint

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "formula.lint")
        findings = formula_lint(ctx, sheet)
        ctx.close()

    # Apply severity filter
    severity_rank = {"info": 0, "warning": 1, "error": 2}
    if severity:
        min_rank = severity_rank.get(severity, 0)
        findings = [f for f in findings if severity_rank.get(f.get("severity", "info"), 0) >= min_rank]

    # Apply category filter
    if category:
        cats = {c.strip() for c in category.split(",") if c.strip()}
        findings = [f for f in findings if f.get("category") in cats]

    # Build summary
    by_category: dict[str, int] = {}
    by_severity: dict[str, int] = {}
    by_sheet: dict[str, int] = {}
    for f in findings:
        cat = f.get("category", "unknown")
        by_category[cat] = by_category.get(cat, 0) + 1
        sev = f.get("severity", "info")
        by_severity[sev] = by_severity.get(sev, 0) + 1
        ref = f.get("ref", "")
        sname = ref.split("!")[0] if "!" in ref else "unknown"
        by_sheet[sname] = by_sheet.get(sname, 0) + 1

    summary_data = {
        "total": len(findings),
        "by_category": by_category,
        "by_severity": by_severity,
        "by_sheet": by_sheet,
    }

    if summary:
        result = {"findings": [], "count": len(findings), "summary": summary_data}
    else:
        result = {"findings": findings, "count": len(findings), "summary": summary_data}

    env = success_envelope("formula.lint", result,
                           target=Target(file=file, sheet=sheet), duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl formula find
# ---------------------------------------------------------------------------
@formula_app.command("find")
def formula_find_cmd(
    file: FilePath,
    pattern: Annotated[str, typer.Option("--pattern", help="Regex pattern to match against formula text (e.g. 'VLOOKUP', 'SUM.*Revenue')")],
    sheet: SheetOpt = None,
    json_out: JsonFlag = True,
):
    """Search for formulas matching a regex pattern.

    Scans all formulas (or a single sheet with `--sheet`) and returns matches
    with cell location and formula text.

    Example: `xl formula find -f data.xlsx --pattern "VLOOKUP"`

    Example: `xl formula find -f data.xlsx --pattern "SUM" --sheet Revenue`
    """
    import re

    from xl.adapters.openpyxl_engine import formula_find

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "formula.find")

        try:
            matches = formula_find(ctx, pattern, sheet)
        except re.error as e:
            ctx.close()
            env = error_envelope("formula.find", "ERR_PATTERN_INVALID", str(e), target=Target(file=file, sheet=sheet))
            _emit(env)
            return
        except KeyError as e:
            ctx.close()
            env = error_envelope("formula.find", "ERR_RANGE_INVALID", str(e), target=Target(file=file, sheet=sheet))
            _emit(env)
            return
        except Exception as e:
            ctx.close()
            env = error_envelope("formula.find", "ERR_INTERNAL", str(e), target=Target(file=file, sheet=sheet))
            _emit(env)
            return
        ctx.close()

    env = success_envelope("formula.find", {"matches": matches, "count": len(matches)},
                           target=Target(file=file, sheet=sheet), duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl format number (CLI wiring for existing adapter)
# ---------------------------------------------------------------------------
@format_app.command("number")
def format_number_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Range as Sheet!A1:D10 or table column as TableName[Column]")],
    style: Annotated[str, typer.Option("--style", help="Format style: number, percent, currency, date, or text")] = "number",
    decimals: Annotated[int, typer.Option("--decimals", help="Decimal places (default: 2)")] = 2,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Apply number format to a range or table column. Mutating.

    Sets the display format for cells. Accepts both range refs and table column refs.

    Example: `xl format number -f data.xlsx --ref "Sheet1!C2:C100" --style currency --decimals 2`

    Example: `xl format number -f data.xlsx --ref "Sales[Revenue]" --style number --decimals 0`

    See also: `xl plan format` to generate a plan instead of applying directly.
    """
    from xl.adapters.openpyxl_engine import format_number, resolve_table_column_ref

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "format.number")

        resolved = resolve_table_column_ref(ctx, ref)
        if resolved:
            sheet_name, cell_ref = resolved
        elif "!" in ref:
            sheet_name, cell_ref = ref.split("!", 1)
        else:
            ctx.close()
            env = error_envelope("format.number", "ERR_RANGE_INVALID", "Ref must include sheet or be Table[Col]", target=Target(file=file))
            _emit(env)
            return

        try:
            change = format_number(ctx, sheet_name, cell_ref, style=style, decimals=decimals)
        except ValueError as e:
            ctx.close()
            env = error_envelope("format.number", "ERR_INVALID_ARGUMENT", str(e), target=Target(file=file, ref=ref))
            _emit(env)
            return

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "format.number", result, target=Target(file=file, ref=ref),
        changes=[change], duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl format width
# ---------------------------------------------------------------------------
@format_app.command("width")
def format_width_cmd(
    file: FilePath,
    sheet: Annotated[str, typer.Option("--sheet", "-s", help="Sheet name (as shown by 'xl sheet ls')")],
    columns: Annotated[str, typer.Option("--columns", help="Comma-separated column letters (e.g. A,B,C)")],
    width: Annotated[float, typer.Option("--width", help="Column width in character units (Excel default is ~8.43)")],
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Set column widths on a sheet. Mutating.

    Example: `xl format width -f data.xlsx --sheet Sheet1 --columns A,B,C --width 15`
    """
    import re

    from openpyxl.utils import column_index_from_string

    from xl.adapters.openpyxl_engine import format_width

    col_list = [c.strip().upper() for c in columns.split(",") if c.strip()]
    if not col_list:
        env = error_envelope("format.width", "ERR_RANGE_INVALID", "Provide at least one column letter", target=Target(file=file, sheet=sheet))
        _emit(env)
        return

    invalid_cols: list[str] = []
    for col in col_list:
        if not re.fullmatch(r"[A-Z]{1,3}", col):
            invalid_cols.append(col)
            continue
        try:
            column_index_from_string(col)
        except ValueError:
            invalid_cols.append(col)

    if invalid_cols:
        env = error_envelope(
            "format.width",
            "ERR_RANGE_INVALID",
            f"Invalid column tokens: {', '.join(invalid_cols)}",
            target=Target(file=file, sheet=sheet),
        )
        _emit(env)
        return

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "format.width")
        change = format_width(ctx, sheet, col_list, width)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "format.width", result, target=Target(file=file, sheet=sheet),
        changes=[change], duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl format freeze
# ---------------------------------------------------------------------------
@format_app.command("freeze")
def format_freeze_cmd(
    file: FilePath,
    sheet: Annotated[str, typer.Option("--sheet", "-s", help="Sheet name (as shown by 'xl sheet ls')")],
    ref: Annotated[Optional[str], typer.Option("--ref", help="Cell below/right of frozen area (e.g. B2 freezes row 1 and column A)")] = None,
    unfreeze: Annotated[bool, typer.Option("--unfreeze", help="Remove all freeze panes from the sheet")] = False,
    backup: Annotated[bool, typer.Option("--backup", help="Create timestamped .bak copy before writing")] = False,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Preview changes without writing to disk")] = False,
    json_out: JsonFlag = True,
):
    """Freeze or unfreeze panes on a sheet. Mutating.

    Freezing at B2 keeps row 1 and column A visible while scrolling.
    Use `--unfreeze` to remove all freeze panes.

    Example: `xl format freeze -f data.xlsx --sheet Sheet1 --ref B2`

    Example: `xl format freeze -f data.xlsx --sheet Sheet1 --unfreeze`
    """
    from xl.adapters.openpyxl_engine import format_freeze

    if unfreeze and ref:
        env = error_envelope("format.freeze", "ERR_INVALID_ARGUMENT", "Use either --ref or --unfreeze, not both", target=Target(file=file, sheet=sheet))
        _emit(env)
        return

    freeze_ref = None if unfreeze else ref
    if not unfreeze and not ref:
        env = error_envelope("format.freeze", "ERR_MISSING_PARAM", "Provide --ref or --unfreeze", target=Target(file=file))
        _emit(env)
        return

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "format.freeze")
        change = format_freeze(ctx, sheet, freeze_ref)

        backup_path = None
        if not dry_run:
            if backup:
                from xl.io.fileops import backup as make_backup
                backup_path = make_backup(file)
            ctx.save(file)
        ctx.close()

    result = {"dry_run": dry_run, "backup_path": backup_path}
    env = success_envelope(
        "format.freeze", result, target=Target(file=file, sheet=sheet),
        changes=[change], duration_ms=t.elapsed_ms,
    )
    _emit(env)


# ---------------------------------------------------------------------------
# xl validate refs
# ---------------------------------------------------------------------------
@validate_app.command("refs")
def validate_refs_cmd(
    file: FilePath,
    ref: Annotated[str, typer.Option("--ref", help="Reference to validate as SheetName!Range (e.g. Sheet1!A1:D10)")],
    json_out: JsonFlag = True,
):
    """Validate that a reference points to a valid sheet and cell range.

    Checks that the sheet exists and the range is valid. Useful before
    running cell/range/formula commands to catch typos early.

    Example: `xl validate refs -f data.xlsx --ref "Sheet1!A1:D10"`
    """
    from xl.adapters.openpyxl_engine import _parse_ref

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "validate.refs")

        checks: list[dict] = []
        if "!" in ref:
            sheet_name, range_ref = ref.split("!", 1)
            if sheet_name in ctx.wb.sheetnames:
                checks.append({"type": "sheet_exists", "target": sheet_name, "passed": True, "message": f"Sheet '{sheet_name}' exists"})
                try:
                    _parse_ref(range_ref)
                    checks.append({"type": "range_valid", "target": ref, "passed": True, "message": f"Range '{range_ref}' is valid"})
                except ValueError as e:
                    checks.append({"type": "range_valid", "target": ref, "passed": False, "message": str(e)})
            else:
                checks.append({"type": "sheet_exists", "target": sheet_name, "passed": False, "message": f"Sheet '{sheet_name}' not found"})
        else:
            checks.append({"type": "ref_format", "target": ref, "passed": False, "message": "Reference must include sheet name (Sheet!A1)"})

        ctx.close()

    valid = all(c.get("passed", True) for c in checks)
    result = ValidationResult(valid=valid, checks=checks)
    env = success_envelope("validate.refs", result.model_dump(), target=Target(file=file, ref=ref), duration_ms=t.elapsed_ms)
    if not valid:
        env.ok = False
        env.errors = [
            ErrorDetail(
                code="ERR_RANGE_INVALID",
                message="Reference validation failed",
                details={"checks": checks},
            )
        ]
    _emit(env)


# ---------------------------------------------------------------------------
# xl validate workflow
# ---------------------------------------------------------------------------
@validate_app.command("workflow")
def validate_workflow_cmd(
    workflow_file: Annotated[str, typer.Option("--workflow", "-w", help="Path to YAML workflow file to validate")],
    json_out: JsonFlag = True,
):
    """Validate a workflow YAML file — no workbook required.

    Checks YAML syntax, top-level keys, step IDs, step commands, and
    structure. Returns pass/fail checks in a ValidationResult.

    Example: `xl validate workflow -w pipeline.yaml`

    See also: `xl run` to execute a workflow.
    """
    from xl.engine.workflow import validate_workflow

    with Timer() as t:
        result = validate_workflow(workflow_file)

    env = success_envelope("validate.workflow", result, duration_ms=t.elapsed_ms)
    if not result.get("valid", True):
        env.ok = False
        failed = [c for c in result.get("checks", []) if not c.get("passed")]
        env.errors = [
            ErrorDetail(
                code="ERR_WORKFLOW_INVALID",
                message="Workflow validation failed",
                details={"checks": failed},
            )
        ]
    _emit(env)


# ---------------------------------------------------------------------------
# xl wb lock-status
# ---------------------------------------------------------------------------
@wb_app.command("lock-status")
def wb_lock_status_cmd(
    file: FilePath,
    json_out: JsonFlag = True,
):
    """Check if a workbook file is locked by another process.

    Returns lock status and lock holder info if available. Check this
    before mutating operations to avoid ERR_LOCK_HELD errors.

    Example: `xl wb lock-status -f data.xlsx`
    """
    from xl.io.fileops import check_lock

    with Timer() as t:
        result = check_lock(file)

    env = success_envelope("wb.lock_status", result, target=Target(file=file), duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl verify assert
# ---------------------------------------------------------------------------
@verify_app.command("assert")
def verify_assert_cmd(
    file: FilePath,
    assertions: Annotated[Optional[str], typer.Option("--assertions", help="Inline JSON array of assertion objects (e.g. '[{\"type\":\"table.column_exists\",...}]')")] = None,
    assertions_file: Annotated[Optional[str], typer.Option("--assertions-file", help="Path to JSON file containing an array of assertion objects")] = None,
    json_out: JsonFlag = True,
):
    """Run post-apply assertions to verify workbook state.

    Checks that the workbook matches expected conditions after mutations.
    Returns `ok: false` if any assertion fails.

    **Assertion types:** `table.column_exists`, `cell.value_equals`, `cell.not_empty`,
    `cell.value_type`, `table.row_count`, `table.row_count.gte` (alias: `row_count.gte`).

    Example: `xl verify assert -f data.xlsx --assertions '[{"type":"table.column_exists","table":"Sales","column":"Margin"}]'`

    Example: `xl verify assert -f data.xlsx --assertions-file checks.json`

    See also: `xl validate workbook` for structural health checks.
    """
    from xl.engine.verify import run_assertions

    if assertions and assertions_file:
        env = error_envelope("verify.assert", "ERR_INVALID_ARGUMENT", "Use either --assertions or --assertions-file", target=Target(file=file))
        _emit(env)
        return

    if not assertions and not assertions_file:
        env = error_envelope("verify.assert", "ERR_MISSING_DATA", "Provide --assertions or --assertions-file", target=Target(file=file))
        _emit(env)
        return

    try:
        if assertions:
            assertion_list = json.loads(assertions)
        else:
            assertion_list = json.loads(read_text_safe(assertions_file))
    except Exception as e:
        env = error_envelope("verify.assert", "ERR_VALIDATION_FAILED", f"Cannot parse assertions: {e}", target=Target(file=file))
        _emit(env)
        return

    if not isinstance(assertion_list, list):
        env = error_envelope("verify.assert", "ERR_VALIDATION_FAILED", "Assertions payload must be a JSON array", target=Target(file=file))
        _emit(env)
        return

    with Timer() as t:
        ctx = _load_ctx_or_emit(file, "verify.assert")
        results = run_assertions(ctx, assertion_list)
        ctx.close()

    all_passed = all(r.get("passed", False) for r in results)
    result = {"passed": all_passed, "assertions": results, "total": len(results), "passed_count": sum(1 for r in results if r.get("passed"))}
    env = success_envelope("verify.assert", result, target=Target(file=file), duration_ms=t.elapsed_ms)
    if not all_passed:
        env.ok = False
        failed = [r for r in results if not r.get("passed", False)]
        env.errors = [
            ErrorDetail(
                code="ERR_ASSERTION_FAILED",
                message=f"{len(failed)} assertion(s) failed",
                details={"failed": failed},
            )
        ]
    _emit(env)


# ---------------------------------------------------------------------------
# xl diff compare
# ---------------------------------------------------------------------------
@diff_app.command("compare")
def diff_compare_cmd(
    file_a: Annotated[str, typer.Option("--file-a", help="First (original/before) workbook path")],
    file_b: Annotated[str, typer.Option("--file-b", help="Second (modified/after) workbook path")],
    sheet: SheetOpt = None,
    include_formulas: Annotated[bool, typer.Option("--include-formulas", help="Include formula text changes in addition to value changes")] = False,
    json_out: JsonFlag = True,
):
    """Compare two workbook files cell by cell.

    Returns cell-level value changes, sheets added/removed, and fingerprint
    comparison. Use `--sheet` to limit comparison to a single sheet.
    Use `--include-formulas` to also compare formula text (loads workbooks twice).

    Example: `xl diff compare --file-a original.xlsx --file-b modified.xlsx`

    Example: `xl diff compare --file-a v1.xlsx --file-b v2.xlsx --sheet Revenue`

    Useful for reviewing changes after `xl apply`.
    """
    from xl.diff.differ import diff_workbooks

    with Timer() as t:
        try:
            result = diff_workbooks(file_a, file_b, sheet_filter=sheet, include_formulas=include_formulas)
        except FileNotFoundError as e:
            env = error_envelope("diff.compare", "ERR_WORKBOOK_NOT_FOUND", str(e))
            _emit(env)
            return
        except ValueError as e:
            env = error_envelope("diff.compare", "ERR_RANGE_INVALID", str(e))
            _emit(env)
            return

    env = success_envelope("diff.compare", result, duration_ms=t.elapsed_ms)
    _emit(env)


# ---------------------------------------------------------------------------
# xl run
# ---------------------------------------------------------------------------
@app.command("run")
def run_cmd(
    workflow_file: Annotated[str, typer.Option("--workflow", "-w", help="Path to YAML workflow file defining steps to execute")],
    file: Annotated[Optional[str], typer.Option("--file", "-f", help="Override the target workbook (instead of workflow's target.file)")] = None,
    json_out: JsonFlag = True,
):
    """Execute a multi-step YAML workflow.

    Runs a sequence of xl commands defined in a YAML file. The workflow
    specifies a target workbook and a list of steps.

    Supported step commands:

    Inspection: wb.inspect, sheet.ls, table.ls, cell.get, range.stat, query, formula.find, formula.lint

    Mutation: table.add_column, table.append_rows, cell.set, formula.set, format.number, format.width, format.freeze, range.clear

    Validation: validate.plan, validate.workbook, validate.refs, verify.assert

    Other: apply, diff.compare

    Example YAML workflow::

        schema_version: "1.0"
        name: pipeline
        target: { file: data.xlsx }
        steps:
          - { id: inspect, run: wb.inspect }
          - { id: add_col, run: table.add_column, args: { table: Sales, name: Margin, formula: "=[@Revenue]-[@Cost]" } }
          - { id: check, run: verify.assert, args: { assertions: [{ type: table.column_exists, table: Sales, column: Margin }] } }

    Example: `xl run --workflow pipeline.yaml -f data.xlsx`

    See also: `xl apply` for single-plan execution.
    """
    from xl.engine.workflow import WorkflowValidationError, execute_workflow, load_workflow

    with Timer() as t:
        try:
            workflow = load_workflow(workflow_file)
        except WorkflowValidationError as e:
            env = error_envelope(
                "run", "ERR_WORKFLOW_INVALID", str(e),
                details={"issues": e.details},
            )
            _emit(env)
            return
        except Exception as e:
            env = error_envelope("run", "ERR_WORKFLOW_INVALID", f"Cannot parse workflow: {e}")
            _emit(env)
            return

        workbook_path = file or workflow.target.get("file", "")
        if not workbook_path:
            env = error_envelope("run", "ERR_MISSING_PARAM", "Provide --file or set target.file in workflow")
            _emit(env)
            return

        try:
            result = execute_workflow(workflow, workbook_path)
        except Exception as e:
            env = error_envelope("run", "ERR_WORKFLOW_FAILED", str(e), target=Target(file=workbook_path))
            _emit(env)
            return

    env = success_envelope("run", result, target=Target(file=workbook_path), duration_ms=t.elapsed_ms)
    if not result.get("ok"):
        env.ok = False
        for step in result.get("steps", []):
            if not step.get("ok") and step.get("error"):
                env.errors.append(ErrorDetail(
                    code="ERR_WORKFLOW_STEP_FAILED",
                    message=f"Step '{step['step_id']}' ({step['run']}): {step['error']}",
                ))
    _emit(env)


# ---------------------------------------------------------------------------
# xl serve --stdio
# ---------------------------------------------------------------------------
@app.command("serve")
def serve_cmd(
    stdio: Annotated[bool, typer.Option("--stdio", help="Use stdin/stdout for JSON request/response (for agent tool integration)")] = True,
):
    """Start stdio server for agent tool integration (MCP/ACP).

    Reads JSON commands from stdin and writes JSON responses to stdout.
    Each line is a JSON object: `{"id": "1", "command": "wb.inspect", "args": {"file": "data.xlsx"}}`

    Example: `xl serve --stdio`
    """
    from xl.server.stdio import StdioServer
    server = StdioServer()
    server.run()


# ---------------------------------------------------------------------------
# Entrypoint (for `python -m xl`)
# ---------------------------------------------------------------------------
def main() -> None:
    try:
        app()
    except SystemExit:
        raise
    except Exception as exc:
        # Catch-all: any unhandled exception gets wrapped in a proper JSON
        # error envelope so machine consumers never see raw tracebacks.
        env = error_envelope(
            "unknown",
            "ERR_INTERNAL",
            str(exc),
        )
        print_response(env)
        raise SystemExit(90) from exc
    finally:
        # Suppress openpyxl PermissionError during temp-file cleanup on Windows.
        # openpyxl registers an atexit handler that may fail with [WinError 32]
        # when temp files are still locked. Registering here (after openpyxl's
        # handler) means atexit LIFO runs ours first, silencing stderr.
        if sys.platform == "win32":
            import atexit
            import io
            atexit.register(lambda: setattr(sys, "stderr", io.StringIO()))


if __name__ == "__main__":
    main()
