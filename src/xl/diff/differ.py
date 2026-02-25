"""Workbook diff logic — compare two workbook files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from xl.io.fileops import fingerprint


def diff_workbooks(
    path_a: str | Path,
    path_b: str | Path,
    sheet_filter: str | None = None,
    *,
    include_formulas: bool = False,
) -> dict[str, Any]:
    """Compare two workbook files and return structured diff."""
    wb_a = openpyxl.load_workbook(str(path_a), data_only=True)
    wb_b = openpyxl.load_workbook(str(path_b), data_only=True)

    fp_a = fingerprint(path_a)
    fp_b = fingerprint(path_b)

    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)

    sheets_added = sorted(sheets_b - sheets_a)
    sheets_removed = sorted(sheets_a - sheets_b)
    sheets_common = sorted(sheets_a & sheets_b)

    if sheet_filter:
        missing_in: list[str] = []
        if sheet_filter not in sheets_a:
            missing_in.append(f"file_a ({path_a})")
        if sheet_filter not in sheets_b:
            missing_in.append(f"file_b ({path_b})")
        if missing_in:
            wb_a.close()
            wb_b.close()
            raise ValueError(f"Sheet '{sheet_filter}' not found in {', '.join(missing_in)}")
        sheets_common = [s for s in sheets_common if s == sheet_filter]

    cell_changes: list[dict[str, Any]] = []

    for sname in sheets_common:
        ws_a = wb_a[sname]
        ws_b = wb_b[sname]

        # Get the union of all cells
        max_row = max(ws_a.max_row or 1, ws_b.max_row or 1)
        max_col = max(ws_a.max_column or 1, ws_b.max_column or 1)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_a = ws_a.cell(row=row, column=col).value
                val_b = ws_b.cell(row=row, column=col).value
                if val_a != val_b:
                    cell_ref = f"{sname}!{get_column_letter(col)}{row}"
                    if val_a is None:
                        change_type = "added"
                    elif val_b is None:
                        change_type = "removed"
                    else:
                        change_type = "modified"
                    cell_changes.append({
                        "ref": cell_ref,
                        "change_type": change_type,
                        "before": val_a,
                        "after": val_b,
                    })

    wb_a.close()
    wb_b.close()

    # Formula-level comparison (loads workbooks with data_only=False)
    formula_changes: list[dict[str, Any]] = []
    if include_formulas:
        wb_a_f = openpyxl.load_workbook(str(path_a), data_only=False)
        wb_b_f = openpyxl.load_workbook(str(path_b), data_only=False)

        for sname in sheets_common:
            ws_a_f = wb_a_f[sname]
            ws_b_f = wb_b_f[sname]
            max_row = max(ws_a_f.max_row or 1, ws_b_f.max_row or 1)
            max_col = max(ws_a_f.max_column or 1, ws_b_f.max_column or 1)

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    val_a = ws_a_f.cell(row=row, column=col).value
                    val_b = ws_b_f.cell(row=row, column=col).value
                    is_formula_a = isinstance(val_a, str) and val_a.startswith("=")
                    is_formula_b = isinstance(val_b, str) and val_b.startswith("=")
                    if (is_formula_a or is_formula_b) and val_a != val_b:
                        cell_ref = f"{sname}!{get_column_letter(col)}{row}"
                        formula_changes.append({
                            "ref": cell_ref,
                            "change_type": "formula_modified",
                            "before": val_a,
                            "after": val_b,
                        })

        wb_a_f.close()
        wb_b_f.close()

    result = {
        "file_a": str(path_a),
        "file_b": str(path_b),
        "fingerprint_a": fp_a,
        "fingerprint_b": fp_b,
        "identical": fp_a == fp_b,
        "sheets_added": sheets_added,
        "sheets_removed": sheets_removed,
        "cell_changes": cell_changes,
        "total_changes": len(cell_changes) + len(sheets_added) + len(sheets_removed),
    }
    if include_formulas:
        result["formula_changes"] = formula_changes
        result["total_changes"] += len(formula_changes)
    return result
