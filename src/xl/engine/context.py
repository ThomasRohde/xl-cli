"""WorkbookContext: loads a workbook, provides metadata and fingerprint."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from xl.contracts.common import Target, WorkbookCorruptError
from xl.contracts.responses import (
    NamedRangeMeta,
    SheetMeta,
    TableColumnMeta,
    TableMeta,
    WorkbookMeta,
)
from xl.io.fileops import fingerprint


class WorkbookContext:
    """Wraps an openpyxl workbook with metadata and helper methods."""

    @classmethod
    def create(cls, path: str | Path, *, sheets: list[str] | None = None) -> "WorkbookContext":
        """Create a new workbook file. Raises FileExistsError if path exists."""
        p = Path(path).resolve()
        if p.exists():
            raise FileExistsError(f"File already exists: {p}")
        wb = Workbook()
        if sheets:
            wb.active.title = sheets[0]
            for name in sheets[1:]:
                wb.create_sheet(name)
        wb.save(str(p))
        wb.close()
        return cls(p)

    def __init__(self, path: str | Path, *, data_only: bool = False) -> None:
        self.path = Path(path).resolve()
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")
        self.fp = fingerprint(self.path)
        try:
            self.wb: Workbook = openpyxl.load_workbook(
                str(self.path), data_only=data_only
            )
        except Exception as e:
            raise WorkbookCorruptError(f"Cannot open workbook {self.path}: {e}") from e

    def target(self, **overrides: str | None) -> Target:
        t = Target(file=str(self.path))
        for k, v in overrides.items():
            if v is not None:
                setattr(t, k, v)
        return t

    def get_workbook_meta(self) -> WorkbookMeta:
        sheets: list[SheetMeta] = []
        for idx, name in enumerate(self.wb.sheetnames):
            ws: Worksheet = self.wb[name]
            vis = "visible"
            if ws.sheet_state == "hidden":
                vis = "hidden"
            elif ws.sheet_state == "veryHidden":
                vis = "veryHidden"
            used = ws.dimensions if ws.dimensions else None
            tbl_count = len(ws.tables) if hasattr(ws, "tables") else 0
            sheets.append(SheetMeta(
                name=name, index=idx, visible=vis,
                used_range=used, table_count=tbl_count,
            ))

        names: list[NamedRangeMeta] = []
        for dn in self.wb.defined_names.values():
            names.append(NamedRangeMeta(
                name=dn.name,
                scope="workbook" if dn.localSheetId is None else self.wb.sheetnames[dn.localSheetId],
                ref=str(dn.attr_text),
            ))

        has_macros = self.path.suffix.lower() == ".xlsm" or self.wb.vba_archive is not None
        has_external = bool(getattr(self.wb, "_external_links", []))

        warnings: list[str] = []
        if has_macros:
            warnings.append("Workbook contains macros (VBA). xl will not execute them.")
        if has_external:
            warnings.append("Workbook contains external links.")

        return WorkbookMeta(
            path=str(self.path),
            fingerprint=self.fp,
            sheets=sheets,
            names=names,
            has_macros=has_macros,
            has_external_links=has_external,
            warnings=warnings,
        )

    def list_sheets(self) -> list[SheetMeta]:
        return self.get_workbook_meta().sheets

    def list_tables(self, sheet: str | None = None) -> list[TableMeta]:
        tables: list[TableMeta] = []
        sheet_names = [sheet] if sheet else self.wb.sheetnames
        for sname in sheet_names:
            if sname not in self.wb.sheetnames:
                if sheet:  # explicitly requested sheet doesn't exist
                    raise ValueError(f"Sheet not found: {sname}")
                continue
            ws: Worksheet = self.wb[sname]
            for tbl in ws._tables.values():
                tbl_name = tbl.displayName
                cols = [
                    TableColumnMeta(name=col.name, index=i)
                    for i, col in enumerate(tbl.tableColumns)
                ]
                ref = tbl.ref or ""
                row_count = 0
                if ref and ":" in ref:
                    parts = ref.split(":")
                    try:
                        import re
                        start_row = int(re.sub(r"[^0-9]", "", parts[0]))
                        end_row = int(re.sub(r"[^0-9]", "", parts[1]))
                        row_count = max(0, end_row - start_row)  # minus header
                    except (ValueError, IndexError):
                        pass

                tables.append(TableMeta(
                    table_id=f"tbl_{sname}_{tbl_name}".lower().replace(" ", "_"),
                    name=tbl_name,
                    sheet=sname,
                    ref=ref,
                    columns=cols,
                    style=tbl.tableStyleInfo.name if tbl.tableStyleInfo else None,
                    totals_row=bool(tbl.totalsRowCount),
                    row_count_estimate=row_count,
                ))
        return tables

    def get_sheet(self, name: str) -> Worksheet:
        if name not in self.wb.sheetnames:
            raise KeyError(f"Sheet not found: {name}")
        return self.wb[name]

    def find_table(self, table_name: str) -> tuple[Worksheet, object] | None:
        """Find a table by name across all sheets. Returns (worksheet, Table) or None."""
        for sname in self.wb.sheetnames:
            ws = self.wb[sname]
            for tbl in ws._tables.values():
                if tbl.displayName == table_name:
                    return ws, tbl
        return None

    def save(self, path: str | Path | None = None) -> bytes:
        """Save workbook to bytes. Optionally save to a path."""
        from io import BytesIO
        buf = BytesIO()
        self.wb.save(buf)
        data = buf.getvalue()
        if path:
            from xl.io.fileops import atomic_write
            atomic_write(path, data)
        return data

    def close(self) -> None:
        self.wb.close()
