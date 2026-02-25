"""Shared test fixtures."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


FIXTURES_DIR = Path(__file__).parent / "fixtures"
WORKBOOKS_DIR = FIXTURES_DIR / "workbooks"
PLANS_DIR = FIXTURES_DIR / "plans"


def _save_and_reload(wb: Workbook, path: Path) -> None:
    """Save workbook and reload so tableColumns get populated."""
    wb.save(str(path))
    # Reload triggers XML roundtrip which populates tableColumns
    wb2 = openpyxl.load_workbook(str(path))
    wb2.save(str(path))
    wb2.close()


@pytest.fixture()
def simple_workbook(tmp_path: Path) -> Path:
    """Create a simple workbook with one sheet and one table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    # Write table data
    headers = ["Region", "Product", "Sales", "Cost"]
    ws.append(headers)
    ws.append(["North", "Widget", 1000, 600])
    ws.append(["South", "Widget", 1500, 900])
    ws.append(["East", "Gadget", 2000, 1100])
    ws.append(["West", "Gadget", 800, 500])

    # Create table
    tab = Table(displayName="Sales", ref="A1:D5")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Add a second sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Sales"
    ws2["B1"] = "=SUM(Revenue!C2:C5)"

    path = tmp_path / "test_workbook.xlsx"
    _save_and_reload(wb, path)
    return path


@pytest.fixture()
def multi_table_workbook(tmp_path: Path) -> Path:
    """Create a workbook with multiple tables."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Table 1: Products
    ws["A1"] = "ProductID"
    ws["B1"] = "Name"
    ws["C1"] = "Price"
    ws["A2"] = 1
    ws["B2"] = "Widget"
    ws["C2"] = 10.99
    ws["A3"] = 2
    ws["B3"] = "Gadget"
    ws["C3"] = 24.99
    ws["A4"] = 3
    ws["B4"] = "Doohickey"
    ws["C4"] = 5.49

    tab1 = Table(displayName="Products", ref="A1:C4")
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(tab1)

    # Table 2: Orders (offset to the right)
    ws["E1"] = "OrderID"
    ws["F1"] = "ProductID"
    ws["G1"] = "Quantity"
    ws["E2"] = 101
    ws["F2"] = 1
    ws["G2"] = 5
    ws["E3"] = 102
    ws["F3"] = 2
    ws["G3"] = 3

    tab2 = Table(displayName="Orders", ref="E1:G3")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(tab2)

    path = tmp_path / "multi_table.xlsx"
    _save_and_reload(wb, path)
    return path


@pytest.fixture()
def raw_data_workbook(tmp_path: Path) -> Path:
    """Create a workbook with raw data (no Excel Tables)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value", "Category"])
    ws.append(["Alpha", 100, "A"])
    ws.append(["Beta", 200, "B"])
    ws.append(["Gamma", 300, "A"])
    path = tmp_path / "raw_data.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture()
def formula_table_workbook(tmp_path: Path) -> Path:
    """Create a workbook with a table that has a formula column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    ws.append(["Name", "Amount", "Tax"])
    ws.append(["Alice", 100, "=[@Amount]*0.1"])
    ws.append(["Bob", 200, "=[@Amount]*0.1"])
    ws.append(["Charlie", 300, "=[@Amount]*0.1"])

    tab = Table(displayName="Payments", ref="A1:C4")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    path = tmp_path / "formula_table.xlsx"
    _save_and_reload(wb, path)
    return path


@pytest.fixture()
def sample_plan(simple_workbook: Path) -> dict:
    """Create a sample patch plan dict."""
    from xl.io.fileops import fingerprint

    fp = fingerprint(simple_workbook)
    return {
        "schema_version": "1.0",
        "plan_id": "pln_test_001",
        "target": {
            "file": str(simple_workbook),
            "fingerprint": fp,
        },
        "options": {
            "recalc_mode": "cached",
            "backup": False,
            "fail_on_external_change": True,
        },
        "preconditions": [
            {"type": "table_exists", "table": "Sales"},
        ],
        "operations": [
            {
                "op_id": "op_1",
                "type": "table.add_column",
                "table": "Sales",
                "name": "Margin",
                "formula": "=[@Sales]-[@Cost]",
            },
        ],
        "postconditions": [
            {"type": "column_exists", "table": "Sales", "column": "Margin"},
        ],
    }


@pytest.fixture()
def sample_plan_file(tmp_path: Path, sample_plan: dict) -> Path:
    """Write sample plan to a JSON file."""
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(sample_plan))
    return plan_path
