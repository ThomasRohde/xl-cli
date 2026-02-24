"""Generate golden workbook fixtures for snapshot testing.

Run this script to (re)create the golden .xlsx files under tests/fixtures/workbooks/.
These files have deterministic content used by golden snapshot tests.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


WORKBOOKS_DIR = Path(__file__).parent / "workbooks"
WORKBOOKS_DIR.mkdir(exist_ok=True)


def _save_and_reload(wb: Workbook, path: Path) -> None:
    wb.save(str(path))
    wb2 = openpyxl.load_workbook(str(path))
    wb2.save(str(path))
    wb2.close()


def create_sales_workbook() -> Path:
    """Standard sales workbook with one table, formulas, and a summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    headers = ["Region", "Product", "Sales", "Cost"]
    ws.append(headers)
    ws.append(["North", "Widget", 1000, 600])
    ws.append(["South", "Widget", 1500, 900])
    ws.append(["East", "Gadget", 2000, 1100])
    ws.append(["West", "Gadget", 800, 500])

    tab = Table(displayName="Sales", ref="A1:D5")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Sales"
    ws2["B1"] = "=SUM(Revenue!C2:C5)"

    path = WORKBOOKS_DIR / "golden_sales.xlsx"
    _save_and_reload(wb, path)
    return path


def create_multi_table_workbook() -> Path:
    """Workbook with two tables on one sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "ProductID"
    ws["B1"] = "Name"
    ws["C1"] = "Price"
    ws["A2"] = 1
    ws["B2"] = "Widget"
    ws["C2"] = 10.99
    ws["A3"] = 2
    ws["B3"] = "Gadget"
    ws["C3"] = 24.99
    ws["A4"] = 3
    ws["B4"] = "Doohickey"
    ws["C4"] = 5.49

    tab1 = Table(displayName="Products", ref="A1:C4")
    tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(tab1)

    ws["E1"] = "OrderID"
    ws["F1"] = "ProductID"
    ws["G1"] = "Quantity"
    ws["E2"] = 101
    ws["F2"] = 1
    ws["G2"] = 5
    ws["E3"] = 102
    ws["F3"] = 2
    ws["G3"] = 3

    tab2 = Table(displayName="Orders", ref="E1:G3")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(tab2)

    path = WORKBOOKS_DIR / "golden_multi_table.xlsx"
    _save_and_reload(wb, path)
    return path


def create_formulas_workbook() -> Path:
    """Workbook with various formula patterns for lint/find testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    ws["A1"] = "Value"
    ws["B1"] = "Formula"
    ws["A2"] = 100
    ws["B2"] = "=A2*2"
    ws["A3"] = 200
    ws["B3"] = "=SUM(A2:A3)"
    ws["A4"] = 300
    ws["B4"] = "=VLOOKUP(A4,A2:B3,2,FALSE)"
    ws["A5"] = "=NOW()"
    ws["B5"] = "=INDIRECT(\"A\"&ROW())"
    ws["A6"] = "=#REF!+1"

    path = WORKBOOKS_DIR / "golden_formulas.xlsx"
    wb.save(str(path))
    return path


def create_empty_workbook() -> Path:
    """Minimal workbook with a single empty sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    path = WORKBOOKS_DIR / "golden_empty.xlsx"
    wb.save(str(path))
    return path


def create_hidden_sheets_workbook() -> Path:
    """Workbook with visible and hidden sheets for hygiene checks."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Visible"
    ws["A1"] = "public data"

    hidden = wb.create_sheet("HiddenSheet")
    hidden["A1"] = "hidden data"
    hidden.sheet_state = "hidden"

    very_hidden = wb.create_sheet("VeryHidden")
    very_hidden["A1"] = "very hidden data"
    very_hidden.sheet_state = "veryHidden"

    path = WORKBOOKS_DIR / "golden_hidden_sheets.xlsx"
    wb.save(str(path))
    return path


if __name__ == "__main__":
    files = [
        create_sales_workbook(),
        create_multi_table_workbook(),
        create_formulas_workbook(),
        create_empty_workbook(),
        create_hidden_sheets_workbook(),
    ]
    for f in files:
        print(f"Created: {f}")
