"""Tests for openpyxl_engine adapter operations."""

from pathlib import Path

import pytest

from xl.adapters.openpyxl_engine import (
    cell_set,
    format_number,
    table_add_column,
    table_append_rows,
    table_create,
)
from xl.engine.context import WorkbookContext


def test_table_add_column(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = table_add_column(ctx, "Sales", "Margin", formula="=[@Sales]-[@Cost]")
    assert change.type == "table.add_column"
    assert "Margin" in change.target

    # Verify column was added
    tables = ctx.list_tables()
    col_names = [c.name for c in tables[0].columns]
    assert "Margin" in col_names
    ctx.close()


def test_table_add_column_not_found(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    with pytest.raises(ValueError, match="Table not found"):
        table_add_column(ctx, "NonExistent", "Col")
    ctx.close()


def test_table_add_column_with_default(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = table_add_column(ctx, "Sales", "Status", default_value="Active")
    assert change.type == "table.add_column"

    # Verify values
    ws = ctx.wb["Revenue"]
    # Header should be at col E (5th column), row 1
    assert ws.cell(row=1, column=5).value == "Status"
    assert ws.cell(row=2, column=5).value == "Active"
    ctx.close()


def test_table_append_rows(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    rows = [
        {"Region": "Central", "Product": "Widget", "Sales": 1200, "Cost": 700},
    ]
    change = table_append_rows(ctx, "Sales", rows)
    assert change.type == "table.append_rows"
    assert change.after["rows_added"] == 1
    ctx.close()


def test_table_append_rows_strict_schema(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    rows = [{"Region": "Central", "Product": "Widget"}]  # missing Sales, Cost
    with pytest.raises(ValueError, match="Missing columns"):
        table_append_rows(ctx, "Sales", rows, schema_mode="strict")
    ctx.close()


def test_table_append_rows_extra_columns(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    rows = [
        {"Region": "Central", "Product": "Widget", "Sales": 100, "Cost": 50, "Extra": "bad"},
    ]
    with pytest.raises(ValueError, match="Extra columns"):
        table_append_rows(ctx, "Sales", rows, schema_mode="strict")
    ctx.close()


def test_cell_set(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = cell_set(ctx, "Revenue", "A2", "Updated")
    assert change.type == "cell.set"
    assert change.before == "North"
    assert change.after == "Updated"
    ctx.close()


def test_cell_set_formula_overwrite_blocked(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    # Summary!B1 has a formula
    with pytest.raises(ValueError, match="formula"):
        cell_set(ctx, "Summary", "B1", 999)
    ctx.close()


def test_cell_set_formula_overwrite_forced(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = cell_set(ctx, "Summary", "B1", 999, force_overwrite_formulas=True)
    assert change.after == 999
    ctx.close()


def test_format_number(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = format_number(ctx, "Revenue", "C2:C5", style="number", decimals=2)
    assert change.type == "format.number"
    assert change.impact["cells"] == 4

    # Verify format was applied
    ws = ctx.wb["Revenue"]
    assert ws.cell(row=2, column=3).number_format == "#,##0.00"
    ctx.close()


def test_format_percent(simple_workbook: Path):
    ctx = WorkbookContext(simple_workbook)
    change = format_number(ctx, "Revenue", "C2:C5", style="percent", decimals=1)
    ws = ctx.wb["Revenue"]
    assert ws.cell(row=2, column=3).number_format == "0.0%"
    ctx.close()


# ---------------------------------------------------------------------------
# table_create tests
# ---------------------------------------------------------------------------
def test_table_create_from_existing_data(raw_data_workbook: Path):
    """Create a table from a range that already has headers and data."""
    ctx = WorkbookContext(raw_data_workbook)
    change = table_create(ctx, "Data", "Metrics", "A1:C4")
    assert change.type == "table.create"
    assert "Metrics" in change.target
    assert change.after["columns"] == ["Name", "Value", "Category"]
    assert change.after["ref"] == "A1:C4"
    assert change.impact["rows"] == 3  # 3 data rows

    # Verify table exists
    tables = ctx.list_tables()
    assert len(tables) == 1
    assert tables[0].name == "Metrics"
    ctx.close()


def test_table_create_with_columns(tmp_path: Path):
    """Create a table with explicit column headers on empty range."""
    from openpyxl import Workbook as WB

    wb = WB()
    ws = wb.active
    ws.title = "Sheet1"
    path = tmp_path / "empty.xlsx"
    wb.save(str(path))
    wb.close()

    ctx = WorkbookContext(path)
    change = table_create(ctx, "Sheet1", "NewTable", "A1:C1",
                          columns=["Col1", "Col2", "Col3"])
    assert change.type == "table.create"
    assert change.after["columns"] == ["Col1", "Col2", "Col3"]

    # Verify headers were written
    ws = ctx.wb["Sheet1"]
    assert ws.cell(row=1, column=1).value == "Col1"
    assert ws.cell(row=1, column=2).value == "Col2"
    assert ws.cell(row=1, column=3).value == "Col3"
    ctx.close()


def test_table_create_duplicate_name(simple_workbook: Path):
    """Error when table name already exists."""
    ctx = WorkbookContext(simple_workbook)
    with pytest.raises(ValueError, match="already exists"):
        table_create(ctx, "Revenue", "Sales", "F1:H3",
                     columns=["X", "Y", "Z"])
    ctx.close()


def test_table_create_overlap(simple_workbook: Path):
    """Error when range overlaps existing table."""
    ctx = WorkbookContext(simple_workbook)
    with pytest.raises(ValueError, match="overlap"):
        table_create(ctx, "Revenue", "NewTable", "A1:D3",
                     columns=["A", "B", "C", "D"])
    ctx.close()


def test_table_create_invalid_name(tmp_path: Path):
    """Error for invalid table name format."""
    from openpyxl import Workbook as WB

    wb = WB()
    ws = wb.active
    ws.title = "Data"
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    wb.close()

    ctx = WorkbookContext(path)
    with pytest.raises(ValueError, match="Invalid table name"):
        table_create(ctx, "Data", "123bad", "A1:C1",
                     columns=["X", "Y", "Z"])
    ctx.close()
