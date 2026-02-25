"""Tests for CLI commands via Typer test runner."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from typer.testing import CliRunner

from xl.cli import app

runner = CliRunner()


def test_version():
    result = runner.invoke(app, ["version"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert "version" in data["result"]


def test_guide():
    result = runner.invoke(app, ["guide"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["command"] == "guide"
    guide = data["result"]
    # Check all top-level sections are present
    assert "overview" in guide
    assert "workflow" in guide
    assert "commands" in guide
    assert "ref_syntax" in guide
    assert "response_format" in guide
    assert "error_codes" in guide
    assert "exit_codes" in guide
    assert "safety" in guide
    assert "examples" in guide
    # Check commands are grouped
    assert "inspection" in guide["commands"]
    assert "mutation" in guide["commands"]
    assert "reading" in guide["commands"]
    # Check workflow has steps
    assert len(guide["workflow"]["steps"]) >= 5
    # Check examples are present
    assert len(guide["examples"]) >= 3


def test_wb_inspect(simple_workbook: Path):
    result = runner.invoke(app, ["wb", "inspect", "--file", str(simple_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["command"] == "wb.inspect"
    assert len(data["result"]["sheets"]) == 2
    assert data["result"]["fingerprint"].startswith("sha256:")


def test_wb_inspect_not_found(tmp_path: Path):
    result = runner.invoke(app, ["wb", "inspect", "--file", str(tmp_path / "nope.xlsx")])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_WORKBOOK_NOT_FOUND"


def test_sheet_ls(simple_workbook: Path):
    result = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    names = [s["name"] for s in data["result"]]
    assert "Revenue" in names
    assert "Summary" in names


def test_table_ls(simple_workbook: Path):
    result = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert len(data["result"]) == 1
    assert data["result"][0]["name"] == "Sales"


def test_table_ls_filter_sheet(simple_workbook: Path):
    result = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook), "--sheet", "Summary"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"] == []


def test_table_add_column_dry_run(simple_workbook: Path):
    result = runner.invoke(app, [
        "table", "add-column",
        "--file", str(simple_workbook),
        "--table", "Sales",
        "--name", "Margin",
        "--formula", "=[@Sales]-[@Cost]",
        "--dry-run",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True
    assert len(data["changes"]) == 1
    assert data["changes"][0]["type"] == "table.add_column"


def test_table_add_column_apply(simple_workbook: Path):
    result = runner.invoke(app, [
        "table", "add-column",
        "--file", str(simple_workbook),
        "--table", "Sales",
        "--name", "Margin",
        "--formula", "=[@Sales]-[@Cost]",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is False

    # Verify column persisted
    result = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    col_names = [c["name"] for c in data["result"][0]["columns"]]
    assert "Margin" in col_names


def test_table_append_rows(simple_workbook: Path):
    rows = json.dumps([
        {"Region": "Central", "Product": "Widget", "Sales": 1200, "Cost": 700},
    ])
    result = runner.invoke(app, [
        "table", "append-rows",
        "--file", str(simple_workbook),
        "--table", "Sales",
        "--data", rows,
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["after"]["rows_added"] == 1


def test_cell_set(simple_workbook: Path):
    result = runner.invoke(app, [
        "cell", "set",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2",
        "--value", "Updated",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["after"] == "Updated"


def test_validate_workbook(simple_workbook: Path):
    result = runner.invoke(app, ["validate", "workbook", "--file", str(simple_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["valid"] is True


def test_validate_plan(simple_workbook: Path, sample_plan_file: Path):
    result = runner.invoke(app, [
        "validate", "plan",
        "--file", str(simple_workbook),
        "--plan", str(sample_plan_file),
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["valid"] is True


def test_plan_show(sample_plan_file: Path):
    result = runner.invoke(app, ["plan", "show", "--plan", str(sample_plan_file)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["plan_id"] == "pln_test_001"


def test_plan_add_column(simple_workbook: Path):
    result = runner.invoke(app, [
        "plan", "add-column",
        "--file", str(simple_workbook),
        "--table", "Sales",
        "--name", "Margin",
        "--formula", "=[@Sales]-[@Cost]",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    plan = data["result"]
    assert plan["target"]["file"] == str(simple_workbook)
    assert len(plan["operations"]) == 1
    assert plan["operations"][0]["type"] == "table.add_column"


def test_apply_dry_run(simple_workbook: Path, sample_plan_file: Path):
    result = runner.invoke(app, [
        "apply",
        "--file", str(simple_workbook),
        "--plan", str(sample_plan_file),
        "--dry-run",
        "--no-backup",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True
    assert data["result"]["applied"] is False
    assert len(data["changes"]) == 1


def test_apply_with_backup(simple_workbook: Path, sample_plan_file: Path):
    result = runner.invoke(app, [
        "apply",
        "--file", str(simple_workbook),
        "--plan", str(sample_plan_file),
        "--backup",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["applied"] is True
    assert data["result"]["backup_path"] is not None
    assert Path(data["result"]["backup_path"]).exists()


def test_apply_fingerprint_conflict(simple_workbook: Path, tmp_path: Path):
    """Plan with wrong fingerprint should fail."""
    plan_data = {
        "schema_version": "1.0",
        "plan_id": "pln_conflict",
        "target": {
            "file": str(simple_workbook),
            "fingerprint": "sha256:0000000000000000000000000000000000000000000000000000000000000000",
        },
        "options": {"fail_on_external_change": True},
        "operations": [],
    }
    plan_path = tmp_path / "bad_plan.json"
    plan_path.write_text(json.dumps(plan_data))

    result = runner.invoke(app, [
        "apply",
        "--file", str(simple_workbook),
        "--plan", str(plan_path),
    ])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert "FINGERPRINT" in data["errors"][0]["code"]


# ---------------------------------------------------------------------------
# wb create
# ---------------------------------------------------------------------------
def test_wb_create_basic(tmp_path: Path):
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["wb", "create", "--file", str(out)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["command"] == "wb.create"
    assert out.exists()
    assert data["result"]["fingerprint"].startswith("sha256:")
    assert len(data["result"]["sheets"]) == 1  # default sheet


def test_wb_create_with_sheets(tmp_path: Path):
    out = tmp_path / "new.xlsx"
    result = runner.invoke(app, ["wb", "create", "--file", str(out), "--sheets", "Revenue,Summary,Costs"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["sheets"] == ["Revenue", "Summary", "Costs"]


def test_wb_create_already_exists(simple_workbook: Path):
    result = runner.invoke(app, ["wb", "create", "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert "FILE_EXISTS" in data["errors"][0]["code"]


def test_wb_create_force(simple_workbook: Path):
    result = runner.invoke(app, ["wb", "create", "--file", str(simple_workbook), "--force"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert simple_workbook.exists()


# ---------------------------------------------------------------------------
# sheet create
# ---------------------------------------------------------------------------
def test_sheet_create_basic(simple_workbook: Path):
    result = runner.invoke(app, ["sheet", "create", "--file", str(simple_workbook), "--name", "NewSheet"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["command"] == "sheet.create"
    assert data["result"]["sheet"] == "NewSheet"

    # Verify the sheet now exists
    ls_result = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    ls_data = json.loads(ls_result.stdout)
    names = [s["name"] for s in ls_data["result"]]
    assert "NewSheet" in names


def test_sheet_create_duplicate(simple_workbook: Path):
    result = runner.invoke(app, ["sheet", "create", "--file", str(simple_workbook), "--name", "Revenue"])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert "SHEET_EXISTS" in data["errors"][0]["code"]


def test_sheet_create_dry_run(simple_workbook: Path):
    result = runner.invoke(app, ["sheet", "create", "--file", str(simple_workbook), "--name", "DrySheet", "--dry-run"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True

    # Verify the sheet was NOT actually created
    ls_result = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    ls_data = json.loads(ls_result.stdout)
    names = [s["name"] for s in ls_data["result"]]
    assert "DrySheet" not in names


def test_sheet_create_position(simple_workbook: Path):
    result = runner.invoke(app, ["sheet", "create", "--file", str(simple_workbook), "--name", "First", "--position", "0"])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True

    # Verify position
    ls_result = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    ls_data = json.loads(ls_result.stdout)
    assert ls_data["result"][0]["name"] == "First"


# ---------------------------------------------------------------------------
# validate workflow
# ---------------------------------------------------------------------------
def test_validate_workflow_valid(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text(
        "schema_version: '1.0'\n"
        "name: test\n"
        "steps:\n"
        "  - id: step1\n"
        "    run: wb.inspect\n"
        "  - id: step2\n"
        "    run: table.ls\n"
        "    args: {sheet: Revenue}\n"
    )
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["valid"] is True


def test_validate_workflow_unknown_keys(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("bogus_key: 1\nsteps:\n  - {id: s1, run: wb.inspect}\n")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    checks = data["result"]["checks"]
    unknown_check = next(c for c in checks if c["type"] == "unknown_keys")
    assert unknown_check["passed"] is False


def test_validate_workflow_invalid_command(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("steps:\n  - {id: s1, run: bogus.cmd}\n")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    data = json.loads(result.stdout)
    assert data["ok"] is False


def test_validate_workflow_missing_id(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("steps:\n  - {run: wb.inspect}\n")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    checks = data["result"]["checks"]
    id_check = next(c for c in checks if c["type"] == "step_id")
    assert id_check["passed"] is False


def test_validate_workflow_duplicate_ids(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("steps:\n  - {id: s1, run: wb.inspect}\n  - {id: s1, run: table.ls}\n")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    checks = data["result"]["checks"]
    dup_check = next(c for c in checks if c["type"] == "step_id_unique")
    assert dup_check["passed"] is False


def test_validate_workflow_bad_yaml(tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("{{not valid yaml")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    data = json.loads(result.stdout)
    assert data["ok"] is False


# ---------------------------------------------------------------------------
# xl run — structured validation errors (Feature 3)
# ---------------------------------------------------------------------------
def test_run_missing_required_arg(simple_workbook: Path, tmp_path: Path):
    """table.add_column requires 'table' and 'name' — omitting them gives structured errors."""
    wf = tmp_path / "wf.yaml"
    wf.write_text(
        "steps:\n"
        "  - id: s1\n"
        "    run: table.add_column\n"
        "    args: {formula: '=1+1'}\n"  # missing table and name
    )
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert "WORKFLOW_INVALID" in data["errors"][0]["code"]
    issues = data["errors"][0]["details"]["issues"]
    missing_args = [i["arg"] for i in issues if i["type"] == "missing_required_arg"]
    assert "table" in missing_args
    assert "name" in missing_args


def test_run_unknown_arg_name(simple_workbook: Path, tmp_path: Path):
    """Typo in arg name (e.g. 'tabble') gives a structured unknown_arg error."""
    wf = tmp_path / "wf.yaml"
    wf.write_text(
        "steps:\n"
        "  - id: s1\n"
        "    run: table.add_column\n"
        "    args: {tabble: Sales, name: Margin}\n"
    )
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    issues = data["errors"][0]["details"]["issues"]
    unknown = [i for i in issues if i["type"] == "unknown_arg"]
    assert any(i["arg"] == "tabble" for i in unknown)


def test_run_structured_error_details(simple_workbook: Path, tmp_path: Path):
    """Multiple issues in one workflow produce multiple structured details."""
    wf = tmp_path / "wf.yaml"
    wf.write_text(
        "steps:\n"
        "  - id: s1\n"
        "    run: cell.set\n"
        "    args: {}\n"  # missing ref and value
    )
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    issues = data["errors"][0]["details"]["issues"]
    assert len(issues) >= 2  # missing ref and value


def test_run_invalid_yaml_type(simple_workbook: Path, tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("just a string\n")
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert "WORKFLOW_INVALID" in data["errors"][0]["code"]


def test_run_missing_steps(simple_workbook: Path, tmp_path: Path):
    wf = tmp_path / "wf.yaml"
    wf.write_text("name: empty\n")
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    data = json.loads(result.stdout)
    assert data["ok"] is False


# ---------------------------------------------------------------------------
# formula lint filtering & summarization (Feature 4)
# ---------------------------------------------------------------------------
@pytest.fixture()
def lint_workbook(tmp_path: Path) -> Path:
    """Workbook with formulas that trigger lint findings."""
    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "=NOW()"              # volatile
    ws["A2"] = "=RAND()"             # volatile
    ws["A3"] = "=SUM(#REF!)"         # broken ref (error severity)
    ws["A4"] = "=OFFSET(A1,1,0)"     # volatile

    ws2 = wb.create_sheet("Other")
    ws2["A1"] = "=TODAY()"            # volatile

    path = tmp_path / "lint_test.xlsx"
    wb.save(str(path))
    wb.close()
    return path


def test_lint_defaults_unchanged(lint_workbook: Path):
    """Default lint returns all findings with summary."""
    result = runner.invoke(app, ["formula", "lint", "--file", str(lint_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["count"] >= 4
    assert "summary" in data["result"]
    assert len(data["result"]["findings"]) == data["result"]["count"]


def test_lint_filter_severity(lint_workbook: Path):
    """--severity error only returns broken_ref findings."""
    result = runner.invoke(app, ["formula", "lint", "--file", str(lint_workbook), "--severity", "error"])
    data = json.loads(result.stdout)
    assert data["ok"] is True
    # Only broken refs have error severity
    assert all(f["severity"] == "error" for f in data["result"]["findings"])
    assert data["result"]["count"] >= 1


def test_lint_filter_category(lint_workbook: Path):
    """--category volatile_function filters to only volatile findings."""
    result = runner.invoke(app, ["formula", "lint", "--file", str(lint_workbook), "--category", "volatile_function"])
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert all(f["category"] == "volatile_function" for f in data["result"]["findings"])
    assert data["result"]["count"] >= 3


def test_lint_summary_mode(lint_workbook: Path):
    """--summary returns empty findings list and populated summary."""
    result = runner.invoke(app, ["formula", "lint", "--file", str(lint_workbook), "--summary"])
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["findings"] == []
    assert data["result"]["count"] >= 4
    assert data["result"]["summary"]["total"] >= 4
    assert "volatile_function" in data["result"]["summary"]["by_category"]
    assert "Data" in data["result"]["summary"]["by_sheet"]


# ---------------------------------------------------------------------------
# Relative formula fill (Feature 5)
# ---------------------------------------------------------------------------
def test_adjust_refs_simple():
    """=A1+B1 shifted by (1,0) → =A2+B2"""
    from xl.adapters.openpyxl_engine import _adjust_formula_refs
    assert _adjust_formula_refs("=A1+B1", 1, 0) == "=A2+B2"


def test_adjust_refs_absolute_col():
    """=$A1 shifted by (1,0) → =$A2  (col locked, row shifts)"""
    from xl.adapters.openpyxl_engine import _adjust_formula_refs
    assert _adjust_formula_refs("=$A1", 1, 0) == "=$A2"


def test_adjust_refs_absolute_row():
    """=A$1 shifted by (1,0) → =A$1  (row locked)"""
    from xl.adapters.openpyxl_engine import _adjust_formula_refs
    assert _adjust_formula_refs("=A$1", 1, 0) == "=A$1"


def test_adjust_refs_fully_absolute():
    """=$A$1 is completely locked — unchanged."""
    from xl.adapters.openpyxl_engine import _adjust_formula_refs
    assert _adjust_formula_refs("=$A$1", 5, 5) == "=$A$1"


def test_adjust_refs_string_literal():
    """Refs inside string literals should NOT be adjusted."""
    from xl.adapters.openpyxl_engine import _adjust_formula_refs
    result = _adjust_formula_refs('=IF(A1>0,"A1",B1)', 1, 0)
    assert '"A1"' in result  # string literal preserved
    assert "A2" in result    # A1 outside string → A2
    assert "B2" in result    # B1 outside string → B2


def test_formula_set_relative_range_cli(simple_workbook: Path):
    """Fill =C2*D2 over E2:E5 in relative mode → each row adjusts."""
    import openpyxl

    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Revenue!E2:E5",
        "--formula", "=C2*D2",
        "--fill-mode", "relative",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True

    # Verify formulas were adjusted
    wb = openpyxl.load_workbook(str(simple_workbook))
    ws = wb["Revenue"]
    assert ws["E2"].value == "=C2*D2"
    assert ws["E3"].value == "=C3*D3"
    assert ws["E4"].value == "=C4*D4"
    assert ws["E5"].value == "=C5*D5"
    wb.close()


# ---------------------------------------------------------------------------
# Dry-run summary reporting (Feature 6)
# ---------------------------------------------------------------------------
def test_apply_dry_run_summary_multi_op(simple_workbook: Path, sample_plan_file: Path):
    """Dry-run apply includes dry_run_summary with operation counts."""
    result = runner.invoke(app, [
        "apply",
        "--file", str(simple_workbook),
        "--plan", str(sample_plan_file),
        "--dry-run",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True
    summary = data["result"]["dry_run_summary"]
    assert summary["total_operations"] >= 1
    assert summary["total_cells_affected"] >= 1
    assert "table.add_column" in summary["by_type"]
    assert len(summary["operations"]) >= 1


def test_dry_run_summary_by_sheet(simple_workbook: Path, sample_plan_file: Path):
    """dry_run_summary.by_sheet groups operations by sheet."""
    result = runner.invoke(app, [
        "apply",
        "--file", str(simple_workbook),
        "--plan", str(sample_plan_file),
        "--dry-run",
    ])
    data = json.loads(result.stdout)
    summary = data["result"]["dry_run_summary"]
    # The sample plan adds a column to Sales table which is on Revenue sheet
    # The target of table.add_column is "Sales[Margin]"
    assert len(summary["by_sheet"]) >= 1


def test_summarize_changes_unit():
    """Unit test for summarize_changes helper."""
    from xl.contracts.common import ChangeRecord
    from xl.engine.dispatcher import summarize_changes

    changes = [
        ChangeRecord(type="cell.set", target="Sheet1!A1", impact={"cells": 1}),
        ChangeRecord(type="cell.set", target="Sheet1!B1", impact={"cells": 1}),
        ChangeRecord(type="format.number", target="Sheet2!C1:C10", impact={"cells": 10}),
    ]
    result = summarize_changes(changes)
    assert result["total_operations"] == 3
    assert result["total_cells_affected"] == 12
    assert result["by_type"]["cell.set"] == 2
    assert result["by_type"]["format.number"] == 1
    assert result["by_sheet"]["Sheet1"] == 2
    assert result["by_sheet"]["Sheet2"] == 1
    assert len(result["operations"]) == 3
