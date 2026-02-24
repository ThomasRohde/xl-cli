"""Performance tests for large workbook operations.

These tests verify that core operations complete within acceptable time bounds
on workbooks with realistic data sizes. They create workbooks with 1000+ rows
and measure operation durations.

Tests are marked with ``pytest.mark.slow`` so they can be skipped in fast CI
runs with ``pytest -m 'not slow'``.
"""

from __future__ import annotations

import json
import time
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from typer.testing import CliRunner

from xl.cli import app
from xl.engine.context import WorkbookContext
from xl.io.fileops import fingerprint

runner = CliRunner()

slow = pytest.mark.slow


def _create_large_workbook(path: Path, rows: int = 1000, cols: int = 10) -> Path:
    """Create a workbook with a large table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = [f"Col{i}" for i in range(1, cols + 1)]
    ws.append(headers)

    for r in range(rows):
        ws.append([r * cols + c for c in range(cols)])

    end_col = get_column_letter(cols)
    ref = f"A1:{end_col}{rows + 1}"
    tab = Table(displayName="BigTable", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tab)

    wb.save(str(path))
    wb2 = openpyxl.load_workbook(str(path))
    wb2.save(str(path))
    wb2.close()
    return path


def _create_multi_sheet_workbook(path: Path, sheets: int = 20, rows_per: int = 100) -> Path:
    """Create a workbook with many sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for i in range(1, sheets + 1):
        if i == 1:
            ws_i = ws
        else:
            ws_i = wb.create_sheet(f"Sheet{i}")
        ws_i.append(["ID", "Value", "Label"])
        for r in range(rows_per):
            ws_i.append([r, r * 1.5, f"item_{r}"])

    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# WorkbookContext loading performance
# ---------------------------------------------------------------------------
@slow
class TestLoadPerformance:
    def test_load_large_workbook(self, tmp_path: Path) -> None:
        """Loading a 1000-row workbook should complete in under 5s."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=1000, cols=10)

        start = time.perf_counter()
        ctx = WorkbookContext(path)
        elapsed = time.perf_counter() - start
        ctx.close()

        assert elapsed < 5.0, f"Load took {elapsed:.2f}s, expected <5s"

    def test_load_multi_sheet_workbook(self, tmp_path: Path) -> None:
        """Loading a 20-sheet workbook should complete quickly."""
        path = _create_multi_sheet_workbook(tmp_path / "multi.xlsx", sheets=20, rows_per=100)

        start = time.perf_counter()
        ctx = WorkbookContext(path)
        elapsed = time.perf_counter() - start
        ctx.close()

        assert elapsed < 5.0, f"Load took {elapsed:.2f}s, expected <5s"


# ---------------------------------------------------------------------------
# Inspection performance
# ---------------------------------------------------------------------------
@slow
class TestInspectPerformance:
    def test_inspect_large_workbook(self, tmp_path: Path) -> None:
        """wb inspect on a large workbook should complete in under 5s."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=1000, cols=10)

        start = time.perf_counter()
        result = runner.invoke(app, ["wb", "inspect", "--file", str(path)])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 5.0, f"Inspect took {elapsed:.2f}s, expected <5s"

    def test_table_ls_large_workbook(self, tmp_path: Path) -> None:
        """table ls on a large table should complete quickly."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=1000, cols=10)

        start = time.perf_counter()
        result = runner.invoke(app, ["table", "ls", "--file", str(path)])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        tables = data["result"]
        assert len(tables) == 1
        assert tables[0]["row_count_estimate"] == 1000
        assert elapsed < 5.0, f"table ls took {elapsed:.2f}s, expected <5s"


# ---------------------------------------------------------------------------
# Mutation performance
# ---------------------------------------------------------------------------
@slow
class TestMutationPerformance:
    def test_add_column_large_table(self, tmp_path: Path) -> None:
        """Adding a column to a 1000-row table should complete in under 10s."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=1000, cols=10)

        start = time.perf_counter()
        result = runner.invoke(app, [
            "table", "add-column",
            "--file", str(path),
            "--table", "BigTable",
            "--name", "NewCol",
            "--default", "test_value",
        ])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 10.0, f"add-column took {elapsed:.2f}s, expected <10s"

    def test_append_many_rows(self, tmp_path: Path) -> None:
        """Appending 100 rows to a table should complete in under 10s."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=100, cols=5)

        # Build 100 rows matching the schema
        rows_data = [
            {f"Col{c}": i * 5 + c for c in range(1, 6)}
            for i in range(100)
        ]
        rows_json = json.dumps(rows_data)

        start = time.perf_counter()
        result = runner.invoke(app, [
            "table", "append-rows",
            "--file", str(path),
            "--table", "BigTable",
            "--data", rows_json,
        ])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 10.0, f"append-rows took {elapsed:.2f}s, expected <10s"

    def test_cell_set_many_cells(self, tmp_path: Path) -> None:
        """Setting cells one at a time in sequence should be functional."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=100, cols=5)

        start = time.perf_counter()
        for i in range(10):
            result = runner.invoke(app, [
                "cell", "set",
                "--file", str(path),
                "--ref", f"Data!A{i + 2}",
                "--value", str(i * 999),
                "--type", "number",
            ])
            assert result.exit_code == 0
        elapsed = time.perf_counter() - start

        assert elapsed < 30.0, f"10 cell sets took {elapsed:.2f}s, expected <30s"


# ---------------------------------------------------------------------------
# Fingerprint performance
# ---------------------------------------------------------------------------
@slow
class TestFingerprintPerformance:
    def test_fingerprint_large_file(self, tmp_path: Path) -> None:
        """Fingerprinting a large workbook should be fast."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=5000, cols=10)

        start = time.perf_counter()
        fp = fingerprint(path)
        elapsed = time.perf_counter() - start

        assert fp.startswith("sha256:")
        assert elapsed < 2.0, f"Fingerprint took {elapsed:.2f}s, expected <2s"


# ---------------------------------------------------------------------------
# Formula lint performance
# ---------------------------------------------------------------------------
@slow
class TestFormulaLintPerformance:
    def test_lint_large_sheet_with_formulas(self, tmp_path: Path) -> None:
        """Linting a sheet with many formulas should complete reasonably."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Formulas"
        ws.append(["A", "B", "C"])
        for i in range(2, 502):
            ws[f"A{i}"] = i
            ws[f"B{i}"] = f"=A{i}*2"
            ws[f"C{i}"] = f"=SUM(A$2:A{i})"

        path = tmp_path / "formulas_large.xlsx"
        wb.save(str(path))

        start = time.perf_counter()
        result = runner.invoke(app, ["formula", "lint", "--file", str(path)])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        assert elapsed < 10.0, f"Lint took {elapsed:.2f}s, expected <10s"


# ---------------------------------------------------------------------------
# Diff performance
# ---------------------------------------------------------------------------
@slow
class TestDiffPerformance:
    def test_diff_large_identical_workbooks(self, tmp_path: Path) -> None:
        """Diffing two identical large workbooks should complete reasonably."""
        import shutil

        path_a = _create_large_workbook(tmp_path / "a.xlsx", rows=500, cols=5)
        path_b = tmp_path / "b.xlsx"
        shutil.copy2(path_a, path_b)

        start = time.perf_counter()
        result = runner.invoke(app, [
            "diff", "compare",
            "--file-a", str(path_a),
            "--file-b", str(path_b),
        ])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 15.0, f"Diff took {elapsed:.2f}s, expected <15s"


# ---------------------------------------------------------------------------
# Query performance (DuckDB)
# ---------------------------------------------------------------------------
@slow
class TestQueryPerformance:
    def test_query_large_table(self, tmp_path: Path) -> None:
        """SQL query over a large table should complete reasonably."""
        # Build a workbook with mixed types (string + number) for DuckDB compat
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["Region", "Product", "Sales", "Cost", "Quantity"]
        ws.append(headers)
        regions = ["North", "South", "East", "West"]
        products = ["Widget", "Gadget", "Doohickey"]
        for i in range(1000):
            ws.append([
                regions[i % len(regions)],
                products[i % len(products)],
                100 + i,
                50 + i // 2,
                i % 20 + 1,
            ])
        ref = f"A1:E1001"
        tab = Table(displayName="SalesData", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
        ws.add_table(tab)
        path = tmp_path / "query_large.xlsx"
        wb.save(str(path))
        wb2 = openpyxl.load_workbook(str(path))
        wb2.save(str(path))
        wb2.close()

        start = time.perf_counter()
        result = runner.invoke(app, [
            "query",
            "--file", str(path),
            "--sql", "SELECT Region, SUM(Sales) as total FROM SalesData GROUP BY Region",
        ])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 15.0, f"Query took {elapsed:.2f}s, expected <15s"


# ---------------------------------------------------------------------------
# Validation performance
# ---------------------------------------------------------------------------
@slow
class TestValidationPerformance:
    def test_validate_large_workbook(self, tmp_path: Path) -> None:
        """Workbook validation on a large workbook should be fast."""
        path = _create_large_workbook(tmp_path / "large.xlsx", rows=1000, cols=10)

        start = time.perf_counter()
        result = runner.invoke(app, ["validate", "workbook", "--file", str(path)])
        elapsed = time.perf_counter() - start

        assert result.exit_code == 0
        data = json.loads(result.stdout)
        assert data["ok"] is True
        assert elapsed < 5.0, f"Validate took {elapsed:.2f}s, expected <5s"
