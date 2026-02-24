"""Tests for Phase 3 features: formula, cell get, range, format commands."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from typer.testing import CliRunner

from xl.cli import app

runner = CliRunner()


# ---------------------------------------------------------------------------
# formula set
# ---------------------------------------------------------------------------
def test_formula_set_cell(simple_workbook: Path):
    """Set a formula on a single empty cell."""
    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Summary!C1",
        "--formula", "=A1&B1",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["type"] == "formula.set"


def test_formula_set_overwrite_blocked(simple_workbook: Path):
    """Setting formula where values exist should be blocked without force flag."""
    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2:A5",
        "--formula", "=ROW()",
    ])
    data = json.loads(result.stdout)
    assert data["ok"] is False


def test_formula_set_force_overwrite_values(simple_workbook: Path):
    """Force overwrite existing values."""
    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2",
        "--formula", "=ROW()",
        "--force-overwrite-values",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True


def test_formula_set_dry_run(simple_workbook: Path):
    """Dry-run should not modify the workbook."""
    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Summary!C1",
        "--formula", "=1+1",
        "--dry-run",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True


# ---------------------------------------------------------------------------
# formula lint
# ---------------------------------------------------------------------------
def test_formula_lint_clean(simple_workbook: Path):
    """Clean workbook should have no major findings."""
    result = runner.invoke(app, ["formula", "lint", "--file", str(simple_workbook)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert isinstance(data["result"]["findings"], list)


def test_formula_lint_volatile(tmp_path: Path):
    """Detect volatile functions."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=NOW()"
    ws["A2"] = "=INDIRECT(B2)"
    path = tmp_path / "volatile.xlsx"
    wb.save(str(path))
    wb.close()

    result = runner.invoke(app, ["formula", "lint", "--file", str(path)])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"]["count"] >= 2
    categories = [f["category"] for f in data["result"]["findings"]]
    assert "volatile_function" in categories


def test_formula_lint_broken_ref(tmp_path: Path):
    """Detect broken #REF! references."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=#REF!+1"
    path = tmp_path / "broken.xlsx"
    wb.save(str(path))
    wb.close()

    result = runner.invoke(app, ["formula", "lint", "--file", str(path)])
    data = json.loads(result.stdout)
    assert data["result"]["count"] >= 1
    assert any(f["category"] == "broken_ref" for f in data["result"]["findings"])


# ---------------------------------------------------------------------------
# formula find
# ---------------------------------------------------------------------------
def test_formula_find(simple_workbook: Path):
    """Find formulas matching a pattern."""
    result = runner.invoke(app, [
        "formula", "find",
        "--file", str(simple_workbook),
        "--pattern", "SUM",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["count"] >= 1
    assert "SUM" in data["result"]["matches"][0]["formula"]


def test_formula_find_no_match(simple_workbook: Path):
    """No match should return empty list."""
    result = runner.invoke(app, [
        "formula", "find",
        "--file", str(simple_workbook),
        "--pattern", "VLOOKUP",
    ])
    data = json.loads(result.stdout)
    assert data["result"]["count"] == 0


# ---------------------------------------------------------------------------
# cell get
# ---------------------------------------------------------------------------
def test_cell_get_value(simple_workbook: Path):
    """Read a cell with a string value."""
    result = runner.invoke(app, [
        "cell", "get",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["value"] == "North"
    assert data["result"]["type"] == "text"


def test_cell_get_number(simple_workbook: Path):
    """Read a cell with a numeric value."""
    result = runner.invoke(app, [
        "cell", "get",
        "--file", str(simple_workbook),
        "--ref", "Revenue!C2",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"]["type"] == "number"
    assert data["result"]["value"] == 1000


def test_cell_get_formula(simple_workbook: Path):
    """Read a cell with a formula."""
    result = runner.invoke(app, [
        "cell", "get",
        "--file", str(simple_workbook),
        "--ref", "Summary!B1",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"]["type"] == "formula"
    assert data["result"]["formula"].startswith("=")


def test_cell_get_empty(simple_workbook: Path):
    """Read an empty cell."""
    result = runner.invoke(app, [
        "cell", "get",
        "--file", str(simple_workbook),
        "--ref", "Summary!Z99",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"]["type"] == "empty"
    assert data["result"]["value"] is None


# ---------------------------------------------------------------------------
# range stat
# ---------------------------------------------------------------------------
def test_range_stat(simple_workbook: Path):
    """Compute statistics for a numeric range."""
    result = runner.invoke(app, [
        "range", "stat",
        "--file", str(simple_workbook),
        "--ref", "Revenue!C2:C5",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    stats = data["result"]
    assert stats["row_count"] == 4
    assert stats["col_count"] == 1
    assert stats["numeric_count"] == 4
    assert stats["sum"] == 1000 + 1500 + 2000 + 800
    assert stats["min"] == 800
    assert stats["max"] == 2000


def test_range_stat_mixed(simple_workbook: Path):
    """Stats on a range with mixed types."""
    result = runner.invoke(app, [
        "range", "stat",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A1:D5",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    stats = data["result"]
    assert stats["row_count"] == 5
    assert stats["col_count"] == 4
    assert stats["non_empty_count"] > 0


# ---------------------------------------------------------------------------
# range clear
# ---------------------------------------------------------------------------
def test_range_clear(simple_workbook: Path):
    """Clear a range of cells."""
    result = runner.invoke(app, [
        "range", "clear",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2:A3",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["type"] == "range.clear"
    assert data["changes"][0]["after"]["cells_cleared"] == 2

    # Verify cells are cleared
    result = runner.invoke(app, [
        "cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!A2",
    ])
    data = json.loads(result.stdout)
    assert data["result"]["value"] is None


def test_range_clear_dry_run(simple_workbook: Path):
    """Dry-run should not modify the workbook."""
    result = runner.invoke(app, [
        "range", "clear",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A2:A3",
        "--dry-run",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["result"]["dry_run"] is True

    # Verify cells are NOT cleared
    result = runner.invoke(app, [
        "cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!A2",
    ])
    data = json.loads(result.stdout)
    assert data["result"]["value"] == "North"


# ---------------------------------------------------------------------------
# format number (CLI command)
# ---------------------------------------------------------------------------
def test_format_number_cmd(simple_workbook: Path):
    result = runner.invoke(app, [
        "format", "number",
        "--file", str(simple_workbook),
        "--ref", "Revenue!C2:C5",
        "--style", "currency",
        "--decimals", "2",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["type"] == "format.number"


# ---------------------------------------------------------------------------
# format width
# ---------------------------------------------------------------------------
def test_format_width(simple_workbook: Path):
    result = runner.invoke(app, [
        "format", "width",
        "--file", str(simple_workbook),
        "--sheet", "Revenue",
        "--columns", "A,B,C",
        "--width", "15.0",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["type"] == "format.width"


# ---------------------------------------------------------------------------
# format freeze
# ---------------------------------------------------------------------------
def test_format_freeze(simple_workbook: Path):
    result = runner.invoke(app, [
        "format", "freeze",
        "--file", str(simple_workbook),
        "--sheet", "Revenue",
        "--ref", "B2",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["changes"][0]["type"] == "format.freeze"
    assert data["changes"][0]["after"] == "B2"


def test_format_unfreeze(simple_workbook: Path):
    # First freeze, then unfreeze
    runner.invoke(app, [
        "format", "freeze", "--file", str(simple_workbook),
        "--sheet", "Revenue", "--ref", "B2",
    ])
    result = runner.invoke(app, [
        "format", "freeze", "--file", str(simple_workbook),
        "--sheet", "Revenue", "--unfreeze",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["changes"][0]["after"] is None


# ---------------------------------------------------------------------------
# validate refs
# ---------------------------------------------------------------------------
def test_validate_refs_valid(simple_workbook: Path):
    result = runner.invoke(app, [
        "validate", "refs",
        "--file", str(simple_workbook),
        "--ref", "Revenue!A1:D5",
    ])
    assert result.exit_code == 0
    data = json.loads(result.stdout)
    assert data["ok"] is True
    assert data["result"]["valid"] is True


def test_validate_refs_bad_sheet(simple_workbook: Path):
    result = runner.invoke(app, [
        "validate", "refs",
        "--file", str(simple_workbook),
        "--ref", "NonExistent!A1",
    ])
    data = json.loads(result.stdout)
    assert data["result"]["valid"] is False
