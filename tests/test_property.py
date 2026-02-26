"""Property-based tests using Hypothesis for patch/apply semantics.

These tests verify invariants that must hold for *any* valid input, not just
specific examples. They exercise:
- ResponseEnvelope round-trip serialization
- PatchPlan model construction and serialization
- Table append-rows schema enforcement
- Cell set/get round-trip integrity
- Fingerprint determinism
- Apply idempotency (dry-run never mutates)
"""

from __future__ import annotations

import json
import re
import shutil
from pathlib import Path

import openpyxl
import pytest
from hypothesis import HealthCheck, given, settings, assume
from hypothesis import strategies as st
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from xl.contracts.common import (
    ChangeRecord,
    ErrorDetail,
    Metrics,
    RecalcInfo,
    ResponseEnvelope,
    Target,
    WarningDetail,
)
from xl.contracts.plans import (
    Operation,
    PatchPlan,
    PlanOptions,
    PlanTarget,
    Postcondition,
    Precondition,
)
from xl.engine.context import WorkbookContext
from xl.io.fileops import fingerprint

# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------

# Safe text that won't break Excel/openpyxl
safe_text = st.text(
    alphabet=st.characters(whitelist_categories=("L", "N", "P", "Z"), min_codepoint=32, max_codepoint=126),
    min_size=1,
    max_size=30,
).filter(lambda s: s.strip() and not s.startswith("="))

safe_header = st.from_regex(r"[A-Za-z][A-Za-z0-9_]{0,19}", fullmatch=True)

cell_values = st.one_of(
    st.integers(min_value=-1_000_000, max_value=1_000_000),
    st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False),
    safe_text,
)

numeric_values = st.one_of(
    st.integers(min_value=-1_000_000, max_value=1_000_000),
    st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False),
)


def _make_workbook_with_table(tmp_path: Path, headers: list[str], rows: list[list]) -> Path:
    """Helper: create a workbook with a single table from given headers/rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    nrows = len(rows) + 1
    ncols = len(headers)
    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(ncols)
    ref = f"A1:{end_col}{nrows}"

    tab = Table(displayName="TestTable", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    ws.add_table(tab)

    path = tmp_path / "prop_test.xlsx"
    wb.save(str(path))
    # Reload to populate tableColumns
    wb2 = openpyxl.load_workbook(str(path))
    wb2.save(str(path))
    wb2.close()
    return path


# ---------------------------------------------------------------------------
# ResponseEnvelope round-trip
# ---------------------------------------------------------------------------
class TestEnvelopeRoundtrip:
    @given(
        ok=st.booleans(),
        command=st.sampled_from(["wb.inspect", "table.ls", "cell.set", "apply", "validate.plan"]),
        duration=st.integers(min_value=0, max_value=100_000),
    )
    @settings(max_examples=50, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_envelope_serialize_deserialize(self, ok: bool, command: str, duration: int) -> None:
        env = ResponseEnvelope(
            ok=ok,
            command=command,
            metrics=Metrics(duration_ms=duration),
        )
        data = env.model_dump(mode="json")
        json_str = json.dumps(data)
        restored = json.loads(json_str)

        assert restored["ok"] == ok
        assert restored["command"] == command
        assert restored["metrics"]["duration_ms"] == duration
        assert isinstance(restored["changes"], list)
        assert isinstance(restored["warnings"], list)
        assert isinstance(restored["errors"], list)

    @given(
        code=st.sampled_from(["ERR_WORKBOOK_NOT_FOUND", "ERR_TABLE_NOT_FOUND", "ERR_SCHEMA_MISMATCH"]),
        message=safe_text,
    )
    @settings(max_examples=30)
    def test_error_envelope_always_has_error_fields(self, code: str, message: str) -> None:
        env = ResponseEnvelope(
            ok=False,
            command="test",
            errors=[ErrorDetail(code=code, message=message)],
        )
        data = env.model_dump(mode="json")
        assert data["ok"] is False
        assert len(data["errors"]) == 1
        assert data["errors"][0]["code"] == code
        assert data["errors"][0]["message"] == message


# ---------------------------------------------------------------------------
# PatchPlan model
# ---------------------------------------------------------------------------
class TestPatchPlanProperties:
    @given(
        plan_id=st.text(min_size=1, max_size=40, alphabet="abcdefghijklmnopqrstuvwxyz0123456789_"),
        table_name=safe_header,
        column_name=safe_header,
    )
    @settings(max_examples=50)
    def test_plan_roundtrip(self, plan_id: str, table_name: str, column_name: str) -> None:
        plan = PatchPlan(
            plan_id=plan_id,
            target=PlanTarget(file="test.xlsx"),
            operations=[
                Operation(
                    op_id="op1",
                    type="table.add_column",
                    table=table_name,
                    name=column_name,
                ),
            ],
            preconditions=[Precondition(type="table_exists", table=table_name)],
            postconditions=[Postcondition(type="column_exists", table=table_name, column=column_name)],
        )
        data = plan.model_dump(mode="json")
        restored = PatchPlan(**data)
        assert restored.plan_id == plan_id
        assert restored.operations[0].table == table_name
        assert restored.operations[0].name == column_name
        assert restored.preconditions[0].table == table_name
        assert restored.postconditions[0].column == column_name


# ---------------------------------------------------------------------------
# Fingerprint determinism
# ---------------------------------------------------------------------------
class TestFingerprintProperties:
    @given(data=st.binary(min_size=1, max_size=10_000))
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_fingerprint_deterministic(self, data: bytes, tmp_path: Path) -> None:
        p = tmp_path / "test_file.bin"
        p.write_bytes(data)
        fp1 = fingerprint(p)
        fp2 = fingerprint(p)
        assert fp1 == fp2
        assert fp1.startswith("sha256:")

    @given(
        data_a=st.binary(min_size=1, max_size=1000),
        data_b=st.binary(min_size=1, max_size=1000),
    )
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_fingerprint_differs_for_different_content(
        self, data_a: bytes, data_b: bytes, tmp_path: Path,
    ) -> None:
        assume(data_a != data_b)
        pa = tmp_path / "a.bin"
        pb = tmp_path / "b.bin"
        pa.write_bytes(data_a)
        pb.write_bytes(data_b)
        assert fingerprint(pa) != fingerprint(pb)


# ---------------------------------------------------------------------------
# Cell set/get round-trip
# ---------------------------------------------------------------------------
class TestCellSetGetRoundtrip:
    @given(value=st.integers(min_value=-1_000_000, max_value=1_000_000))
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_integer_roundtrip(self, value: int, tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import cell_get, cell_set

        path = _make_workbook_with_table(tmp_path, ["A"], [[0]])
        ctx = WorkbookContext(path)
        cell_set(ctx, "Data", "A2", value, force_overwrite_formulas=True)
        ctx.save(path)
        ctx.close()

        ctx2 = WorkbookContext(path)
        result = cell_get(ctx2, "Data", "A2")
        ctx2.close()
        assert result["value"] == value
        assert result["type"] == "number"

    @given(value=st.floats(min_value=-1e6, max_value=1e6, allow_nan=False, allow_infinity=False))
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_float_roundtrip(self, value: float, tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import cell_get, cell_set

        path = _make_workbook_with_table(tmp_path, ["A"], [[0]])
        ctx = WorkbookContext(path)
        cell_set(ctx, "Data", "A2", value, force_overwrite_formulas=True)
        ctx.save(path)
        ctx.close()

        ctx2 = WorkbookContext(path)
        result = cell_get(ctx2, "Data", "A2")
        ctx2.close()
        assert abs(result["value"] - value) < 1e-10
        assert result["type"] == "number"

    @given(value=safe_text)
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_text_roundtrip(self, value: str, tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import cell_get, cell_set

        path = _make_workbook_with_table(tmp_path, ["A"], [["placeholder"]])
        ctx = WorkbookContext(path)
        cell_set(ctx, "Data", "A2", value)
        ctx.save(path)
        ctx.close()

        ctx2 = WorkbookContext(path)
        result = cell_get(ctx2, "Data", "A2")
        ctx2.close()
        assert result["value"] == value
        assert result["type"] == "text"


# ---------------------------------------------------------------------------
# Table append-rows schema enforcement
# ---------------------------------------------------------------------------
class TestAppendRowsSchemaProperty:
    @given(
        extra_col=safe_header.filter(lambda s: s not in ("Col1", "Col2")),
    )
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_strict_mode_rejects_extra_columns(self, extra_col: str, tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import table_append_rows

        path = _make_workbook_with_table(tmp_path, ["Col1", "Col2"], [["a", "b"]])
        ctx = WorkbookContext(path)
        with pytest.raises(ValueError, match="Extra columns"):
            table_append_rows(ctx, "TestTable", [{"Col1": "x", "Col2": "y", extra_col: "z"}], schema_mode="strict")
        ctx.close()

    @given(
        values=st.lists(
            st.fixed_dictionaries({"Col1": safe_text, "Col2": safe_text}),
            min_size=1,
            max_size=5,
        ),
    )
    @settings(max_examples=20, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_strict_mode_accepts_matching_schema(self, values: list[dict], tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import table_append_rows

        path = _make_workbook_with_table(tmp_path, ["Col1", "Col2"], [["a", "b"]])
        ctx = WorkbookContext(path)
        change = table_append_rows(ctx, "TestTable", values, schema_mode="strict")
        ctx.close()
        assert change.type == "table.append_rows"
        assert change.after["rows_added"] == len(values)


# ---------------------------------------------------------------------------
# Add column never loses existing data
# ---------------------------------------------------------------------------
class TestAddColumnPreservesData:
    @given(
        col_name=safe_header,
        default_val=st.one_of(st.none(), st.integers(min_value=0, max_value=1000)),
    )
    @settings(max_examples=15, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_add_column_preserves_existing(self, col_name: str, default_val, tmp_path: Path) -> None:
        from xl.adapters.openpyxl_engine import table_add_column

        assume(col_name not in ("Region", "Sales"))
        path = _make_workbook_with_table(tmp_path, ["Region", "Sales"], [["North", 100], ["South", 200]])
        ctx = WorkbookContext(path)

        # Read original data
        ws = ctx.wb["Data"]
        orig_a2 = ws["A2"].value
        orig_b2 = ws["B2"].value

        table_add_column(ctx, "TestTable", col_name, default_value=default_val)

        # Original data should be unchanged
        assert ws["A2"].value == orig_a2
        assert ws["B2"].value == orig_b2
        ctx.close()


# ---------------------------------------------------------------------------
# Dry-run never mutates the workbook file
# ---------------------------------------------------------------------------
class TestDryRunNeverMutates:
    @given(
        col_name=safe_header.filter(lambda s: s not in ("Region", "Product", "Sales", "Cost")),
    )
    @settings(max_examples=10, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_dry_run_preserves_fingerprint(self, col_name: str, tmp_path: Path) -> None:
        from typer.testing import CliRunner
        from xl.cli import app

        runner = CliRunner()
        path = _make_workbook_with_table(tmp_path, ["Region", "Sales"], [["North", 100]])
        fp_before = fingerprint(path)

        runner.invoke(app, [
            "table", "add-column",
            "--file", str(path),
            "--table", "TestTable",
            "--name", col_name,
            "--dry-run",
        ])

        fp_after = fingerprint(path)
        assert fp_before == fp_after, "Dry-run must not modify the workbook file"


# ---------------------------------------------------------------------------
# Plan validate: valid plan always validates
# ---------------------------------------------------------------------------
class TestPlanValidateConsistency:
    @given(
        col_name=safe_header.filter(lambda s: s not in ("Region", "Sales")),
    )
    @settings(max_examples=10, deadline=500, suppress_health_check=[HealthCheck.function_scoped_fixture])
    def test_valid_plan_validates_successfully(self, col_name: str, tmp_path: Path) -> None:
        from xl.validation.validators import validate_plan

        path = _make_workbook_with_table(tmp_path, ["Region", "Sales"], [["North", 100]])
        ctx = WorkbookContext(path)

        plan = PatchPlan(
            plan_id="test_plan",
            target=PlanTarget(file=str(path), fingerprint=ctx.fp),
            preconditions=[Precondition(type="table_exists", table="TestTable")],
            operations=[
                Operation(op_id="op1", type="table.add_column", table="TestTable", name=col_name),
            ],
        )
        result = validate_plan(ctx, plan)
        ctx.close()
        assert result.valid is True, f"Valid plan should validate: {result.checks}"


# ---------------------------------------------------------------------------
# ChangeRecord model properties
# ---------------------------------------------------------------------------
class TestChangeRecordProperties:
    @given(
        op_type=st.sampled_from(["table.add_column", "table.append_rows", "cell.set", "format.number"]),
        target=safe_text,
        cells=st.integers(min_value=0, max_value=10_000),
    )
    @settings(max_examples=30)
    def test_change_record_serialization(self, op_type: str, target: str, cells: int) -> None:
        cr = ChangeRecord(
            type=op_type,
            target=target,
            impact={"cells": cells},
        )
        data = cr.model_dump(mode="json")
        assert data["type"] == op_type
        assert data["target"] == target
        assert data["impact"]["cells"] == cells
