"""Regression tests for fixes from REPORT.md."""

from __future__ import annotations

import json
from pathlib import Path

import yaml
from typer.testing import CliRunner

from xl.cli import app
from xl.io.fileops import fingerprint

runner = CliRunner()


def _json(stdout: str) -> dict:
    return json.loads(stdout)


def test_formula_find_invalid_regex_returns_envelope(simple_workbook: Path):
    result = runner.invoke(
        app,
        ["formula", "find", "--file", str(simple_workbook), "--pattern", "["],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_PATTERN_INVALID"


def test_table_add_column_duplicate_rejected(simple_workbook: Path):
    first = runner.invoke(
        app,
        [
            "table",
            "add-column",
            "--file",
            str(simple_workbook),
            "--table",
            "Sales",
            "--name",
            "Margin",
            "--formula",
            "=[@Sales]-[@Cost]",
        ],
    )
    assert first.exit_code == 0

    second = runner.invoke(
        app,
        [
            "table",
            "add-column",
            "--file",
            str(simple_workbook),
            "--table",
            "Sales",
            "--name",
            "Margin",
            "--formula",
            "=[@Sales]-[@Cost]",
        ],
    )
    data = _json(second.stdout)
    assert second.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_COLUMN_EXISTS"


def test_formula_set_table_column_skips_header(simple_workbook: Path):
    add = runner.invoke(
        app,
        [
            "table",
            "add-column",
            "--file",
            str(simple_workbook),
            "--table",
            "Sales",
            "--name",
            "NewCalc",
            "--default",
            "0",
        ],
    )
    assert add.exit_code == 0

    set_formula = runner.invoke(
        app,
        [
            "formula",
            "set",
            "--file",
            str(simple_workbook),
            "--ref",
            "Sales[NewCalc]",
            "--formula",
            "=[@Sales]-[@Cost]",
            "--force-overwrite-values",
        ],
    )
    assert set_formula.exit_code == 0

    header = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!E1"])
    body = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!E2"])
    header_data = _json(header.stdout)
    body_data = _json(body.stdout)
    assert header_data["result"]["value"] == "NewCalc"
    assert body_data["result"]["formula"] == "=[@Sales]-[@Cost]"


def test_plan_validate_accepts_envelope_input(simple_workbook: Path, tmp_path: Path):
    """Envelope files are auto-unwrapped: validate extracts .result as the plan body."""
    envelope_plan = tmp_path / "envelope_plan.json"
    gen = runner.invoke(
        app,
        [
            "plan",
            "add-column",
            "--file",
            str(simple_workbook),
            "--table",
            "Sales",
            "--name",
            "EnvelopeCol",
            "--formula",
            "=[@Sales]-[@Cost]",
        ],
    )
    assert gen.exit_code == 0
    envelope_plan.write_text(gen.stdout)

    validate = runner.invoke(
        app,
        ["validate", "plan", "--file", str(simple_workbook), "--plan", str(envelope_plan)],
    )
    data = _json(validate.stdout)
    assert validate.exit_code == 0
    assert data["ok"] is True


def test_plan_append_writes_to_disk(simple_workbook: Path, tmp_path: Path):
    plan_file = tmp_path / "append_plan.json"
    add = runner.invoke(
        app,
        [
            "plan",
            "add-column",
            "--file",
            str(simple_workbook),
            "--table",
            "Sales",
            "--name",
            "Margin1",
            "--append",
            str(plan_file),
        ],
    )
    assert add.exit_code == 0
    assert plan_file.exists()
    p1 = json.loads(plan_file.read_text())
    assert len(p1["operations"]) == 1

    fmt = runner.invoke(
        app,
        [
            "plan",
            "format",
            "--file",
            str(simple_workbook),
            "--ref",
            "Sales[Sales]",
            "--style",
            "currency",
            "--append",
            str(plan_file),
        ],
    )
    assert fmt.exit_code == 0
    p2 = json.loads(plan_file.read_text())
    assert len(p2["operations"]) == 2


def test_verify_assert_supports_row_count_gte_and_value_alias(simple_workbook: Path):
    assertions = json.dumps(
        [
            {"type": "row_count.gte", "table": "Sales", "min_rows": 4},
            {"type": "cell.value_equals", "ref": "Revenue!A2", "value": "North"},
        ]
    )
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True


def test_verify_assert_failure_exit_code_is_validation(simple_workbook: Path):
    assertions = json.dumps([{"type": "cell.value_equals", "ref": "Revenue!A2", "expected": "WRONG"}])
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_ASSERTION_FAILED"


def test_validate_refs_invalid_returns_validation_error(simple_workbook: Path):
    result = runner.invoke(
        app,
        ["validate", "refs", "--file", str(simple_workbook), "--ref", "Missing!A1"],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_RANGE_INVALID"


def test_format_freeze_mutually_exclusive_flags(simple_workbook: Path):
    result = runner.invoke(
        app,
        [
            "format",
            "freeze",
            "--file",
            str(simple_workbook),
            "--sheet",
            "Revenue",
            "--ref",
            "B2",
            "--unfreeze",
        ],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_INVALID_ARGUMENT"


def test_format_width_rejects_non_letter_columns(simple_workbook: Path):
    result = runner.invoke(
        app,
        [
            "format",
            "width",
            "--file",
            str(simple_workbook),
            "--sheet",
            "Revenue",
            "--columns",
            "1,2",
            "--width",
            "10",
            "--dry-run",
        ],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_RANGE_INVALID"


def test_range_clear_formats_only_preserves_cell_value(simple_workbook: Path):
    fmt = runner.invoke(
        app,
        [
            "format",
            "number",
            "--file",
            str(simple_workbook),
            "--ref",
            "Revenue!C2:C2",
            "--style",
            "currency",
        ],
    )
    assert fmt.exit_code == 0

    clear = runner.invoke(
        app,
        [
            "range",
            "clear",
            "--file",
            str(simple_workbook),
            "--ref",
            "Revenue!C2:C2",
            "--formats",
        ],
    )
    clear_data = _json(clear.stdout)
    assert clear.exit_code == 0
    assert clear_data["changes"][0]["after"]["contents"] is False
    assert clear_data["changes"][0]["after"]["formats"] is True

    cell = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!C2"])
    cell_data = _json(cell.stdout)
    assert cell_data["result"]["value"] == 1000


def test_diff_compare_missing_sheet_is_validation_error(simple_workbook: Path):
    result = runner.invoke(
        app,
        [
            "diff",
            "compare",
            "--file-a",
            str(simple_workbook),
            "--file-b",
            str(simple_workbook),
            "--sheet",
            "Missing",
        ],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_RANGE_INVALID"


def test_query_invalid_sql_returns_envelope(simple_workbook: Path):
    result = runner.invoke(
        app,
        ["query", "--file", str(simple_workbook), "--sql", "SELECT Missing FROM Sales"],
    )
    data = _json(result.stdout)
    assert result.exit_code == 90
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_QUERY_FAILED"


def test_run_invalid_workflow_structure_does_not_mutate_file(simple_workbook: Path, tmp_path: Path):
    wf = tmp_path / "invalid_workflow.yaml"
    wf.write_text("foo: bar\n")

    before = fingerprint(simple_workbook)
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    after = fingerprint(simple_workbook)
    data = _json(result.stdout)

    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_WORKFLOW_INVALID"
    assert before == after


def test_run_empty_steps_rejected_without_mutation(simple_workbook: Path, tmp_path: Path):
    wf = tmp_path / "empty_steps.yaml"
    wf.write_text(yaml.safe_dump({"schema_version": "1.0", "target": {"file": str(simple_workbook)}, "steps": []}))

    before = fingerprint(simple_workbook)
    result = runner.invoke(app, ["run", "--workflow", str(wf), "--file", str(simple_workbook)])
    after = fingerprint(simple_workbook)
    data = _json(result.stdout)

    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_WORKFLOW_INVALID"
    assert before == after


def test_run_non_mutating_workflow_does_not_save(simple_workbook: Path, tmp_path: Path):
    wf_path = tmp_path / "read_only.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "read_only",
        "target": {"file": str(simple_workbook)},
        "defaults": {"dry_run": False},
        "steps": [{"id": "list_tables", "run": "table.ls", "args": {}}],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    before = fingerprint(simple_workbook)
    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    after = fingerprint(simple_workbook)
    data = _json(result.stdout)

    assert result.exit_code == 0
    assert data["ok"] is True
    assert before == after


# ---------------------------------------------------------------------------
# Finding 1: UTF-8 BOM tolerance
# ---------------------------------------------------------------------------

def test_load_plan_with_utf8_bom(simple_workbook: Path, sample_plan: dict, tmp_path: Path):
    """Plan JSON files with a UTF-8 BOM should parse without error."""
    plan_path = tmp_path / "bom_plan.json"
    bom = b"\xef\xbb\xbf"
    plan_path.write_bytes(bom + json.dumps(sample_plan).encode("utf-8"))

    result = runner.invoke(
        app,
        ["validate", "plan", "--file", str(simple_workbook), "--plan", str(plan_path)],
    )
    data = _json(result.stdout)
    assert data["ok"] is True


def test_workflow_yaml_with_utf8_bom(simple_workbook: Path, tmp_path: Path):
    """Workflow YAML files with a UTF-8 BOM should load correctly."""
    workflow = {
        "schema_version": "1.0",
        "name": "bom_test",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "s1", "run": "wb.inspect", "args": {}}],
    }
    wf_path = tmp_path / "bom_workflow.yaml"
    bom = b"\xef\xbb\xbf"
    wf_path.write_bytes(bom + yaml.safe_dump(workflow).encode("utf-8"))

    result = runner.invoke(
        app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)]
    )
    data = _json(result.stdout)
    assert data["ok"] is True


def test_assertions_file_with_utf8_bom(simple_workbook: Path, tmp_path: Path):
    """Assertions JSON files with a UTF-8 BOM should parse without error."""
    assertions = [{"type": "table.column_exists", "table": "Sales", "column": "Region"}]
    af = tmp_path / "bom_assertions.json"
    bom = b"\xef\xbb\xbf"
    af.write_bytes(bom + json.dumps(assertions).encode("utf-8"))

    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions-file", str(af)],
    )
    data = _json(result.stdout)
    assert data["ok"] is True


def test_data_file_with_utf8_bom(simple_workbook: Path, tmp_path: Path):
    """--data-file JSON with a UTF-8 BOM should parse without error."""
    rows = [{"Region": "Central", "Product": "Widget", "Sales": 500, "Cost": 300}]
    df = tmp_path / "bom_rows.json"
    bom = b"\xef\xbb\xbf"
    df.write_bytes(bom + json.dumps(rows).encode("utf-8"))

    result = runner.invoke(
        app,
        [
            "table", "append-rows",
            "--file", str(simple_workbook),
            "--table", "Sales",
            "--data-file", str(df),
        ],
    )
    data = _json(result.stdout)
    assert data["ok"] is True


# ---------------------------------------------------------------------------
# Finding 5: WARN_UNCACHED_FORMULA on cell get --data-only
# ---------------------------------------------------------------------------

def test_cell_get_data_only_warns_on_uncached_formula(simple_workbook: Path):
    """When --data-only returns null for a formula cell, a warning should appear."""
    result = runner.invoke(
        app,
        ["cell", "get", "--file", str(simple_workbook), "--ref", "Summary!B1", "--data-only"],
    )
    data = _json(result.stdout)
    assert data["ok"] is True
    # If the value is None (no cached value), we expect the warning
    if data["result"]["value"] is None:
        assert any(w["code"] == "WARN_UNCACHED_FORMULA" for w in data.get("warnings", []))


# ---------------------------------------------------------------------------
# Finding 6: diff compare --include-formulas
# ---------------------------------------------------------------------------

def test_diff_compare_include_formulas(simple_workbook: Path, tmp_path: Path):
    """diff compare --include-formulas should report formula text changes."""
    import shutil
    import openpyxl as oxl

    copy_path = tmp_path / "modified.xlsx"
    shutil.copy2(simple_workbook, copy_path)
    wb = oxl.load_workbook(str(copy_path))
    wb["Summary"]["B1"] = "=SUM(Revenue!C2:C4)"  # changed range
    wb.save(str(copy_path))
    wb.close()

    result = runner.invoke(
        app,
        [
            "diff", "compare",
            "--file-a", str(simple_workbook),
            "--file-b", str(copy_path),
            "--include-formulas",
        ],
    )
    data = _json(result.stdout)
    assert data["ok"] is True
    assert "formula_changes" in data["result"]
    assert len(data["result"]["formula_changes"]) > 0
    fc = data["result"]["formula_changes"][0]
    assert fc["change_type"] == "formula_modified"


# ---------------------------------------------------------------------------
# Finding 3: Workflow dispatcher expansion
# ---------------------------------------------------------------------------

def test_workflow_query_step(simple_workbook: Path, tmp_path: Path):
    """Workflow query step should execute SQL against tables."""
    workflow = {
        "schema_version": "1.0",
        "name": "query_test",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "q1", "run": "query", "args": {"sql": "SELECT COUNT(*) as cnt FROM Sales"}}],
    }
    wf_path = tmp_path / "workflow_query.yaml"
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True
    assert data["result"]["steps"][0]["result"]["row_count"] == 1


def test_workflow_cell_get_step(simple_workbook: Path, tmp_path: Path):
    """Workflow cell.get step should read a cell value."""
    workflow = {
        "schema_version": "1.0",
        "name": "cell_get_test",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "r1", "run": "cell.get", "args": {"ref": "Revenue!A2"}}],
    }
    wf_path = tmp_path / "workflow_cell_get.yaml"
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["result"]["value"] == "North"


def test_workflow_formula_find_step(simple_workbook: Path, tmp_path: Path):
    """Workflow formula.find step should find formulas matching a pattern."""
    workflow = {
        "schema_version": "1.0",
        "name": "ff_test",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "f1", "run": "formula.find", "args": {"pattern": "SUM"}}],
    }
    wf_path = tmp_path / "workflow_ff.yaml"
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True


def test_workflow_invalid_step_rejected_at_parse_time(simple_workbook: Path, tmp_path: Path):
    """Workflow with unknown step command should fail at parse time."""
    workflow = {
        "schema_version": "1.0",
        "name": "bad_step",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "s1", "run": "nonexistent.command", "args": {}}],
    }
    wf_path = tmp_path / "bad_step.yaml"
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_WORKFLOW_INVALID"


# ---------------------------------------------------------------------------
# REPORT.md v1.3.0 blind-test fixes
# ---------------------------------------------------------------------------


def test_error_output_not_duplicated(simple_workbook: Path):
    """§4.1 — Error JSON should appear exactly once, not twice."""
    result = runner.invoke(app, ["wb", "inspect", "--file", "nonexistent_file.xlsx"])
    # Count occurrences of '"ok"' in combined output
    combined = (result.stdout or "") + (result.stderr or "")
    assert combined.count('"ok"') == 1, f"Expected 1 JSON envelope, got {combined.count('\"ok\"')}"


def test_cell_set_date_stores_as_date(tmp_path: Path):
    """§4.4 — cell set --type date should store as Excel serial date, not text."""
    wb_path = tmp_path / "date_test.xlsx"
    runner.invoke(app, ["wb", "create", "--file", str(wb_path)])

    result = runner.invoke(app, [
        "cell", "set", "--file", str(wb_path),
        "--ref", "Sheet!A1", "--value", "2026-02-25", "--type", "date",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True

    get_result = runner.invoke(app, ["cell", "get", "--file", str(wb_path), "--ref", "Sheet!A1"])
    get_data = _json(get_result.stdout)
    assert get_data["result"]["type"] == "date"
    assert "2026-02-25" in get_data["result"]["value"]
    assert "YYYY" in get_data["result"]["number_format"] or "yyyy" in get_data["result"]["number_format"]


def test_cell_set_date_with_time(tmp_path: Path):
    """§4.4 — cell set --type date should accept YYYY-MM-DDTHH:MM:SS."""
    wb_path = tmp_path / "datetime_test.xlsx"
    runner.invoke(app, ["wb", "create", "--file", str(wb_path)])

    result = runner.invoke(app, [
        "cell", "set", "--file", str(wb_path),
        "--ref", "Sheet!A1", "--value", "2026-02-25T14:30:00", "--type", "date",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True

    get_result = runner.invoke(app, ["cell", "get", "--file", str(wb_path), "--ref", "Sheet!A1"])
    get_data = _json(get_result.stdout)
    assert get_data["result"]["type"] == "date"
    assert "2026-02-25" in get_data["result"]["value"]
    assert "14:30:00" in get_data["result"]["value"]


def test_table_append_rows_omit_formula_column(formula_table_workbook: Path):
    """§4.3 — Formula columns should not be required in append-rows."""
    import openpyxl
    rows = json.dumps([{"Name": "Diana", "Amount": 400}])
    result = runner.invoke(app, [
        "table", "append-rows",
        "--file", str(formula_table_workbook),
        "--table", "Payments",
        "--data", rows,
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["changes"][0]["after"]["rows_added"] == 1

    # Verify formula was copied into the new row
    wb = openpyxl.load_workbook(str(formula_table_workbook))
    ws = wb["Payments"]
    tax_cell = ws.cell(row=5, column=3).value  # row 5 = new row (header + 3 data + 1 new)
    assert isinstance(tax_cell, str) and tax_cell.startswith("=")
    wb.close()


def test_table_append_rows_explicit_null_overrides_formula(formula_table_workbook: Path):
    """§4.3 — Explicit null for formula column should use that value."""
    rows = json.dumps([{"Name": "Eve", "Amount": 500, "Tax": None}])
    result = runner.invoke(app, [
        "table", "append-rows",
        "--file", str(formula_table_workbook),
        "--table", "Payments",
        "--data", rows,
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True


def test_table_ls_shows_formula_columns(formula_table_workbook: Path):
    """§4.3 — table ls should indicate which columns are formula columns."""
    result = runner.invoke(app, [
        "table", "ls",
        "--file", str(formula_table_workbook),
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    table = data["result"][0]
    tax_col = next(c for c in table["columns"] if c["name"] == "Tax")
    assert tax_col["is_formula"] is True
    assert tax_col["formula"] is not None

    name_col = next(c for c in table["columns"] if c["name"] == "Name")
    assert name_col["is_formula"] is False


# ---------------------------------------------------------------------------
# REPORT.md blind-test regression fixes
# ---------------------------------------------------------------------------


def test_verify_cell_value_type_accepts_expected_alias(simple_workbook: Path):
    """§4.7 — cell.value_type should accept 'expected' as alias for 'expected_type'."""
    assertions = json.dumps([
        {"type": "cell.value_type", "ref": "Revenue!C2", "expected": "number"},
    ])
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["result"]["assertions"][0]["passed"] is True


def test_workflow_cell_set_with_type_arg(simple_workbook: Path, tmp_path: Path):
    """§4.3 — Workflow cell.set should accept 'type' arg for coercion."""
    wf_path = tmp_path / "wf_cell_type.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "cell_type_test",
        "target": {"file": str(simple_workbook)},
        "steps": [
            {"id": "s1", "run": "cell.set", "args": {"ref": "Summary!C1", "value": "42", "type": "number"}},
        ],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True

    get_result = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Summary!C1"])
    get_data = _json(get_result.stdout)
    assert get_data["result"]["value"] == 42
    assert get_data["result"]["type"] == "number"


def test_workflow_format_width_comma_separated_columns(simple_workbook: Path, tmp_path: Path):
    """§4.2 — Workflow format.width should accept columns as comma-separated string."""
    wf_path = tmp_path / "wf_width.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "width_test",
        "target": {"file": str(simple_workbook)},
        "steps": [
            {"id": "s1", "run": "format.width", "args": {"sheet": "Revenue", "columns": "A,B", "width": 20}},
        ],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True


def test_formula_set_range_defaults_to_relative(simple_workbook: Path):
    """§4.1 — Range formula set without --fill-mode should default to relative."""
    import openpyxl

    result = runner.invoke(app, [
        "formula", "set",
        "--file", str(simple_workbook),
        "--ref", "Revenue!E2:E5",
        "--formula", "=C2*D2",
        # No --fill-mode flag: should default to relative
    ])
    assert result.exit_code == 0
    data = _json(result.stdout)
    assert data["ok"] is True

    wb = openpyxl.load_workbook(str(simple_workbook))
    ws = wb["Revenue"]
    assert ws["E2"].value == "=C2*D2"
    assert ws["E3"].value == "=C3*D3"
    assert ws["E4"].value == "=C4*D4"
    assert ws["E5"].value == "=C5*D5"
    wb.close()


# ---------------------------------------------------------------------------
# Bug #1 regression: Workflow format.number / formula.set with Table[Column]
# ---------------------------------------------------------------------------


def test_workflow_format_number_table_column_ref(simple_workbook: Path, tmp_path: Path):
    """Bug #1: workflow format.number should resolve Sales[Sales] to cell range."""
    wf_path = tmp_path / "wf_fmt_tbl.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "fmt_table_ref",
        "target": {"file": str(simple_workbook)},
        "steps": [
            {
                "id": "fmt",
                "run": "format.number",
                "args": {"ref": "Sales[Sales]", "style": "currency", "decimals": 2},
            },
        ],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True


def test_workflow_formula_set_table_column_ref(simple_workbook: Path, tmp_path: Path):
    """Bug #1: workflow formula.set should resolve Sales[NewCol] to cell range (no header)."""
    # First add the column via a workflow step, then set formula on it
    wf_path = tmp_path / "wf_formula_tbl.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "formula_table_ref",
        "target": {"file": str(simple_workbook)},
        "steps": [
            {
                "id": "add",
                "run": "table.add_column",
                "args": {"table": "Sales", "name": "Margin"},
            },
            {
                "id": "set",
                "run": "formula.set",
                "args": {"ref": "Sales[Margin]", "formula": "=[@Sales]-[@Cost]"},
            },
        ],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True
    assert data["result"]["steps"][1]["ok"] is True


def test_workflow_format_number_sheet_range_still_works(simple_workbook: Path, tmp_path: Path):
    """Bug #1 guard: existing Sheet!Range syntax must still work."""
    wf_path = tmp_path / "wf_fmt_sheet.yaml"
    workflow = {
        "schema_version": "1.0",
        "name": "fmt_sheet_ref",
        "target": {"file": str(simple_workbook)},
        "steps": [
            {
                "id": "fmt",
                "run": "format.number",
                "args": {"ref": "Revenue!C2:C5", "style": "number", "decimals": 2},
            },
        ],
    }
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["steps"][0]["ok"] is True


# ---------------------------------------------------------------------------
# Bug #2 regression: table.row_count.gte should accept "expected" parameter
# ---------------------------------------------------------------------------


def test_verify_row_count_gte_with_expected_param(simple_workbook: Path):
    """Bug #2: table.row_count.gte should accept 'expected' as alias for min_rows."""
    assertions = json.dumps([
        {"type": "table.row_count.gte", "table": "Sales", "expected": 3},
    ])
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["result"]["assertions"][0]["passed"] is True


def test_verify_row_count_gte_with_expected_fails_correctly(simple_workbook: Path):
    """Bug #2: table.row_count.gte with expected > actual should fail."""
    assertions = json.dumps([
        {"type": "table.row_count.gte", "table": "Sales", "expected": 999},
    ])
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["result"]["assertions"][0]["passed"] is False


def test_verify_row_count_gte_min_rows_still_works(simple_workbook: Path):
    """Bug #2 guard: existing min_rows param must still work."""
    assertions = json.dumps([
        {"type": "table.row_count.gte", "table": "Sales", "min_rows": 4},
    ])
    result = runner.invoke(
        app,
        ["verify", "assert", "--file", str(simple_workbook), "--assertions", assertions],
    )
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True


# ---------------------------------------------------------------------------
# REPORT.md v1.3.2 — Section 4.2 Issue Fixes
# ---------------------------------------------------------------------------


# Issue #7: format number trailing dot with 0 decimals
def test_format_number_zero_decimals_no_trailing_dot(simple_workbook: Path):
    """Issue #7: --decimals 0 should NOT produce a trailing dot in format string."""
    for style, expected_no_dot in [("currency", "$#,##0"), ("number", "#,##0"), ("percent", "0%")]:
        result = runner.invoke(app, [
            "format", "number", "--file", str(simple_workbook),
            "--ref", "Revenue!C2:C2", "--style", style, "--decimals", "0",
        ])
        data = _json(result.stdout)
        assert result.exit_code == 0
        fmt = data["changes"][0]["after"]["format"]
        assert fmt == expected_no_dot, f"Style '{style}' with 0 decimals got '{fmt}', expected '{expected_no_dot}'"


def test_format_number_positive_decimals_still_works(simple_workbook: Path):
    """Issue #7 regression guard: --decimals 2 should still produce correct format."""
    result = runner.invoke(app, [
        "format", "number", "--file", str(simple_workbook),
        "--ref", "Revenue!C2:C2", "--style", "currency", "--decimals", "2",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["changes"][0]["after"]["format"] == "$#,##0.00"


# Issue #9: double JSON output
def test_validate_workflow_error_single_output(simple_workbook: Path, tmp_path: Path):
    """Issue #9: validate workflow error should output JSON exactly once."""
    wf = tmp_path / "bad.yaml"
    wf.write_text("not: a: valid: yaml: [workflow")
    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf)])
    combined = (result.stdout or "") + (result.stderr or "")
    assert combined.count('"ok"') == 1


def test_verify_assert_parse_error_single_output(simple_workbook: Path):
    """Issue #9: verify assert with bad JSON should output exactly once."""
    result = runner.invoke(app, [
        "verify", "assert", "--file", str(simple_workbook),
        "--assertions", "not valid json",
    ])
    combined = (result.stdout or "") + (result.stderr or "")
    assert combined.count('"ok"') == 1


# Issue #2: inline JSON assertions error message
def test_verify_assert_parse_error_suggests_file(simple_workbook: Path):
    """Issue #2: JSON parse error should suggest --assertions-file."""
    result = runner.invoke(app, [
        "verify", "assert", "--file", str(simple_workbook),
        "--assertions", "{bad json}",
    ])
    data = _json(result.stdout)
    assert data["ok"] is False
    assert "assertions-file" in data["errors"][0]["message"]


# Issue #4: diff includes formulas by default
def test_diff_includes_formulas_by_default(simple_workbook: Path, tmp_path: Path):
    """Issue #4: diff compare should include formula_changes by default."""
    import shutil
    import openpyxl as oxl

    copy_path = tmp_path / "modified.xlsx"
    shutil.copy2(simple_workbook, copy_path)
    wb = oxl.load_workbook(str(copy_path))
    wb["Summary"]["B1"] = "=SUM(Revenue!C2:C4)"
    wb.save(str(copy_path))
    wb.close()

    result = runner.invoke(app, [
        "diff", "compare",
        "--file-a", str(simple_workbook),
        "--file-b", str(copy_path),
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert "formula_changes" in data["result"]


def test_diff_no_formulas_flag(simple_workbook: Path, tmp_path: Path):
    """Issue #4: --no-formulas should exclude formula_changes."""
    import shutil
    copy_path = tmp_path / "modified2.xlsx"
    shutil.copy2(simple_workbook, copy_path)

    result = runner.invoke(app, [
        "diff", "compare",
        "--file-a", str(simple_workbook),
        "--file-b", str(copy_path),
        "--no-formulas",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert "formula_changes" not in data["result"]


# Issue #1: query --data-only
def test_query_data_only_flag(simple_workbook: Path):
    """Issue #1: --data-only should not return formula strings starting with '='."""
    result = runner.invoke(app, [
        "query", "--file", str(simple_workbook),
        "--table", "Sales", "--data-only",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    # With data-only, formula cells should return cached values (or None), not formula strings
    for row in data["result"]["rows"]:
        for val in row.values():
            if isinstance(val, str):
                assert not val.startswith("="), f"Got formula string '{val}' with --data-only"


# Issue #3: workflow args hint
def test_validate_workflow_detects_missing_args(simple_workbook: Path, tmp_path: Path):
    """Issue #3: validate workflow should catch missing required args."""
    workflow = {
        "schema_version": "1.0",
        "name": "missing_args",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "s1", "run": "cell.get", "args": {}}],
    }
    wf_path = tmp_path / "missing_args.yaml"
    wf_path.write_text(yaml.safe_dump(workflow))

    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf_path)])
    data = _json(result.stdout)
    assert data["result"]["valid"] is False
    failed = [c for c in data["result"]["checks"] if not c["passed"]]
    assert any("ref" in c["message"] for c in failed)


def test_validate_workflow_hints_misplaced_args(simple_workbook: Path, tmp_path: Path):
    """Issue #3: when args are at step level, error should hint to move inside args:."""
    wf_path = tmp_path / "misplaced.yaml"
    content = yaml.safe_dump({
        "schema_version": "1.0",
        "name": "misplaced",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "s1", "run": "cell.get", "ref": "Revenue!A1", "args": {}}],
    })
    wf_path.write_text(content)

    result = runner.invoke(app, ["validate", "workflow", "--workflow", str(wf_path)])
    data = _json(result.stdout)
    assert data["result"]["valid"] is False
    failed = [c for c in data["result"]["checks"] if not c["passed"]]
    assert any("step level" in c["message"] for c in failed)


def test_run_workflow_hints_misplaced_args(simple_workbook: Path, tmp_path: Path):
    """Issue #3: xl run should also hint about misplaced args."""
    wf_path = tmp_path / "misplaced_run.yaml"
    content = yaml.safe_dump({
        "schema_version": "1.0",
        "name": "misplaced_run",
        "target": {"file": str(simple_workbook)},
        "steps": [{"id": "s1", "run": "cell.get", "ref": "Revenue!A1", "args": {}}],
    })
    wf_path.write_text(content)

    result = runner.invoke(app, ["run", "--workflow", str(wf_path), "--file", str(simple_workbook)])
    data = _json(result.stdout)
    assert data["ok"] is False
    err_msg = data["errors"][0]["message"]
    # The detailed issues should contain the hint
    details = data["errors"][0].get("details", {})
    issues = details.get("issues", [])
    assert any("step level" in str(i.get("message", "")) for i in issues)


# Issue #8: plan create-table header detection
def test_plan_create_table_detects_headers(simple_workbook: Path):
    """Issue #8: plan create-table should detect column headers from workbook."""
    result = runner.invoke(app, [
        "plan", "create-table",
        "--file", str(simple_workbook),
        "--table", "NewTable",
        "--sheet", "Revenue",
        "--ref", "A1:D5",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    plan_result = data["result"]
    # Should have detected headers
    assert plan_result.get("detected_headers") is not None
    assert len(plan_result["detected_headers"]) == 4
    # Operation should have columns populated
    op = plan_result["operations"][0]
    assert op["columns"] is not None


def test_plan_create_table_explicit_columns_override(simple_workbook: Path):
    """Issue #8: explicit --columns should override header detection."""
    result = runner.invoke(app, [
        "plan", "create-table",
        "--file", str(simple_workbook),
        "--table", "NewTable",
        "--sheet", "Revenue",
        "--ref", "A1:D5",
        "--columns", "X,Y,Z,W",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    op = data["result"]["operations"][0]
    assert op["columns"] == ["X", "Y", "Z", "W"]
    # No detected_headers when explicit columns provided
    assert data["result"].get("detected_headers") is None


# Issue #5: sheet delete and rename
def test_sheet_delete_success(simple_workbook: Path):
    """Issue #5: sheet delete should remove a sheet."""
    result = runner.invoke(app, [
        "sheet", "delete", "--file", str(simple_workbook), "--name", "Summary",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["command"] == "sheet.delete"
    assert data["result"]["sheet"] == "Summary"

    # Verify sheet is gone
    ls = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    names = [s["name"] for s in _json(ls.stdout)["result"]]
    assert "Summary" not in names


def test_sheet_delete_last_sheet_blocked(tmp_path: Path):
    """Issue #5: cannot delete the last remaining sheet."""
    wb_path = tmp_path / "single.xlsx"
    runner.invoke(app, ["wb", "create", "--file", str(wb_path)])

    result = runner.invoke(app, [
        "sheet", "delete", "--file", str(wb_path), "--name", "Sheet",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["ok"] is False
    assert data["errors"][0]["code"] == "ERR_LAST_SHEET"


def test_sheet_delete_not_found(simple_workbook: Path):
    """Issue #5: deleting nonexistent sheet returns ERR_SHEET_NOT_FOUND."""
    result = runner.invoke(app, [
        "sheet", "delete", "--file", str(simple_workbook), "--name", "Nonexistent",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["errors"][0]["code"] == "ERR_SHEET_NOT_FOUND"


def test_sheet_delete_dry_run(simple_workbook: Path):
    """Issue #5: sheet delete --dry-run should not persist."""
    result = runner.invoke(app, [
        "sheet", "delete", "--file", str(simple_workbook),
        "--name", "Summary", "--dry-run",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True

    ls = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    names = [s["name"] for s in _json(ls.stdout)["result"]]
    assert "Summary" in names


def test_sheet_rename_success(simple_workbook: Path):
    """Issue #5: sheet rename should change the sheet name."""
    result = runner.invoke(app, [
        "sheet", "rename", "--file", str(simple_workbook),
        "--name", "Summary", "--new-name", "Overview",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["result"]["old_name"] == "Summary"
    assert data["result"]["new_name"] == "Overview"

    ls = runner.invoke(app, ["sheet", "ls", "--file", str(simple_workbook)])
    names = [s["name"] for s in _json(ls.stdout)["result"]]
    assert "Overview" in names
    assert "Summary" not in names


def test_sheet_rename_target_exists(simple_workbook: Path):
    """Issue #5: renaming to existing name returns ERR_SHEET_EXISTS."""
    result = runner.invoke(app, [
        "sheet", "rename", "--file", str(simple_workbook),
        "--name", "Summary", "--new-name", "Revenue",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["errors"][0]["code"] == "ERR_SHEET_EXISTS"


def test_sheet_rename_not_found(simple_workbook: Path):
    """Issue #5: renaming nonexistent sheet returns ERR_SHEET_NOT_FOUND."""
    result = runner.invoke(app, [
        "sheet", "rename", "--file", str(simple_workbook),
        "--name", "Nonexistent", "--new-name", "NewName",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["errors"][0]["code"] == "ERR_SHEET_NOT_FOUND"


# Issue #6: table delete and column delete
def test_table_delete_success(simple_workbook: Path):
    """Issue #6: table delete should remove table definition, preserve data."""
    import openpyxl

    # Read a cell value before deletion
    get_before = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!A2"])
    val_before = _json(get_before.stdout)["result"]["value"]

    result = runner.invoke(app, [
        "table", "delete", "--file", str(simple_workbook), "--table", "Sales",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["command"] == "table.delete"

    # Verify table is gone
    ls = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    tables = [t["name"] for t in _json(ls.stdout)["result"]]
    assert "Sales" not in tables

    # Verify cell data is preserved
    get_after = runner.invoke(app, ["cell", "get", "--file", str(simple_workbook), "--ref", "Revenue!A2"])
    val_after = _json(get_after.stdout)["result"]["value"]
    assert val_after == val_before


def test_table_delete_not_found(simple_workbook: Path):
    """Issue #6: deleting nonexistent table returns error."""
    result = runner.invoke(app, [
        "table", "delete", "--file", str(simple_workbook), "--table", "Nonexistent",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["errors"][0]["code"] == "ERR_TABLE_NOT_FOUND"


def test_table_delete_dry_run(simple_workbook: Path):
    """Issue #6: table delete --dry-run should not persist."""
    result = runner.invoke(app, [
        "table", "delete", "--file", str(simple_workbook),
        "--table", "Sales", "--dry-run",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True

    ls = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    tables = [t["name"] for t in _json(ls.stdout)["result"]]
    assert "Sales" in tables


def test_table_delete_column_success(simple_workbook: Path):
    """Issue #6: table delete-column should remove the column."""
    # First add a column so we have something to delete
    runner.invoke(app, [
        "table", "add-column", "--file", str(simple_workbook),
        "--table", "Sales", "--name", "ToDelete", "--default", "0",
    ])

    result = runner.invoke(app, [
        "table", "delete-column", "--file", str(simple_workbook),
        "--table", "Sales", "--name", "ToDelete",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 0
    assert data["ok"] is True
    assert data["command"] == "table.delete_column"

    # Verify column is gone
    ls = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    tables = _json(ls.stdout)["result"]
    sales = next(t for t in tables if t["name"] == "Sales")
    col_names = [c["name"] for c in sales["columns"]]
    assert "ToDelete" not in col_names


def test_table_delete_column_not_found(simple_workbook: Path):
    """Issue #6: deleting nonexistent column returns error."""
    result = runner.invoke(app, [
        "table", "delete-column", "--file", str(simple_workbook),
        "--table", "Sales", "--name", "Nonexistent",
    ])
    data = _json(result.stdout)
    assert result.exit_code == 10
    assert data["errors"][0]["code"] == "ERR_COLUMN_NOT_FOUND"


def test_table_delete_column_dry_run(simple_workbook: Path):
    """Issue #6: delete-column --dry-run should not persist."""
    result = runner.invoke(app, [
        "table", "delete-column", "--file", str(simple_workbook),
        "--table", "Sales", "--name", "Region", "--dry-run",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    assert data["result"]["dry_run"] is True

    ls = runner.invoke(app, ["table", "ls", "--file", str(simple_workbook)])
    tables = _json(ls.stdout)["result"]
    sales = next(t for t in tables if t["name"] == "Sales")
    col_names = [c["name"] for c in sales["columns"]]
    assert "Region" in col_names


# Plan commands for new operations
def test_plan_delete_sheet(simple_workbook: Path):
    """Issue #5: plan delete-sheet should generate valid plan."""
    result = runner.invoke(app, [
        "plan", "delete-sheet", "--file", str(simple_workbook), "--sheet", "Summary",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    op = data["result"]["operations"][0]
    assert op["type"] == "sheet.delete"
    assert op["sheet"] == "Summary"


def test_plan_rename_sheet(simple_workbook: Path):
    """Issue #5: plan rename-sheet should generate valid plan."""
    result = runner.invoke(app, [
        "plan", "rename-sheet", "--file", str(simple_workbook),
        "--sheet", "Summary", "--new-name", "Overview",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    op = data["result"]["operations"][0]
    assert op["type"] == "sheet.rename"
    assert op["new_name"] == "Overview"


def test_plan_delete_table(simple_workbook: Path):
    """Issue #6: plan delete-table should generate valid plan."""
    result = runner.invoke(app, [
        "plan", "delete-table", "--file", str(simple_workbook), "--table", "Sales",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    op = data["result"]["operations"][0]
    assert op["type"] == "table.delete"
    assert op["table"] == "Sales"


def test_plan_delete_column(simple_workbook: Path):
    """Issue #6: plan delete-column should generate valid plan."""
    result = runner.invoke(app, [
        "plan", "delete-column", "--file", str(simple_workbook),
        "--table", "Sales", "--column", "Region",
    ])
    data = _json(result.stdout)
    assert data["ok"] is True
    op = data["result"]["operations"][0]
    assert op["type"] == "table.delete_column"
    assert op["column"] == "Region"
